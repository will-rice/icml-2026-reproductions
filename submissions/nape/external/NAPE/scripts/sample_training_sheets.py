"""
Worksheet Sampling Script for Training Data Collection.

Samples N qualifying worksheets from a directory of Excel workbooks.
For each workbook, at most one qualifying sheet is selected based on:
  - Non-empty content (at least one cell with a value)
  - Formatting attribute diversity (N distinct non-default attribute types)
  - Formatting coverage (X% of used range has non-default formatting)
  - No cross-sheet formula dependencies
  - No pivot tables or pivot charts
  - No Excel Tables (Ctrl+T / ListObjects)
  - No chart sheets

Outputs:
  - Copies of selected .xlsx files to dest_dir/xlsx/
  - Sheet images (PNG) to dest_dir/sheet_images/
  - A tracking JSON file to prevent re-sampling on subsequent runs

Usage:
    python sample_training_sheets.py --source_dir /path/to/workbooks --dest_dir /path/to/output
    python sample_training_sheets.py --source_dir /path/to/workbooks --dest_dir /path/to/output --num_sheets 500
    python sample_training_sheets.py --source_dir /path/to/workbooks --dest_dir /path/to/output --previous_tracking prev_run/sampling_tracker.json
"""

from __future__ import annotations

import argparse
import hashlib
import json
import logging
import os
import random
import shutil
import subprocess
import sys
import threading
import time
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

# ---------------------------------------------------------------------------
# Ensure the package is importable when running as a script
# ---------------------------------------------------------------------------
_SCRIPT_DIR = Path(__file__).resolve().parent
_PROJECT_ROOT = _SCRIPT_DIR.parent
_SRC_DIR = _PROJECT_ROOT / "src"
if str(_SRC_DIR) not in sys.path:
    sys.path.insert(0, str(_SRC_DIR))

from tqdm import tqdm

# ---------------------------------------------------------------------------
# Monkey-patch: openpyxl ColumnDimension to ignore unknown kwargs (e.g. widthPt)
# ---------------------------------------------------------------------------
import inspect
from openpyxl.worksheet.dimensions import ColumnDimension

_orig_col_dim_init = ColumnDimension.__init__
_col_dim_params = set(inspect.signature(_orig_col_dim_init).parameters.keys())

def _patched_col_dim_init(self, *args, **kwargs):
    filtered = {k: v for k, v in kwargs.items() if k in _col_dim_params}
    _orig_col_dim_init(self, *args, **filtered)

ColumnDimension.__init__ = _patched_col_dim_init
# ---------------------------------------------------------------------------

# Reuse screenshot utilities (guarded import -- xlwings is optional)
from next_action_pred_eval.utils.image_utils.screenshots import (
    create_configured_app,
    ensure_app_alive,
    save_images_add_headings,
    _safe_close_app,
)

logger = logging.getLogger(__name__)

# =============================================================================
# CONFIGURATION DEFAULTS
# =============================================================================

DEFAULT_NUM_SHEETS = 1000
DEFAULT_MIN_FILLED_CELLS = 1
DEFAULT_MIN_FORMAT_ATTRIBUTES = 8
DEFAULT_MIN_FORMAT_COVERAGE = 0.20
DEFAULT_SEED = None
DEFAULT_MAX_INDEX = 100

# Sheet names to always reject (e.g., Spire.XLS trial watermark sheets)
BLACKLISTED_SHEET_NAMES = {"Evaluation Warning", "Evaluation_Warning"}


# =============================================================================
# DATA CLASSES
# =============================================================================


@dataclass
class SampledSheet:
    """Record of a single sampled worksheet."""

    workbook_path: str  # Relative path from source_dir
    workbook_abs_path: str  # Absolute path (internal use)
    sheet_name: str
    dest_name: str  # Unique output name
    sampled_at: str  # ISO timestamp
    format_attributes_found: int
    format_coverage: float
    total_cells: int
    formatted_cells: int


@dataclass
class TrackingData:
    """Persistent tracking file structure."""

    version: int = 1
    source_dir: str = ""
    samples: List[Dict[str, Any]] = field(default_factory=list)

    def sampled_workbooks(self) -> Set[str]:
        """Return set of workbook relative paths already sampled."""
        return {s["workbook"] for s in self.samples}


# =============================================================================
# TRACKING FILE I/O
# =============================================================================


def load_tracking_file(path: Optional[Path]) -> TrackingData:
    """Load tracking data from a JSON file. Returns empty TrackingData if missing."""
    if path is None or not path.exists():
        return TrackingData()
    with open(path, "r", encoding="utf-8") as f:
        raw = json.load(f)
    return TrackingData(
        version=raw.get("version", 1),
        source_dir=raw.get("source_dir", ""),
        samples=raw.get("samples", []),
    )


def save_tracking_file(path: Path, data: TrackingData) -> None:
    """Save tracking data to a JSON file (atomic write via temp + rename)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "version": data.version,
        "source_dir": data.source_dir,
        "samples": data.samples,
    }
    tmp_path = path.with_suffix(".tmp")
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)
    tmp_path.replace(path)


# =============================================================================
# WORKBOOK DISCOVERY
# =============================================================================


def discover_workbooks(source_dir: Path) -> List[Path]:
    """Recursively find all .xlsx files, skipping ~$ temp files."""
    workbooks = []
    for root, _dirs, files in os.walk(source_dir):
        for f in files:
            if f.endswith(".xlsx") and not f.startswith("~$"):
                workbooks.append(Path(root) / f)
    workbooks.sort()
    return workbooks


# =============================================================================
# NAMING HELPERS
# =============================================================================


def sanitize_name(name: str) -> str:
    """Sanitize a name for use in file/folder names."""
    return name.replace(" ", "_").replace("/", "_").replace("\\", "_")


def make_dest_name(workbook_path: Path, sheet_name: str, existing_names: Set[str]) -> str:
    """Generate a unique destination name: {stem}_{sheet_name}, with hash suffix on collision."""
    base = f"{sanitize_name(workbook_path.stem)}_{sanitize_name(sheet_name)}"
    if base not in existing_names:
        return base
    path_hash = hashlib.md5(str(workbook_path).encode()).hexdigest()[:8]
    return f"{base}_{path_hash}"


# =============================================================================
# ZIP-LEVEL PRE-CHECK (PIVOTS, TABLES, CHART SHEETS)
# =============================================================================

# XML namespaces used in xlsx internals
_NS_SPREADSHEET = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_RELATIONSHIPS = "http://schemas.openxmlformats.org/package/2006/relationships"
_NS_OFFICEREL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

# Relationship types that disqualify a sheet
_REL_PIVOT_TABLE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotTable"
)
_REL_TABLE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/table"
)


def _get_sheet_name_to_xml_map(zf: zipfile.ZipFile) -> Dict[str, str]:
    """
    Parse xl/workbook.xml to build a mapping of sheet name -> sheetN.xml filename.

    Returns e.g. {"Sheet1": "sheet1.xml", "Data": "sheet2.xml"}.
    """
    mapping: Dict[str, str] = {}
    try:
        wb_xml = zf.read("xl/workbook.xml")
        root = ET.fromstring(wb_xml)
    except (KeyError, ET.ParseError):
        return mapping

    # Get sheet elements with their rId references
    sheets_el = root.find(f"{{{_NS_SPREADSHEET}}}sheets")
    if sheets_el is None:
        return mapping

    rid_to_name: Dict[str, str] = {}
    for sheet_el in sheets_el.findall(f"{{{_NS_SPREADSHEET}}}sheet"):
        name = sheet_el.get("name", "")
        rid = sheet_el.get(f"{{{_NS_OFFICEREL}}}id", "")
        if name and rid:
            rid_to_name[rid] = name

    # Parse workbook relationships to map rId -> sheetN.xml target
    try:
        rels_xml = zf.read("xl/_rels/workbook.xml.rels")
        rels_root = ET.fromstring(rels_xml)
    except (KeyError, ET.ParseError):
        return mapping

    for rel in rels_root.findall(f"{{{_NS_RELATIONSHIPS}}}Relationship"):
        rid = rel.get("Id", "")
        target = rel.get("Target", "")
        if rid in rid_to_name and "worksheets/" in target:
            # target is like "worksheets/sheet1.xml"
            xml_filename = target.split("/")[-1]
            mapping[rid_to_name[rid]] = xml_filename

    return mapping


def _get_chart_sheet_names(zf: zipfile.ZipFile) -> Set[str]:
    """Identify chart sheet names by checking xl/workbook.xml relationships."""
    chart_sheets: Set[str] = set()
    try:
        wb_xml = zf.read("xl/workbook.xml")
        root = ET.fromstring(wb_xml)
    except (KeyError, ET.ParseError):
        return chart_sheets

    sheets_el = root.find(f"{{{_NS_SPREADSHEET}}}sheets")
    if sheets_el is None:
        return chart_sheets

    rid_to_name: Dict[str, str] = {}
    for sheet_el in sheets_el.findall(f"{{{_NS_SPREADSHEET}}}sheet"):
        name = sheet_el.get("name", "")
        rid = sheet_el.get(f"{{{_NS_OFFICEREL}}}id", "")
        if name and rid:
            rid_to_name[rid] = name

    try:
        rels_xml = zf.read("xl/_rels/workbook.xml.rels")
        rels_root = ET.fromstring(rels_xml)
    except (KeyError, ET.ParseError):
        return chart_sheets

    for rel in rels_root.findall(f"{{{_NS_RELATIONSHIPS}}}Relationship"):
        rid = rel.get("Id", "")
        target = rel.get("Target", "")
        if rid in rid_to_name and "chartsheets/" in target:
            chart_sheets.add(rid_to_name[rid])

    return chart_sheets


def _get_sheets_with_disqualifying_features(
    zf: zipfile.ZipFile, sheet_name_to_xml: Dict[str, str]
) -> Set[str]:
    """
    Check worksheet _rels files for pivot table and table relationships.

    Returns a set of sheet names that have pivots or tables.
    """
    disqualified: Set[str] = set()

    # Invert the mapping: xml_filename -> sheet_name
    xml_to_name: Dict[str, str] = {v: k for k, v in sheet_name_to_xml.items()}

    for xml_filename, sheet_name in xml_to_name.items():
        rels_path = f"xl/worksheets/_rels/{xml_filename}.rels"
        try:
            rels_xml = zf.read(rels_path)
        except KeyError:
            continue  # No relationships file = no pivots/tables

        try:
            rels_root = ET.fromstring(rels_xml)
        except ET.ParseError:
            continue

        for rel in rels_root.findall(f"{{{_NS_RELATIONSHIPS}}}Relationship"):
            rel_type = rel.get("Type", "")
            if rel_type in (_REL_PIVOT_TABLE, _REL_TABLE):
                disqualified.add(sheet_name)
                break  # No need to check more rels for this sheet

    return disqualified


def _get_password_protected_sheets(
    zf: zipfile.ZipFile, sheet_name_to_xml: Dict[str, str]
) -> Set[str]:
    """
    Check worksheet XML for <sheetProtection> with a password attribute.

    Returns a set of sheet names that have password protection.
    """
    protected: Set[str] = set()

    for sheet_name, xml_filename in sheet_name_to_xml.items():
        xml_path = f"xl/worksheets/{xml_filename}"
        try:
            ws_xml = zf.read(xml_path)
        except KeyError:
            continue

        try:
            root = ET.fromstring(ws_xml)
        except ET.ParseError:
            continue

        # Look for <sheetProtection password="..." sheet="1" .../>
        prot_el = root.find(f"{{{_NS_SPREADSHEET}}}sheetProtection")
        if prot_el is not None:
            has_password = prot_el.get("password") or prot_el.get("algorithmName")
            is_protected = prot_el.get("sheet", "0") == "1"
            if has_password and is_protected:
                protected.add(sheet_name)

    return protected


def zip_precheck(workbook_path: Path) -> Set[str]:
    """
    Fast zip-level pre-check. Returns a set of sheet names to EXCLUDE
    (sheets with pivots, tables, chart sheets, or password protection).

    Returns empty set if the workbook is clean or can't be read as a zip.
    """
    excluded: Set[str] = set()

    try:
        with zipfile.ZipFile(str(workbook_path), "r") as zf:
            namelist = zf.namelist()

            # Check for chart sheets
            has_chartsheets = any(n.startswith("xl/chartsheets/") for n in namelist)
            if has_chartsheets:
                excluded.update(_get_chart_sheet_names(zf))

            # Build sheet map (needed for pivots, tables, and password check)
            sheet_map = _get_sheet_name_to_xml_map(zf)

            # Check for pivots or tables
            has_pivots = any(n.startswith("xl/pivotTables/") for n in namelist)
            has_tables = any(n.startswith("xl/tables/") for n in namelist)

            if (has_pivots or has_tables) and sheet_map:
                excluded.update(
                    _get_sheets_with_disqualifying_features(zf, sheet_map)
                )

            # Check for password-protected sheets
            if sheet_map:
                excluded.update(_get_password_protected_sheets(zf, sheet_map))

    except (zipfile.BadZipFile, OSError, Exception) as e:
        logger.debug(f"Zip pre-check failed for {workbook_path}: {e}")

    return excluded


# =============================================================================
# CELL-LEVEL CONDITION CHECKS
# =============================================================================


def classify_format_attributes(cell) -> Set[Tuple[str, str]]:
    """
    Identify unique (property_name, property_value) tuples of non-default
    formatting on a cell. This measures true formatting *diversity* — a sheet
    with 500 bold cells scores 1, while a sheet with bold, italic, red, blue
    scores 4.

    Inlines the logic from sheet_to_state._extract_font/fill/alignment/borders
    but collects (name, value) pairs rather than building value dicts.
    """
    attrs: Set[Tuple[str, str]] = set()

    # Font
    font = cell.font
    if font is not None:
        if font.name and font.name != "Calibri":
            attrs.add(("font_name", str(font.name)))
        if font.size and font.size != 11:
            attrs.add(("font_size", str(font.size)))
        if font.bold:
            attrs.add(("font_bold", "True"))
        if font.italic:
            attrs.add(("font_italic", "True"))
        if font.underline and font.underline != "none":
            attrs.add(("font_underline", str(font.underline)))
        if font.color and font.color.rgb and font.color.rgb != "00000000":
            attrs.add(("font_color", str(font.color.rgb)))

    # Fill
    fill = cell.fill
    if fill is not None and fill.fill_type == "solid":
        if fill.fgColor and fill.fgColor.rgb and fill.fgColor.rgb != "00000000":
            attrs.add(("fill_color", str(fill.fgColor.rgb)))

    # Alignment
    alignment = cell.alignment
    if alignment is not None:
        if alignment.horizontal and alignment.horizontal != "general":
            attrs.add(("horizontal_alignment", str(alignment.horizontal)))
        if alignment.vertical and alignment.vertical != "bottom":
            attrs.add(("vertical_alignment", str(alignment.vertical)))
        if alignment.wrap_text:
            attrs.add(("wrap_text", "True"))
        if alignment.text_rotation and alignment.text_rotation != 0:
            attrs.add(("text_rotation", str(alignment.text_rotation)))

    # Borders
    border = cell.border
    if border is not None:
        for side_name in ("left", "right", "top", "bottom"):
            side = getattr(border, side_name, None)
            if side and side.style and side.style != "none":
                attrs.add((f"border_{side_name}", str(side.style)))

    # Number format
    if cell.number_format and cell.number_format != "General":
        attrs.add(("number_format", str(cell.number_format)))

    return attrs


def check_sheet_qualifies(
    ws,
    min_filled_cells: int = DEFAULT_MIN_FILLED_CELLS,
    min_format_attributes: int = DEFAULT_MIN_FORMAT_ATTRIBUTES,
    min_format_coverage: float = DEFAULT_MIN_FORMAT_COVERAGE,
) -> Tuple[bool, Dict[str, Any]]:
    """
    Check whether a worksheet meets all cell-level selection conditions:
      - Condition 1: non-empty (at least min_filled_cells cells with a value)
      - Condition 2: format diversity (>= min_format_attributes distinct attribute types)
      - Condition 3: format coverage (>= min_format_coverage fraction of cells formatted)
      - Condition 4: no cross-sheet formulas (no formula containing '!')

    Returns (qualifies, stats_dict).
    """
    has_value = False
    all_format_types: Set[Tuple[str, str]] = set()
    total_cells = 0
    formatted_cells = 0
    condition2_met = False
    has_cross_sheet_formula = False

    for row in ws.iter_rows():
        if has_cross_sheet_formula:
            break

        for cell in row:
            total_cells += 1

            # Condition 1: non-empty
            if cell.value is not None:
                has_value = True

                # Condition 4: cross-sheet formula check
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    if "!" in cell.value:
                        has_cross_sheet_formula = True
                        break

            # Formatting (conditions 2 & 3)
            cell_attrs = classify_format_attributes(cell)
            if cell_attrs:
                formatted_cells += 1
                if not condition2_met:
                    all_format_types.update(cell_attrs)
                    if len(all_format_types) >= min_format_attributes:
                        condition2_met = True

    stats = {
        "total_cells": total_cells,
        "formatted_cells": formatted_cells,
        "format_coverage": formatted_cells / total_cells if total_cells > 0 else 0.0,
        "format_attribute_types": len(all_format_types),
        "format_attributes_found": sorted(f"{k}={v}" for k, v in all_format_types),
        "has_value": has_value,
        "has_cross_sheet_formula": has_cross_sheet_formula,
    }

    if not has_value:
        return False, stats
    if has_cross_sheet_formula:
        return False, stats
    if not condition2_met:
        return False, stats
    if total_cells == 0 or (formatted_cells / total_cells) < min_format_coverage:
        return False, stats

    return True, stats


# =============================================================================
# SHEET SELECTION LOGIC
# =============================================================================


def select_qualifying_sheet(
    workbook_path: Path,
    min_filled_cells: int,
    min_format_attributes: int,
    min_format_coverage: float,
    rng: random.Random,
) -> Optional[Tuple[str, Dict[str, Any]]]:
    """
    Open a workbook and find a qualifying sheet.

    1. Zip pre-check to exclude sheets with pivots/tables/chart sheets.
    2. openpyxl read_only scan for cell-level conditions.

    Returns (sheet_name, stats) if found, else None.
    """
    from openpyxl import load_workbook

    # Phase A: Zip pre-check (pivots, tables, chart sheets, password protection)
    excluded_sheets = zip_precheck(workbook_path)

    # Phase B: Integrity check — verify workbook opens in non-read_only mode
    # This catches openpyxl version incompatibilities (e.g., ColumnDimension widthPt)
    # and other structural issues that would cause failures during generation.
    try:
        wb_check = load_workbook(str(workbook_path), read_only=False, data_only=False)
        wb_check.close()
    except Exception as e:
        logger.debug(f"Integrity check failed for {workbook_path}: {e}")
        return None

    # Phase C: openpyxl read_only scan for cell-level conditions
    try:
        wb = load_workbook(str(workbook_path), read_only=True, data_only=False)
    except Exception as e:
        logger.warning(f"Failed to open {workbook_path}: {e}")
        return None

    try:
        # Get regular worksheet names (exclude chart sheets, blacklisted names)
        sheet_names = []
        for name in wb.sheetnames:
            if name in excluded_sheets:
                continue
            if name in BLACKLISTED_SHEET_NAMES:
                continue
            try:
                if hasattr(wb, "chartsheets"):
                    chartsheet_names = {cs.title for cs in wb.chartsheets}
                    if name in chartsheet_names:
                        continue
            except Exception:
                pass
            sheet_names.append(name)

        rng.shuffle(sheet_names)

        for sheet_name in sheet_names:
            try:
                ws = wb[sheet_name]
                qualifies, stats = check_sheet_qualifies(
                    ws,
                    min_filled_cells=min_filled_cells,
                    min_format_attributes=min_format_attributes,
                    min_format_coverage=min_format_coverage,
                )
                if qualifies:
                    return sheet_name, stats
            except Exception as e:
                logger.debug(f"Error checking sheet '{sheet_name}' in {workbook_path}: {e}")
                continue
    finally:
        wb.close()

    return None


# =============================================================================
# MAIN PIPELINE
# =============================================================================


# ---------------------------------------------------------------------------
# Timeout-guarded image capture (watchdog pattern)
# ---------------------------------------------------------------------------

_CAPTURE_TIMEOUT_SECONDS = 15  # Max seconds per workbook for image capture


def _kill_excel_by_pid(app) -> None:
    """Force-kill the Excel process behind *app* using taskkill."""
    try:
        pid = app.pid
        subprocess.run(
            ["taskkill", "/F", "/PID", str(pid)],
            capture_output=True,
            timeout=10,
        )
        logger.info(f"Force-killed Excel process (PID {pid})")
    except Exception as exc:
        logger.warning(f"Could not kill Excel by PID: {exc}")
    # Nuclear option: kill ALL Excel processes to ensure nothing lingers
    try:
        subprocess.run(
            ["taskkill", "/F", "/IM", "EXCEL.EXE"],
            capture_output=True,
            timeout=10,
        )
    except Exception:
        pass


class _CaptureWatchdog:
    """
    Watchdog timer that kills the Excel app if a capture takes too long.

    Usage::

        wd = _CaptureWatchdog(app, timeout=60)
        wd.start()
        try:
            save_images_add_headings(...)
            wd.cancel()          # capture finished in time
        except Exception:
            wd.cancel()
            if wd.fired:
                # Excel was killed by watchdog
                ...
            raise
    """

    def __init__(self, app, timeout: int = _CAPTURE_TIMEOUT_SECONDS, label: str = ""):
        self.app = app
        self.timeout = timeout
        self.label = label
        self.fired = False
        self._timer: Optional[threading.Timer] = None

    def _on_timeout(self) -> None:
        self.fired = True
        logger.warning(
            f"Watchdog fired after {self.timeout}s for {self.label}. "
            f"Killing Excel process."
        )
        _kill_excel_by_pid(self.app)

    def start(self) -> None:
        self._timer = threading.Timer(self.timeout, self._on_timeout)
        self._timer.daemon = True
        self._timer.start()

    def cancel(self) -> None:
        if self._timer is not None:
            self._timer.cancel()


def run_sampling(
    source_dir: Path,
    dest_dir: Path,
    num_sheets: int = DEFAULT_NUM_SHEETS,
    tracking_file: Optional[Path] = None,
    previous_tracking_file: Optional[Path] = None,
    min_filled_cells: int = DEFAULT_MIN_FILLED_CELLS,
    min_format_attributes: int = DEFAULT_MIN_FORMAT_ATTRIBUTES,
    min_format_coverage: float = DEFAULT_MIN_FORMAT_COVERAGE,
    seed: Optional[int] = DEFAULT_SEED,
    max_index: int = DEFAULT_MAX_INDEX,
    skip_images: bool = False,
    images_only: bool = False,
) -> None:
    """
    Main sampling pipeline.

    Phase 1: Discovery & Selection (openpyxl read_only)
    Phase 2: File copy (shutil)
    Phase 3: Image capture (xlwings)

    If images_only=True, skip Phase 1 & 2 and only capture missing images
    for samples already recorded in the tracking file.
    """
    rng = random.Random(seed)
    start_time = time.time()

    # --- Resolve tracking file path ---
    if tracking_file is None:
        tracking_file = dest_dir / "sampling_tracker.json"

    # --- Load tracking data ---
    tracking = load_tracking_file(tracking_file)
    if previous_tracking_file and previous_tracking_file != tracking_file:
        prev_tracking = load_tracking_file(previous_tracking_file)
        existing_workbooks = tracking.sampled_workbooks()
        for sample in prev_tracking.samples:
            if sample["workbook"] not in existing_workbooks:
                tracking.samples.append(sample)

    tracking.source_dir = str(source_dir)
    already_sampled = tracking.sampled_workbooks()
    logger.info(f"Loaded tracking data: {len(already_sampled)} workbooks already sampled")

    # --- Images-only mode: reconstruct selected list from tracking file ---
    if images_only:
        print("=" * 70)
        print("IMAGES-ONLY MODE: Capturing missing images from tracking file")
        print("=" * 70)

        selected = []
        for s in tracking.samples:
            wb_abs = str((source_dir / s["workbook"]).resolve())
            selected.append(SampledSheet(
                workbook_path=s["workbook"],
                workbook_abs_path=wb_abs,
                sheet_name=s["sheet"],
                dest_name=s["dest_name"],
                sampled_at=s.get("sampled_at", ""),
                format_attributes_found=0,
                format_coverage=0.0,
                total_cells=0,
                formatted_cells=0,
            ))
        print(f"Loaded {len(selected)} samples from tracking file")

    else:
        # ===== PHASE 1: DISCOVERY & SELECTION =====
        print("=" * 70)
        print("PHASE 1: DISCOVERY & SELECTION")
        print("=" * 70)

        all_workbooks = discover_workbooks(source_dir)
        print(f"Found {len(all_workbooks)} workbooks in {source_dir}")

        candidates = []
        for wb_path in all_workbooks:
            rel_path = str(wb_path.relative_to(source_dir))
            if rel_path not in already_sampled:
                candidates.append(wb_path)
        print(f"Candidates after filtering already-sampled: {len(candidates)}")

        if len(candidates) == 0:
            print("No new workbooks to sample from. Exiting.")
            return

        rng.shuffle(candidates)

        selected: List[SampledSheet] = []
        existing_dest_names: Set[str] = {s["dest_name"] for s in tracking.samples}
        skipped = 0

        print(f"Scanning workbooks for qualifying sheets (target: {num_sheets})...")
        with tqdm(total=min(num_sheets, len(candidates)), desc="Selecting", unit="sheet") as pbar:
            for wb_path in candidates:
                if len(selected) >= num_sheets:
                    break

                result = select_qualifying_sheet(
                    wb_path,
                    min_filled_cells=min_filled_cells,
                    min_format_attributes=min_format_attributes,
                    min_format_coverage=min_format_coverage,
                    rng=rng,
                )

                if result is None:
                    skipped += 1
                    continue

                sheet_name, stats = result
                dest_name = make_dest_name(wb_path, sheet_name, existing_dest_names)
                existing_dest_names.add(dest_name)

                rel_path = str(wb_path.relative_to(source_dir))
                sample = SampledSheet(
                    workbook_path=rel_path,
                    workbook_abs_path=str(wb_path),
                    sheet_name=sheet_name,
                    dest_name=dest_name,
                    sampled_at=datetime.now().isoformat(),
                    format_attributes_found=stats["format_attribute_types"],
                    format_coverage=stats["format_coverage"],
                    total_cells=stats["total_cells"],
                    formatted_cells=stats["formatted_cells"],
                )
                selected.append(sample)
                pbar.update(1)

        print(f"\nSelected {len(selected)} qualifying sheets")
        print(f"Scanned {len(selected) + skipped} workbooks ({skipped} did not qualify)")

        if not selected:
            print("No qualifying sheets found. Exiting.")
            return

        # --- Update tracking file immediately ---
        for s in selected:
            tracking.samples.append({
                "workbook": s.workbook_path,
                "sheet": s.sheet_name,
                "dest_name": s.dest_name,
                "sampled_at": s.sampled_at,
            })
        save_tracking_file(tracking_file, tracking)
        print(f"Tracking file updated: {tracking_file}")

        # ===== PHASE 2: COPY XLSX FILES =====
        print("\n" + "=" * 70)
        print("PHASE 2: COPYING XLSX FILES")
        print("=" * 70)

        xlsx_dir = dest_dir / "xlsx"
        xlsx_dir.mkdir(parents=True, exist_ok=True)

        copy_errors = 0
        for s in tqdm(selected, desc="Copying", unit="file"):
            src = Path(s.workbook_abs_path)
            dst = xlsx_dir / f"{s.dest_name}.xlsx"
            try:
                shutil.copy2(str(src), str(dst))
            except Exception as e:
                logger.error(f"Failed to copy {src} -> {dst}: {e}")
                copy_errors += 1

        print(f"Copied {len(selected) - copy_errors} files to {xlsx_dir}")
        if copy_errors:
            print(f"  ({copy_errors} copy errors)")

    # ===== PHASE 3: CAPTURE SHEET IMAGES =====
    if skip_images:
        print("\n[SKIP] Phase 3: Image capture (--skip-images)")
    else:
        print("\n" + "=" * 70)
        print("PHASE 3: CAPTURING SHEET IMAGES")
        print("=" * 70)

        images_dir = dest_dir / "sheet_images"
        images_dir.mkdir(parents=True, exist_ok=True)

        app = create_configured_app(visible=False, add_book=False)
        if app is None:
            print("WARNING: xlwings not available. Skipping image capture.")
        else:
            capture_errors = 0
            timeout_errors = 0
            skipped_existing = 0
            new_captures = 0
            _APP_RESTART_INTERVAL = 500
            try:
                for i, s in enumerate(tqdm(selected, desc="Capturing", unit="image")):
                    src = Path(s.workbook_abs_path)
                    output_path = images_dir / f"{s.dest_name}.png"

                    # Skip already-captured images (resume support)
                    if output_path.exists():
                        skipped_existing += 1
                        continue

                    # Restart Excel every N captures to prevent memory buildup
                    if new_captures > 0 and new_captures % _APP_RESTART_INTERVAL == 0:
                        logger.info(f"Restarting Excel app after {new_captures} new captures to free memory")
                        _safe_close_app(app)
                        app = create_configured_app(visible=False, add_book=False)

                    # Capture with watchdog timeout (2 attempts: 1 try + 1 retry)
                    _MAX_ATTEMPTS = 2
                    succeeded = False
                    for attempt in range(_MAX_ATTEMPTS):
                        watchdog = _CaptureWatchdog(
                            app, timeout=_CAPTURE_TIMEOUT_SECONDS, label=s.dest_name
                        )
                        try:
                            app = ensure_app_alive(app)
                            watchdog.start()
                            save_images_add_headings(
                                book=str(src),
                                output=str(output_path),
                                sheets=s.sheet_name,
                                app=app,
                                max_index=max_index,
                            )
                            watchdog.cancel()
                            new_captures += 1
                            succeeded = True
                            break
                        except Exception as e:
                            watchdog.cancel()
                            if watchdog.fired:
                                logger.error(
                                    f"Timeout capturing {s.dest_name} "
                                    f"(killed after {_CAPTURE_TIMEOUT_SECONDS}s)"
                                )
                                timeout_errors += 1
                                # Process already killed — don't call _safe_close_app
                                # (COM calls on dead app would block for minutes)
                                app = None
                                app = create_configured_app(visible=False, add_book=False)
                                break  # skip retry for timeouts
                            elif attempt == 0:
                                logger.warning(f"Attempt 1 failed for {s.dest_name}: {e} — retrying")
                            else:
                                logger.error(f"Failed to capture {s.dest_name} after 2 attempts: {e}")
                            # Restart Excel to clear bad COM state
                            _safe_close_app(app)
                            app = create_configured_app(visible=False, add_book=False)
                    if not succeeded:
                        capture_errors += 1
            finally:
                _safe_close_app(app)

            print(f"Captured {new_captures} new images to {images_dir}")
            if skipped_existing:
                print(f"  ({skipped_existing} already existed, skipped)")
            if capture_errors:
                print(f"  ({capture_errors} capture errors, {timeout_errors} timeouts)")

    # ===== SUMMARY =====
    elapsed = time.time() - start_time
    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)
    xlsx_dir = dest_dir / "xlsx"
    print(f"  Selected:  {len(selected)} sheets")
    print(f"  XLSX dir:  {xlsx_dir}")
    if not skip_images:
        print(f"  Images:    {dest_dir / 'sheet_images'}")
    print(f"  Tracking:  {tracking_file}")
    print(f"  Time:      {elapsed:.1f}s")


# =============================================================================
# CLI ENTRY POINT
# =============================================================================


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Sample qualifying worksheets from a directory of Excel workbooks.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--source_dir",
        type=str,
        required=True,
        help="Directory containing .xlsx workbooks (searched recursively).",
    )
    parser.add_argument(
        "--dest_dir",
        type=str,
        required=True,
        help="Destination directory for output (xlsx/, sheet_images/, tracking file).",
    )
    parser.add_argument(
        "--num_sheets",
        type=int,
        default=DEFAULT_NUM_SHEETS,
        help=f"Number of sheets to sample (default: {DEFAULT_NUM_SHEETS}).",
    )
    parser.add_argument(
        "--tracking_file",
        type=str,
        default=None,
        help="Path to tracking JSON file. Default: dest_dir/sampling_tracker.json.",
    )
    parser.add_argument(
        "--previous_tracking",
        type=str,
        default=None,
        help="Path to a previous tracking file to merge (avoids re-sampling those workbooks).",
    )
    parser.add_argument(
        "--min_format_attributes",
        type=int,
        default=DEFAULT_MIN_FORMAT_ATTRIBUTES,
        help=f"Min distinct formatting attribute types (default: {DEFAULT_MIN_FORMAT_ATTRIBUTES}).",
    )
    parser.add_argument(
        "--min_format_coverage",
        type=float,
        default=DEFAULT_MIN_FORMAT_COVERAGE,
        help=f"Min fraction of cells with non-default formatting (default: {DEFAULT_MIN_FORMAT_COVERAGE}).",
    )
    parser.add_argument(
        "--seed",
        type=int,
        default=DEFAULT_SEED,
        help="Random seed for reproducibility.",
    )
    parser.add_argument(
        "--max_index",
        type=int,
        default=DEFAULT_MAX_INDEX,
        help=f"Max row/col index for image capture (default: {DEFAULT_MAX_INDEX}).",
    )
    parser.add_argument(
        "--skip-images",
        action="store_true",
        help="Skip image capture phase (useful for selection-only runs).",
    )
    parser.add_argument(
        "--images-only",
        action="store_true",
        help="Skip selection & copy, only capture missing images from tracking file.",
    )
    parser.add_argument(
        "-v",
        "--verbose",
        action="store_true",
        help="Enable verbose logging.",
    )
    return parser.parse_args()


def main():
    args = parse_args()

    log_level = logging.DEBUG if args.verbose else logging.INFO
    logging.basicConfig(
        level=log_level,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    source_dir = Path(args.source_dir).resolve()
    dest_dir = Path(args.dest_dir).resolve()

    if not source_dir.is_dir():
        print(f"ERROR: Source directory does not exist: {source_dir}")
        sys.exit(1)

    tracking_file = Path(args.tracking_file) if args.tracking_file else None
    previous_tracking = Path(args.previous_tracking) if args.previous_tracking else None

    run_sampling(
        source_dir=source_dir,
        dest_dir=dest_dir,
        num_sheets=args.num_sheets,
        tracking_file=tracking_file,
        previous_tracking_file=previous_tracking,
        min_filled_cells=DEFAULT_MIN_FILLED_CELLS,
        min_format_attributes=args.min_format_attributes,
        min_format_coverage=args.min_format_coverage,
        seed=args.seed,
        max_index=args.max_index,
        skip_images=args.skip_images,
        images_only=args.images_only,
    )


if __name__ == "__main__":
    main()
