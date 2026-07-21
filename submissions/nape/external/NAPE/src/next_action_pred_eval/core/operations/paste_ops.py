"""
Paste operations - PasteFrom operation.
"""

import copy
import json
from typing import Any, Dict, List

from next_action_pred_eval.core.cell_range import CellRange
from next_action_pred_eval.core.operation import Operation
from next_action_pred_eval.utils.cell_utils import expand_range, get_cell_address
from next_action_pred_eval.core.operations._helpers import (
    _ensure_sheet,
    _ensure_cell,
)


def adjust_formula_references(formula: str, row_offset: int, col_offset: int, source_sheet: str) -> str:
    """Adjust formula cell references by the given row and column offsets.

    This is a simplified implementation that handles basic A1-style references.
    For full Excel formula support, consider using a proper formula parser.

    Args:
        formula: The formula to adjust
        row_offset: Number of rows to shift references
        col_offset: Number of columns to shift references
        source_sheet: The source sheet name (for cross-sheet reference handling)

    Returns:
        Adjusted formula string
    """
    import re
    from openpyxl.utils import get_column_letter, column_index_from_string

    if not formula or not formula.startswith('='):
        return formula

    def adjust_cell_ref(match):
        """Adjust a single cell reference."""
        full_match = match.group(0)

        # Check for absolute references ($)
        col_part = match.group(1) or ''
        col_abs = match.group(2) or ''
        col = match.group(3)
        row_abs = match.group(4) or ''
        row = match.group(5)

        # Handle column adjustment
        if col_abs != '$' and col_offset != 0:
            try:
                col_idx = column_index_from_string(col)
                new_col_idx = col_idx + col_offset
                if new_col_idx < 1:
                    new_col_idx = 1
                col = get_column_letter(new_col_idx)
            except (ValueError, TypeError):
                pass

        # Handle row adjustment
        if row_abs != '$' and row_offset != 0:
            try:
                new_row = int(row) + row_offset
                if new_row < 1:
                    new_row = 1
                row = str(new_row)
            except (ValueError, TypeError):
                pass

        return f'{col_part}{col_abs}{col}{row_abs}{row}'

    # Pattern matches cell references like A1, $A1, A$1, $A$1
    # Also handles sheet references like Sheet1!A1
    # (?<![A-Za-z]) prevents matching within function names (e.g., DAYS360)
    # {1,3} limits column to valid Excel range (A–XFD)
    cell_pattern = r"((?:[A-Za-z_][A-Za-z0-9_]*!)?)(?<![A-Za-z])(\$?)([A-Z]{1,3})(\$?)(\d+)"

    return re.sub(cell_pattern, adjust_cell_ref, formula)


# ============= Operation Class =============

class PasteFrom(Operation):
    """Paste/copy operation from source range to destination range."""
    source_range: str  # Store as string representation like "Sheet1!A1:B2"

    def to_symbolic(self) -> str:
        if self.is_inverse:
            return f"PASTE_FROM | {self.cell_range} | clear"
        return f"PASTE_FROM | {self.cell_range} | {self.source_range} | {self.value}"

    def to_officejs(self, sheet_var: str = "sheet") -> str:
        if '!' in self.source_range:
            source_sheet, source_range_addr = self.source_range.split('!', 1)
            source_sheet = source_sheet.strip("'")
            source_range_ref = f'context.workbook.worksheets.getItem("{source_sheet}").getRange("{source_range_addr}")'
        else:
            source_range_ref = f'{sheet_var}.getRange("{self.source_range}")'

        mode_map = {
            'all': 'Excel.RangeCopyType.all',
            'values': 'Excel.RangeCopyType.values',
            'formats': 'Excel.RangeCopyType.formats',
            'formulas': 'Excel.RangeCopyType.formulas',
        }
        copy_type = mode_map.get(self.value.lower(), 'Excel.RangeCopyType.all')

        return f'{sheet_var}.getRange("{self.cell_range.range}").copyFrom({source_range_ref}, {copy_type});'

    def to_openpyxl(self, sheet_var: str = "ws") -> str:
        return f'# PasteFrom operation not supported in openpyxl - requires manual cell-by-cell copy'

    def to_xlwings(self, sheet_var: str = "sheet") -> str:
        if '!' in self.source_range:
            source_sheet, source_range_addr = self.source_range.split('!', 1)
            source_sheet = source_sheet.strip("'")
            source_range_ref = f'{sheet_var}.book.sheets["{source_sheet}"].range("{source_range_addr}")'
        else:
            source_range_ref = f'{sheet_var}.range("{self.source_range}")'

        return f'copy_range({source_range_ref}, {sheet_var}.range("{self.cell_range.range}"), paste="{self.value.lower()}")'

    @classmethod
    def from_symbolic(cls, symbolic: str) -> 'PasteFrom':
        """Create PasteFrom operation from symbolic representation."""
        parts = [p.strip() for p in symbolic.split('|', 3)]
        if len(parts) < 2 or parts[0] != 'PASTE_FROM':
            raise ValueError(f"Invalid PASTE_FROM symbolic format: {symbolic}")

        cell_range = CellRange.from_string(parts[1])

        if len(parts) > 2 and parts[2] == 'clear':
            return cls(
                cell_range=cell_range,
                source_range="",
                value="all",
                is_inverse=True
            )

        if len(parts) < 4:
            raise ValueError(f"Invalid PASTE_FROM symbolic format (missing source_range or value): {symbolic}")

        return cls(
            cell_range=cell_range,
            source_range=parts[2],
            value=parts[3],
            is_inverse=False
        )

    def apply_to_state(self, state: Dict[str, Any]) -> None:
        """Apply PasteFrom to state by copying data from source to destination range.

        Behavior differs by paste mode:
        - 'all', 'values', 'formulas': Use source dimensions
        - 'formats':
          - Single-cell destination: Use source dimensions
          - Multi-cell destination: Tile source pattern to fill destination
        """
        source_cell_range = CellRange.from_string(self.source_range)

        src_rows, src_cols = source_cell_range.get_dimensions()
        dst_rows, dst_cols = self.cell_range.get_dimensions()

        dst_start_row, dst_start_col, dst_end_row, dst_end_col = self.cell_range.get_coordinates()
        src_start_row, src_start_col, src_end_row, src_end_col = source_cell_range.get_coordinates()

        row_offset = dst_start_row - src_start_row
        col_offset = dst_start_col - src_start_col

        source_sheet = _ensure_sheet(state, source_cell_range.sheet)
        dest_sheet = _ensure_sheet(state, self.cell_range.sheet)

        paste_mode = self.value.lower()

        is_single_cell_dest = (dst_rows == 1 and dst_cols == 1)

        if paste_mode == 'formats' and not is_single_cell_dest:
            iterate_rows = dst_rows
            iterate_cols = dst_cols
            use_tiling = True
        else:
            iterate_rows = src_rows
            iterate_cols = src_cols
            use_tiling = False

        for row_idx in range(iterate_rows):
            for col_idx in range(iterate_cols):
                dst_row = dst_start_row + row_idx
                dst_col = dst_start_col + col_idx
                dst_cell_addr = get_cell_address(dst_row, dst_col)

                if use_tiling:
                    src_row_idx = row_idx % src_rows
                    src_col_idx = col_idx % src_cols
                else:
                    src_row_idx = row_idx
                    src_col_idx = col_idx

                src_row = src_start_row + src_row_idx
                src_col = src_start_col + src_col_idx
                src_cell_addr = get_cell_address(src_row, src_col)

                src_cell = source_sheet["cells"].get(src_cell_addr, {})
                dst_cell = _ensure_cell(state, self.cell_range.sheet, dst_cell_addr)

                if paste_mode in ['all', 'values']:
                    if 'value' in src_cell:
                        dst_cell['value'] = copy.deepcopy(src_cell['value'])
                    elif 'value' in dst_cell:
                        del dst_cell['value']

                    if 'datatype' in src_cell:
                        dst_cell['datatype'] = src_cell['datatype']
                    elif 'datatype' in dst_cell and paste_mode == 'all':
                        del dst_cell['datatype']

                if paste_mode in ['all', 'formulas']:
                    if 'formula' in src_cell:
                        formula = src_cell['formula']
                        adjusted_formula = adjust_formula_references(
                            formula, row_offset, col_offset, source_cell_range.sheet
                        )
                        dst_cell['formula'] = adjusted_formula
                    elif 'formula' in dst_cell:
                        del dst_cell['formula']

                if paste_mode in ['all', 'formats']:
                    if 'Format' in src_cell:
                        dst_cell['Format'] = copy.deepcopy(src_cell['Format'])
                    elif 'Format' in dst_cell:
                        del dst_cell['Format']

        if paste_mode in ['all', 'formats']:
            self._copy_merge_cells(state, source_cell_range, row_offset, col_offset,
                                   src_rows, src_cols, iterate_rows, iterate_cols, use_tiling)

    def _copy_merge_cells(self, state: Dict[str, Any], source_range: CellRange,
                         row_offset: int, col_offset: int, src_rows: int, src_cols: int,
                         iterate_rows: int, iterate_cols: int, use_tiling: bool) -> None:
        """Copy merge cell information from source to destination."""
        source_sheet = _ensure_sheet(state, source_range.sheet)
        dest_sheet = _ensure_sheet(state, self.cell_range.sheet)

        src_start_row, src_start_col, src_end_row, src_end_col = source_range.get_coordinates()
        dst_start_row, dst_start_col, _, _ = self.cell_range.get_coordinates()
        dst_end_row = dst_start_row + iterate_rows - 1
        dst_end_col = dst_start_col + iterate_cols - 1

        source_merged = source_sheet.get("worksheetProperties", {}).get("merged_cells", [])

        source_merges_in_range = []
        for merge_info in source_merged:
            m_start_row = merge_info["start_row"]
            m_start_col = merge_info["start_col"]
            m_end_row = merge_info["end_row"]
            m_end_col = merge_info["end_col"]

            if (m_start_row >= src_start_row and m_end_row <= src_end_row and
                m_start_col >= src_start_col and m_end_col <= src_end_col):
                source_merges_in_range.append({
                    "rel_start_row": m_start_row - src_start_row,
                    "rel_start_col": m_start_col - src_start_col,
                    "rel_end_row": m_end_row - src_start_row,
                    "rel_end_col": m_end_col - src_start_col
                })

        if not source_merges_in_range:
            return

        if "worksheetProperties" not in dest_sheet:
            dest_sheet["worksheetProperties"] = {"merged_cells": []}
        if "merged_cells" not in dest_sheet["worksheetProperties"]:
            dest_sheet["worksheetProperties"]["merged_cells"] = []

        if use_tiling:
            tile_rows = (iterate_rows + src_rows - 1) // src_rows
            tile_cols = (iterate_cols + src_cols - 1) // src_cols

            for tile_row in range(tile_rows):
                for tile_col in range(tile_cols):
                    tile_row_offset = tile_row * src_rows
                    tile_col_offset = tile_col * src_cols

                    for rel_merge in source_merges_in_range:
                        new_merge = {
                            "start_row": dst_start_row + tile_row_offset + rel_merge["rel_start_row"],
                            "start_col": dst_start_col + tile_col_offset + rel_merge["rel_start_col"],
                            "end_row": dst_start_row + tile_row_offset + rel_merge["rel_end_row"],
                            "end_col": dst_start_col + tile_col_offset + rel_merge["rel_end_col"]
                        }

                        if (new_merge["start_row"] >= dst_start_row and
                            new_merge["end_row"] <= dst_end_row and
                            new_merge["start_col"] >= dst_start_col and
                            new_merge["end_col"] <= dst_end_col):

                            if new_merge not in dest_sheet["worksheetProperties"]["merged_cells"]:
                                dest_sheet["worksheetProperties"]["merged_cells"].append(new_merge)
        else:
            for rel_merge in source_merges_in_range:
                new_merge = {
                    "start_row": dst_start_row + rel_merge["rel_start_row"],
                    "start_col": dst_start_col + rel_merge["rel_start_col"],
                    "end_row": dst_start_row + rel_merge["rel_end_row"],
                    "end_col": dst_start_col + rel_merge["rel_end_col"]
                }

                if (new_merge["start_row"] >= dst_start_row and
                    new_merge["end_row"] <= dst_end_row and
                    new_merge["start_col"] >= dst_start_col and
                    new_merge["end_col"] <= dst_end_col):

                    if new_merge not in dest_sheet["worksheetProperties"]["merged_cells"]:
                        dest_sheet["worksheetProperties"]["merged_cells"].append(new_merge)

    def get_inverse(self) -> 'Operation':
        """Return operation to clear the pasted content/formatting."""
        from next_action_pred_eval.core.operations.value_ops import SetValue, SetFormula

        paste_mode = self.value.lower()

        if paste_mode == 'values':
            return SetValue(cell_range=self.cell_range, value=None, is_inverse=True)
        elif paste_mode == 'formats':
            return SetValue(cell_range=self.cell_range, value=None, is_inverse=True)
        elif paste_mode == 'formulas':
            return SetFormula(cell_range=self.cell_range, value=None, is_inverse=True)
        else:
            return SetValue(cell_range=self.cell_range, value=None, is_inverse=True)

    @property
    def modifies_format(self) -> bool:
        return self.value.lower() in ['formats', 'all']
