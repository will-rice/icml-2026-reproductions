"""
AutoFill operations - AutoFill (drag-fill) operation.

Simulates Excel's AutoFill behavior: given a source range and a destination
range, extends patterns from the source into the fill area.

Supported patterns:
- Arithmetic number series (constant step)
- Text copy/cycle
- Text+number extension (with bounce at 0 for negative steps)
- Custom lists: months (full/short), days (full/short), Q1-Q4 (case-preserving)
- Formula reference adjustment (row or column offset)
- Format tiling from source
- Position-based mixed-type independence

Exempted edge cases (documented):
- Diagonal single-cell fill (rejected by geometry validation)
- Growth/geometric trends (Excel's xlGrowthTrend only, not default)
- Non-linear extrapolation (uses average-of-diffs, not regression)
- Leading-zero text becoming numbers ("001","002" → numeric 3,4,5)
- Date-aware monthly fill (uses day intervals, not month-end logic)
"""

import copy
import re
from typing import Any, Dict, List, Optional, Tuple

from pydantic import model_validator

from next_action_pred_eval.core.cell_range import CellRange
from next_action_pred_eval.core.operation import Operation
from next_action_pred_eval.utils.cell_utils import get_cell_address
from next_action_pred_eval.core.operations._helpers import (
    _ensure_sheet,
    _ensure_cell,
    _get_cells_in_range,
)
from next_action_pred_eval.core.operations.paste_ops import adjust_formula_references


# ============================================================================
# Custom list definitions (matching Excel's built-in AutoFill lists)
# ============================================================================

_MONTHS_FULL = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]
_MONTHS_SHORT = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                 "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
_DAYS_FULL = ["Monday", "Tuesday", "Wednesday", "Thursday",
              "Friday", "Saturday", "Sunday"]
_DAYS_SHORT = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
_QUARTER_LABELS = ["Q1", "Q2", "Q3", "Q4"]
_QUARTER_FULL = ["Quarter 1", "Quarter 2", "Quarter 3", "Quarter 4"]

_ALL_CUSTOM_LISTS = [
    _MONTHS_FULL, _MONTHS_SHORT, _DAYS_FULL, _DAYS_SHORT,
    _QUARTER_LABELS, _QUARTER_FULL,
]


# ============================================================================
# Pattern detection helpers
# ============================================================================

def _match_custom_list(values: List[str]) -> Optional[Tuple[List[str], int]]:
    """Check if values match a known custom list (months, days, quarters).

    Returns (list, start_index) if matched, else None.
    Matching is case-insensitive; the original list casing is used for output.
    """
    if not values or not all(isinstance(v, str) for v in values):
        return None

    for lst in _ALL_CUSTOM_LISTS:
        lower_lst = [x.lower() for x in lst]
        first_lower = values[0].lower()
        if first_lower not in lower_lst:
            continue
        start_idx = lower_lst.index(first_lower)
        expected = [lower_lst[(start_idx + i) % len(lst)] for i in range(len(values))]
        actual = [v.lower() for v in values]
        if actual == expected:
            return lst, start_idx

    return None


def _detect_case_style(sample: str) -> str:
    """Detect if text is UPPER, lower, or Title case."""
    if sample.isupper():
        return "upper"
    elif sample.islower():
        return "lower"
    return "title"


def _apply_case_style(text: str, style: str) -> str:
    """Apply case style to text."""
    if style == "upper":
        return text.upper()
    elif style == "lower":
        return text.lower()
    return text  # title case is the default in our lists


def _detect_text_number_pattern(values: List[str]) -> Optional[Tuple[str, List[int], int]]:
    """Detect text+number suffix pattern like 'Item 1', 'Item 2'.

    Only matches trailing numeric suffix.
    Returns (prefix, numbers_list, step) or None.
    """
    if not values or not all(isinstance(v, str) for v in values):
        return None

    prefixes = []
    numbers = []
    for v in values:
        match = re.match(r'^(.*?)(\d+)$', v)
        if not match:
            return None
        prefixes.append(match.group(1))
        numbers.append(int(match.group(2)))

    if len(set(prefixes)) != 1:
        return None

    if len(numbers) == 1:
        return prefixes[0], numbers, 1  # single value, step=1

    diffs = [numbers[i + 1] - numbers[i] for i in range(len(numbers) - 1)]
    if len(set(diffs)) == 1:
        return prefixes[0], numbers, diffs[0]

    return None


def _extend_custom_list(lst: List[str], start_idx: int, n_source: int,
                        n_fill: int, case_style: str, direction: int) -> List[Any]:
    """Extend a custom list sequence."""
    result = []
    for fill_idx in range(n_fill):
        if direction == 1:
            idx = (start_idx + n_source + fill_idx) % len(lst)
        else:
            idx = (start_idx - 1 - fill_idx) % len(lst)
        result.append(_apply_case_style(lst[idx], case_style))
    return result


def _extend_text_number(prefix: str, numbers: List[int], step: int,
                        n_fill: int, direction: int) -> List[str]:
    """Extend a text+number pattern, bouncing at 0 for negative steps.

    Text+number patterns bounce at 0: they clamp to the nearest non-negative
    value and then reverse direction, matching Excel's behavior.
    Raw numbers (without text prefix) do NOT bounce — they go negative freely.
    """
    result = []
    last_num = numbers[-1] if direction == 1 else numbers[0]
    current_step = step * direction

    for _ in range(n_fill):
        next_num = last_num + current_step
        if next_num < 0:
            # Bounce: reflect around 0 and reverse step
            next_num = abs(next_num)
            current_step = abs(current_step)
        result.append(f"{prefix}{next_num}")
        last_num = next_num

    return result


def _extend_number_series(values: List[float], n_fill: int, direction: int) -> List[float]:
    """Extend an arithmetic number series.

    Uses constant difference if uniform, otherwise average of differences
    (linear extrapolation). Numbers can go negative (no bounce).
    """
    if len(values) == 1:
        return [values[0]] * n_fill

    diffs = [values[i + 1] - values[i] for i in range(len(values) - 1)]
    step = diffs[0] if len(set(diffs)) == 1 else sum(diffs) / len(diffs)

    result = []
    last = values[-1] if direction == 1 else values[0]
    effective_step = step * direction

    for _ in range(n_fill):
        last = last + effective_step
        result.append(last)

    return result


def _extend_uniform_values(values: List[Any], n_fill: int, direction: int) -> List[Any]:
    """Extend values of a single uniform type."""
    if not values or n_fill == 0:
        return []

    if all(v is None for v in values):
        return [None] * n_fill

    non_none = [v for v in values if v is not None]
    if not non_none:
        return [None] * n_fill

    sample = non_none[0]

    # Boolean: cycle
    if isinstance(sample, bool):
        return [values[i % len(values)] for i in range(n_fill)]

    # Numbers (int or float, but not bool)
    if isinstance(sample, (int, float)) and not isinstance(sample, bool):
        num_values = [float(v) for v in values if v is not None]
        extended = _extend_number_series(num_values, n_fill, direction)
        # Preserve int type if all source values were int
        if all(isinstance(v, int) for v in values if v is not None):
            extended = [int(round(v)) if v == int(round(v)) else v for v in extended]
        return extended

    # Strings
    if isinstance(sample, str):
        str_values = [v for v in values if isinstance(v, str)]

        # Check custom lists first
        custom = _match_custom_list(str_values)
        if custom:
            lst, start_idx = custom
            case_style = _detect_case_style(str_values[0])
            return _extend_custom_list(lst, start_idx, len(str_values), n_fill,
                                       case_style, direction)

        # Check text+number pattern
        tn_pattern = _detect_text_number_pattern(str_values)
        if tn_pattern:
            prefix, numbers, step = tn_pattern
            return _extend_text_number(prefix, numbers, step, n_fill, direction)

        # Default: cycle text values
        return [str_values[i % len(str_values)] for i in range(n_fill)]

    # Default: cycle
    return [values[i % len(values)] for i in range(n_fill)]


def _detect_and_extend_values(values: List[Any], n_fill: int,
                              direction: int, n_source: int) -> List[Any]:
    """Detect pattern in values and extend by n_fill items.

    Handles position-based mixed-type logic: each position in the source
    cycle is treated independently when types differ across positions.
    """
    if not values:
        return [None] * n_fill

    if all(v is None for v in values):
        return [None] * n_fill

    # Determine types present
    all_types = set()
    for v in values:
        if v is None:
            all_types.add(type(None))
        elif isinstance(v, bool):
            all_types.add(bool)
        elif isinstance(v, (int, float)):
            all_types.add(float)
        elif isinstance(v, str):
            all_types.add(str)
        else:
            all_types.add(type(v))

    # Uniform type: analyze as a single sequence
    if len(all_types) == 1:
        return _extend_uniform_values(values, n_fill, direction)

    # Mixed types: each position is independent
    result = []
    pos_extensions: Dict[int, List[Any]] = {}

    for pos in range(n_source):
        pos_vals = [values[i] for i in range(len(values)) if i % n_source == pos]
        n_pos_fill = sum(1 for i in range(n_fill) if i % n_source == pos)
        pos_extensions[pos] = _extend_uniform_values(pos_vals, n_pos_fill, direction)

    pos_counters = {pos: 0 for pos in range(n_source)}
    for fill_idx in range(n_fill):
        pos = fill_idx % n_source
        if pos in pos_extensions and pos_counters[pos] < len(pos_extensions[pos]):
            result.append(pos_extensions[pos][pos_counters[pos]])
            pos_counters[pos] += 1
        else:
            result.append(values[fill_idx % len(values)])

    return result


def _fill_column_values(
    source_vals: List[Any],
    source_formulas: List[Optional[str]],
    n_fill: int,
    direction: int,
    n_source: int,
    is_horizontal: bool = False,
) -> List[Dict[str, Any]]:
    """Fill values for a single column (or row for horizontal fill).

    Handles position-based mixed-type patterns where each position
    in the source cycle is analyzed independently.

    Args:
        source_vals: Source cell values
        source_formulas: Source cell formulas (None if no formula)
        n_fill: Number of cells to fill
        direction: +1 (down/right) or -1 (up/left)
        n_source: Number of source cells
        is_horizontal: True for horizontal fill (adjusts columns instead of rows)

    Returns:
        List of dicts with 'value' and 'formula' keys
    """
    result: List[Dict[str, Any]] = [{"value": None, "formula": None}] * n_fill

    # Pass 1: Handle formula slots
    for fill_idx in range(n_fill):
        slot = fill_idx % n_source
        src_formula = source_formulas[slot]

        if src_formula and src_formula.startswith('='):
            cycles = fill_idx // n_source + 1
            offset = cycles * n_source * direction
            if is_horizontal:
                adjusted = adjust_formula_references(src_formula, 0, offset, "")
            else:
                adjusted = adjust_formula_references(src_formula, offset, 0, "")
            result[fill_idx] = {"value": None, "formula": adjusted}

    # Pass 2: Handle value slots
    value_slot_indices = []
    for slot in range(n_source):
        if not (source_formulas[slot] and source_formulas[slot].startswith('=')):
            value_slot_indices.append(slot)

    if not value_slot_indices:
        return result  # All formulas

    value_slot_vals = [source_vals[i] for i in value_slot_indices]

    if len(value_slot_indices) == n_source:
        # All slots are values (no formulas mixed in)
        extended = _detect_and_extend_values(value_slot_vals, n_fill, direction, n_source)
        ext_idx = 0
        for fill_idx in range(n_fill):
            if ext_idx < len(extended):
                result[fill_idx] = {"value": extended[ext_idx], "formula": None}
                ext_idx += 1
    else:
        # Mixed formula + value slots: extend each value slot independently
        fill_counts_per_slot = {}
        for slot in value_slot_indices:
            fill_counts_per_slot[slot] = sum(
                1 for fi in range(n_fill) if fi % n_source == slot
            )

        for slot in value_slot_indices:
            slot_val = source_vals[slot]
            n_slot_fill = fill_counts_per_slot[slot]
            slot_extended = _extend_uniform_values([slot_val], n_slot_fill, direction)

            ext_idx = 0
            for fill_idx in range(n_fill):
                if fill_idx % n_source == slot and ext_idx < len(slot_extended):
                    result[fill_idx] = {"value": slot_extended[ext_idx], "formula": None}
                    ext_idx += 1

    return result


# ============================================================================
# AutoFill Operation Class
# ============================================================================

class AutoFill(Operation):
    """AutoFill (drag-fill) operation.

    Simulates Excel's AutoFill: given a source range and a destination range
    (which includes the source), fills the destination by extending patterns
    from the source. Direction is inferred from geometry.

    The cell_range field holds the **destination** range (the full target).
    The value field holds the **source** range as a string (the seed cells).

    Symbolic format: AUTOFILL | Sheet1!A1:C8 | Sheet1!A1:C3
    """

    @model_validator(mode='before')
    @classmethod
    def validate_geometry(cls, data: Any) -> Any:
        """Validate that source ⊂ destination and direction is inferrable."""
        if isinstance(data, dict):
            cell_range = data.get('cell_range')  # destination
            value = data.get('value')  # source
            is_inverse = data.get('is_inverse', False)

            if is_inverse or cell_range is None or value is None:
                return data

            if isinstance(value, str) and value:
                try:
                    source = CellRange.from_string(value)
                except Exception:
                    return data

                if cell_range.sheet != source.sheet:
                    raise ValueError(
                        f"Destination sheet '{cell_range.sheet}' must match "
                        f"source sheet '{source.sheet}'"
                    )

                src_r1, src_c1, src_r2, src_c2 = source.get_coordinates()
                dst_r1, dst_c1, dst_r2, dst_c2 = cell_range.get_coordinates()

                if not source.is_subset(cell_range):
                    raise ValueError(
                        f"Source range {source} must be a subset of "
                        f"destination range {cell_range}"
                    )

                cols_match = (src_c1 == dst_c1 and src_c2 == dst_c2)
                rows_match = (src_r1 == dst_r1 and src_r2 == dst_r2)

                if cols_match and rows_match:
                    raise ValueError("Source and destination are the same range")

                if not cols_match and not rows_match:
                    raise ValueError(
                        f"AutoFill requires same columns (vertical) or same rows "
                        f"(horizontal). Source {source} and dest {cell_range} differ "
                        f"on both axes."
                    )

        return data

    def _infer_direction(self) -> str:
        """Infer fill direction from source and destination geometry."""
        source = CellRange.from_string(self.value)
        src_r1, src_c1, src_r2, src_c2 = source.get_coordinates()
        dst_r1, dst_c1, dst_r2, dst_c2 = self.cell_range.get_coordinates()

        cols_match = (src_c1 == dst_c1 and src_c2 == dst_c2)
        if cols_match:
            return 'down' if dst_r2 > src_r2 else 'up'
        else:
            return 'right' if dst_c2 > src_c2 else 'left'

    def _get_fill_range(self) -> CellRange:
        """Get the CellRange of cells that will be filled (excluding source)."""
        from openpyxl.utils import get_column_letter

        source = CellRange.from_string(self.value)
        direction = self._infer_direction()
        src_r1, src_c1, src_r2, src_c2 = source.get_coordinates()
        dst_r1, dst_c1, dst_r2, dst_c2 = self.cell_range.get_coordinates()

        if direction == 'down':
            fill_range = f"{get_column_letter(src_c1)}{src_r2 + 1}:{get_column_letter(src_c2)}{dst_r2}"
        elif direction == 'up':
            fill_range = f"{get_column_letter(src_c1)}{dst_r1}:{get_column_letter(src_c2)}{src_r1 - 1}"
        elif direction == 'right':
            fill_range = f"{get_column_letter(src_c2 + 1)}{src_r1}:{get_column_letter(dst_c2)}{src_r2}"
        else:  # left
            fill_range = f"{get_column_letter(dst_c1)}{src_r1}:{get_column_letter(src_c1 - 1)}{src_r2}"

        return CellRange(sheet=self.cell_range.sheet, range=fill_range)

    def to_symbolic(self) -> str:
        if self.is_inverse:
            return f"AUTOFILL | {self.cell_range} | clear"
        return f"AUTOFILL | {self.cell_range} | {self.value}"

    @classmethod
    def from_symbolic(cls, symbolic: str) -> 'AutoFill':
        parts = [p.strip() for p in symbolic.split('|', 2)]
        if len(parts) < 3 or parts[0] != 'AUTOFILL':
            raise ValueError(f"Invalid AUTOFILL symbolic format: {symbolic}")

        cell_range = CellRange.from_string(parts[1])

        if parts[2] == 'clear':
            return cls(cell_range=cell_range, value="", is_inverse=True)

        return cls(cell_range=cell_range, value=parts[2].strip(), is_inverse=False)

    def apply_to_state(self, state: Dict[str, Any]) -> None:
        """Apply AutoFill by extending patterns from source into fill area.

        This operation is state-dependent: it reads source cell values, formulas,
        and formatting from the current state to detect patterns.
        """
        if self.is_inverse:
            fill_range = self._get_fill_range()
            for cell_addr in _get_cells_in_range(fill_range):
                cell = _ensure_cell(state, fill_range.sheet, cell_addr)
                cell.pop("value", None)
                cell.pop("formula", None)
                cell.pop("datatype", None)
                cell.pop("number_format", None)
                cell.pop("Format", None)
            return

        direction = self._infer_direction()
        source = CellRange.from_string(self.value)
        src_r1, src_c1, src_r2, src_c2 = source.get_coordinates()
        dst_r1, dst_c1, dst_r2, dst_c2 = self.cell_range.get_coordinates()

        sheet_name = self.cell_range.sheet
        sheet_data = _ensure_sheet(state, sheet_name)
        is_vertical = direction in ('down', 'up')

        if is_vertical:
            self._fill_vertical(state, sheet_data, sheet_name, direction,
                                src_r1, src_c1, src_r2, src_c2, dst_r1, dst_r2)
        else:
            self._fill_horizontal(state, sheet_data, sheet_name, direction,
                                  src_r1, src_c1, src_r2, src_c2, dst_c1, dst_c2)

    def _fill_vertical(self, state, sheet_data, sheet_name, direction,
                       src_r1, src_c1, src_r2, src_c2, dst_r1, dst_r2):
        """Fill vertically (down or up)."""
        n_source = src_r2 - src_r1 + 1
        n_fill = (dst_r2 - src_r2) if direction == 'down' else (src_r1 - dst_r1)
        dir_sign = 1 if direction == 'down' else -1

        for col in range(src_c1, src_c2 + 1):
            source_vals, source_formulas, source_formats, source_nformats, fav = (
                self._read_source_column(sheet_data, src_r1, src_r2, col, is_vertical=True)
            )

            fill_data = _fill_column_values(
                source_vals, source_formulas, n_fill, dir_sign, n_source
            )

            for fill_idx in range(n_fill):
                dst_row = (src_r2 + 1 + fill_idx) if direction == 'down' else (src_r1 - 1 - fill_idx)
                self._write_fill_cell(
                    state, sheet_name, dst_row, col,
                    fill_data[fill_idx], source_formats, source_nformats,
                    fill_idx, n_source, formulas_as_values=fav
                )

    def _fill_horizontal(self, state, sheet_data, sheet_name, direction,
                         src_r1, src_c1, src_r2, src_c2, dst_c1, dst_c2):
        """Fill horizontally (right or left)."""
        n_source = src_c2 - src_c1 + 1
        n_fill = (dst_c2 - src_c2) if direction == 'right' else (src_c1 - dst_c1)
        dir_sign = 1 if direction == 'right' else -1

        for row in range(src_r1, src_r2 + 1):
            source_vals, source_formulas, source_formats, source_nformats, fav = (
                self._read_source_column(sheet_data, src_c1, src_c2, row, is_vertical=False)
            )

            fill_data = _fill_column_values(
                source_vals, source_formulas, n_fill, dir_sign,
                n_source, is_horizontal=True
            )

            for fill_idx in range(n_fill):
                dst_col = (src_c2 + 1 + fill_idx) if direction == 'right' else (src_c1 - 1 - fill_idx)
                self._write_fill_cell(
                    state, sheet_name, row, dst_col,
                    fill_data[fill_idx], source_formats, source_nformats,
                    fill_idx, n_source, formulas_as_values=fav
                )

    def _read_source_column(self, sheet_data, start, end, fixed_coord, is_vertical):
        """Read source values, formulas, formats from state for one column/row."""
        vals, formulas, formats, nformats = [], [], [], []
        formulas_from_values = False
        for offset in range(end - start + 1):
            if is_vertical:
                addr = get_cell_address(start + offset, fixed_coord)
            else:
                addr = get_cell_address(fixed_coord, start + offset)
            cell_data = sheet_data.get("cells", {}).get(addr, {})
            val = cell_data.get("value")
            formula = cell_data.get("formula")
            # Detect formula-like values stored via SetInput (INPUT ops)
            if formula is None and isinstance(val, str) and val.startswith("="):
                formula = val
                val = None
                formulas_from_values = True
            vals.append(val)
            formulas.append(formula)
            formats.append(copy.deepcopy(cell_data.get("Format")))
            nformats.append(cell_data.get("number_format"))
        return vals, formulas, formats, nformats, formulas_from_values

    @staticmethod
    def _write_fill_cell(state, sheet_name, row, col, fill_datum,
                         source_formats, source_nformats, fill_idx, n_source,
                         formulas_as_values=False):
        """Write a single filled cell to state."""
        dst_cell_addr = get_cell_address(row, col)
        dst_cell = _ensure_cell(state, sheet_name, dst_cell_addr)

        if fill_datum.get("formula"):
            if formulas_as_values:
                # Source had formulas stored as values (via SetInput/INPUT),
                # so write filled cells the same way for state consistency.
                dst_cell["value"] = fill_datum["formula"]
                dst_cell.pop("formula", None)
            else:
                dst_cell["formula"] = fill_datum["formula"]
                dst_cell.pop("value", None)
        elif fill_datum.get("value") is not None:
            dst_cell["value"] = fill_datum["value"]
            dst_cell.pop("formula", None)
        else:
            dst_cell.pop("value", None)
            dst_cell.pop("formula", None)

        # Tile formatting from source
        fmt_idx = fill_idx % n_source
        src_fmt = source_formats[fmt_idx]
        if src_fmt:
            dst_cell["Format"] = copy.deepcopy(src_fmt)
        else:
            dst_cell.pop("Format", None)

        src_nf = source_nformats[fmt_idx]
        if src_nf:
            dst_cell["number_format"] = src_nf
        else:
            dst_cell.pop("number_format", None)

    def get_inverse(self) -> 'Operation':
        """Return operation that clears the filled cells (values + formatting)."""
        return AutoFill(cell_range=self.cell_range, value=self.value, is_inverse=True)

    @property
    def modifies_format(self) -> bool:
        return True

    def to_officejs(self, sheet_var: str = "sheet") -> str:
        if self.is_inverse:
            fill_range = self._get_fill_range()
            return f'{sheet_var}.getRange("{fill_range.range}").clear(Excel.ClearApplyTo.all);'
        source = CellRange.from_string(self.value)
        return (
            f'{sheet_var}.getRange("{source.range}")'
            f'.autoFill({sheet_var}.getRange("{self.cell_range.range}"), '
            f'Excel.AutoFillType.fillDefault);'
        )

    def to_openpyxl(self, sheet_var: str = "ws") -> str:
        return '# AutoFill not natively supported in openpyxl — requires manual pattern extension'

    def to_xlwings(self, sheet_var: str = "sheet") -> str:
        if self.is_inverse:
            fill_range = self._get_fill_range()
            return f'{sheet_var}.range("{fill_range.range}").clear()'
        source = CellRange.from_string(self.value)
        return (
            f'{sheet_var}.range("{source.range}").api.AutoFill('
            f'{sheet_var}.range("{self.cell_range.range}").api, 0)'
        )
