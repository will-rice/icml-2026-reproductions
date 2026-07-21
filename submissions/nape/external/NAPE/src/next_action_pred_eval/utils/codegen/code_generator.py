from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple, Union
import re
from next_action_pred_eval.core.operation import Operation
from next_action_pred_eval.core.operations import MergeCells
INDENT = "    "

class OfficeJSGenerator:
    def generate(self, operations: List[Operation], comment_cell_ranges: bool = True,
                 add_wrap: bool = True, add_helper_functions: bool = True) -> str:
        """Generate Office.js code from operations"""
        if operations is None:
            operations = self.operations

        # Group operations by sheet
        sheet_operations = {}
        for op in operations:
            sheet_name = op.cell_range.sheet
            if sheet_name not in sheet_operations:
                sheet_operations[sheet_name] = []
            sheet_operations[sheet_name].append(op)

        # Generate code
        code_lines = []

        for sheet_name, ops in sheet_operations.items():
            var_name = self._sanitize_sheet_name(sheet_name)

            code_lines.extend(
                [
                    f"// Sheet: {sheet_name}",
                    f"let {var_name} = context.workbook.worksheets.getItemOrNullObject('{sheet_name}');",
                    "await context.sync();",
                    f"if ({var_name}.isNullObject) {{",
                    f"{INDENT}{var_name} = context.workbook.worksheets.add('{sheet_name}');",
                    "}",
                    "",
                ]
            )

            # Group similar operations for better readability
            current_cell = None
            for op in ops:
                if op.cell_range.range != current_cell:
                    if current_cell:
                        code_lines.append("")  # Add blank line between cells
                    current_cell = op.cell_range.range
                    if comment_cell_ranges:
                        code_lines.append(f"// Cell: {current_cell}")

                code_lines.append(f"{op.to_officejs(var_name)}")

            code_lines.append("")

        code = "\n".join(code_lines)

        if add_wrap:
            code = OfficeJSGenerator.wrap_officejs(code)
            if add_helper_functions:
                code += """

function setOutsideBorders(range, weight, style, color) {
    const sides = ['EdgeTop', 'EdgeBottom', 'EdgeLeft', 'EdgeRight'];
    sides.forEach(side => {
        range.format.borders.getItem(side).style = style;
        range.format.borders.getItem(side).weight = weight;
        if (color) {
            range.format.borders.getItem(side).color = color;
        }
    });
}

function setAllBorders(range, weight, style, color) {
    const sides = [
        'EdgeTop', 'EdgeBottom', 'EdgeLeft', 'EdgeRight',
        'InsideHorizontal', 'InsideVertical'
    ];
    sides.forEach(side => {
        range.format.borders.getItem(side).style = style;
        range.format.borders.getItem(side).weight = weight;
        if (color) {
            range.format.borders.getItem(side).color = color;
        }
    });
}"""

        return code.strip()

    @staticmethod
    def wrap_officejs(code: str) -> str:
        if not code.startswith("Excel.run(async (context) =>"):
            code = "\n".join(f"{INDENT}{line}" for line in code.splitlines())
            code = f"""\
await Excel.run(async (context) => {{
{code}
{INDENT}await context.sync();
}});
""".strip()
        return code

    def _sanitize_sheet_name(self, name: str) -> str:
        """Convert sheet name to valid JavaScript variable name"""
        # Remove non-alphanumeric characters and convert to camelCase
        parts = re.split(r"[^a-zA-Z0-9]", name)
        if not parts:
            return "sheet"

        # First part lowercase, rest title case
        sanitized = parts[0].lower() if parts[0] else "sheet"
        for part in parts[1:]:
            if part:
                sanitized += part[0].upper() + part[1:].lower()

        # Ensure it starts with a letter
        if sanitized and not sanitized[0].isalpha():
            sanitized = "sheet" + sanitized

        return sanitized or "sheet"

class PythonGenerator:
    def generate(self, operations: List[Operation], comment_cell_ranges: bool = True,
                 add_wrap: bool = True, workbook_path: Union[str, Path] = "output.xlsx",
                 add_helper_functions: bool = True) -> str:
        """Generate openpyxl Python code from operations"""
        if operations is None:
            return ""

        # Group operations by sheet
        sheet_operations = {}
        for op in operations:
            sheet_name = op.cell_range.sheet
            if sheet_name not in sheet_operations:
                sheet_operations[sheet_name] = []
            sheet_operations[sheet_name].append(op)

        # Generate code
        code_lines = []

        for sheet_name, ops in sheet_operations.items():
            var_name = self._sanitize_sheet_name(sheet_name)
            code_lines.append(f"# Sheet: {sheet_name}")
            code_lines.append(f"if '{sheet_name}' in wb.sheetnames:")
            code_lines.append(f"{INDENT}{var_name} = wb['{sheet_name}']")
            code_lines.append(f"else:")
            code_lines.append(f"{INDENT}{var_name} = wb.create_sheet('{sheet_name}')")
            code_lines.append("")

            current_cell = None
            for op in ops:
                if op.cell_range.range != current_cell:
                    if current_cell:
                        code_lines.append("")  # Add blank line between cells
                    current_cell = op.cell_range.range
                    if comment_cell_ranges:
                        code_lines.append(f"# Cell: {current_cell}")

                code_lines.append(f"{op.to_openpyxl(var_name)}")

            code_lines.append("")

        code_lines.append(f"# Save workbook")
        code_lines.append(f"wb.save('{workbook_path}')")

        code = "\n".join(code_lines)

        if add_wrap:
            code = PythonGenerator.wrap_python(code)

        return code.strip()

    @staticmethod
    def wrap_python(code: str, add_helper_functions: bool = True) -> str:
        """Wrap Python code with imports and function definition"""
        if not code.startswith("import openpyxl"):
            code = "\n".join(f"{line}" for line in code.splitlines()) # removed indentation
            code = f"""
import openpyxl
from datetime import datetime, time
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl_utils.cellrange import cell_range

wb = openpyxl.Workbook()
{code}
""" # global scope
        return code

    def _sanitize_sheet_name(self, name: str) -> str:
        """Convert sheet name to valid Python variable name"""
        parts = re.split(r"[^a-zA-Z0-9]", name)
        if not parts:
            return "sheet"
        sanitized = parts[0].lower() if parts[0] else "sheet"
        for part in parts[1:]:
            if part:
                sanitized += part[0].upper() + part[1:].lower()
        if sanitized and not sanitized[0].isalpha():
            sanitized = "sheet" + sanitized
        return sanitized or "sheet"

class XlwingsGenerator:
    def generate(self, operations: List[Operation], comment_cell_ranges: bool = True,
                 add_wrap: bool = True, workbook_path: Optional[Union[str, Path]] = None,
                 add_helper_functions: bool = True, add_snapshots: bool = False, snapshot_path: Optional[Union[str, Path]] = None) -> str:
        """Generate xlwings Python code from operations"""
        if operations is None:
            return ""

        # Group operations by sheet
        sheet_operations = {}
        for op in operations:
            sheet_name = op.cell_range.sheet
            if sheet_name not in sheet_operations:
                sheet_operations[sheet_name] = []
            sheet_operations[sheet_name].append(op)

        # Generate code
        code_lines = []

        for sheet_name, ops in sheet_operations.items():
            var_name = self._sanitize_sheet_name(sheet_name)
            code_lines.append(f"# Sheet: {sheet_name}")
            code_lines.append(f"try:")
            code_lines.append(f"{INDENT}{var_name} = wb.sheets['{sheet_name}']")
            code_lines.append(f"except:")
            code_lines.append(f"{INDENT}{var_name} = wb.sheets.add('{sheet_name}')")
            code_lines.append("")

            if add_snapshots and snapshot_path:
                code_lines.append(f"step_counter = 0")
                code_lines.append("")

            current_cell = None
            for op in ops:
                if op.cell_range.range != current_cell:
                    if current_cell:
                        code_lines.append("")  # Add blank line between cells
                    current_cell = op.cell_range.range
                    if comment_cell_ranges:
                        code_lines.append(f"# Cell: {current_cell}")

                code_lines.append(f"{op.to_xlwings(var_name)}")
                if add_snapshots and snapshot_path:
                    # For MergeCells operations, autofit columns and rows to fix wrap_text height issues
                    if isinstance(op, MergeCells) and op.value:  # Only for merge (not unmerge)
                        cell_range = op.cell_range.range
                        code_lines.append(f"# Autofit columns and rows for merged range to fix wrap_text height issues")
                        code_lines.append(f"{var_name}.range('{cell_range}').columns.autofit()")
                        code_lines.append(f"{var_name}.range('{cell_range}').rows.autofit()")
                    code_lines.append(f"# Take snapshot after operation")
                    code_lines.append(f"step_counter += 1")
                    code_lines.append(f"# Capture from A1 to ensure full content is included")
                    code_lines.append(f"{var_name}.range('A1', {var_name}.used_range.last_cell).to_png(rf'{snapshot_path}/{sheet_name}_step_' + str(step_counter) + '.png')")

            code_lines.append("")
        if workbook_path:
            code_lines.append(f"# Save workbook")
            code_lines.append(f"wb.save('{workbook_path}')")
        code_lines.append(f"wb.close()")
        code_lines.append(f"app.quit()")

        code = "\n".join(code_lines)

        if add_wrap:
            code = XlwingsGenerator.wrap_xlwings(code)

        return code.strip()

    @staticmethod
    def wrap_xlwings(code: str, add_helper_functions: bool = True) -> str:
        """Wrap xlwings code with imports and setup"""
        if not code.startswith("import xlwings"):
            code = "\n".join(f"{line}" for line in code.splitlines())

            helper_functions = ""
            if add_helper_functions:
                helper_functions = """
def copy_range(source, dest, paste='all'):
    source.copy()
    dest.paste(paste=paste)
    wb.app.api.CutCopyMode = False

def set_border_side(rng, border_index, line_style, weight, color):
    \"\"\"Set a single border side with style, weight, and color\"\"\"
    rng.api.Borders(border_index).LineStyle = line_style
    rng.api.Borders(border_index).Weight = weight
    if color and color != 'None':
        color_hex = color.lstrip('#')
        r, g, b = int(color_hex[0:2], 16), int(color_hex[2:4], 16), int(color_hex[4:6], 16)
        rng.api.Borders(border_index).Color = r + (g * 256) + (b * 65536)

def set_border_outside(rng, line_style, weight, color):
    \"\"\"Set outside borders (top, bottom, left, right) for a range\"\"\"
    sides = [
        BordersIndex.xlEdgeTop,
        BordersIndex.xlEdgeBottom,
        BordersIndex.xlEdgeLeft,
        BordersIndex.xlEdgeRight
    ]
    for side in sides:
        set_border_side(rng, side, line_style, weight, color)

def set_border_all(rng, line_style, weight, color):
    \"\"\"Set all borders (outside + inside) for a range\"\"\"
    sides = [
        BordersIndex.xlEdgeTop,
        BordersIndex.xlEdgeBottom,
        BordersIndex.xlEdgeLeft,
        BordersIndex.xlEdgeRight,
        BordersIndex.xlInsideHorizontal,
        BordersIndex.xlInsideVertical
    ]
    for side in sides:
        set_border_side(rng, side, line_style, weight, color)
"""

            code = f"""
import xlwings as xw
from xlwings.constants import HAlign, VAlign, LineStyle, BordersIndex, UnderlineStyle

{helper_functions}
# Create or open workbook with dialog suppression
app = xw.App(visible=False)
app.display_alerts = False
app.screen_updating = False
try:
    app.api.AskToUpdateLinks = False
except AttributeError:
    pass
try:
    app.api.EnableEvents = False
except AttributeError:
    pass
wb = app.books.add()

{code}
"""
        return code

    def _sanitize_sheet_name(self, name: str) -> str:
        """Convert sheet name to valid Python variable name"""
        parts = re.split(r"[^a-zA-Z0-9]", name)
        if not parts:
            return "sheet"
        sanitized = parts[0].lower() if parts[0] else "sheet"
        for part in parts[1:]:
            if part:
                sanitized += part[0].upper() + part[1:].lower()
        if sanitized and not sanitized[0].isalpha():
            sanitized = "sheet" + sanitized
        return sanitized or "sheet"
