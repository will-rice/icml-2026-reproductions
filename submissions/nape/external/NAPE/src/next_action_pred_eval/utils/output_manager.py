"""
Output Manager - Organizes structured output directories for experiments.

Manages experiment directory structures, saves LLM responses with metadata,
tracks token usage, and generates summary reports.

Pure Python utility with no internal dependencies on next_action_pred_eval modules.
"""

import json
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional
import shutil
import openpyxl
from dataclasses import dataclass, asdict


@dataclass
class TokenUsage:
    """Token usage information"""
    total_tokens: int
    completion_tokens: int
    prompt_tokens: int


@dataclass
class LLMCallMetadata:
    """Metadata for a single LLM call"""
    timestamp: str
    model_name: str
    time_taken: float
    token_usage: TokenUsage
    prompt_length: int
    response_length: int
    success: bool
    error_message: Optional[str] = None


@dataclass
class RegionMetadata:
    """Metadata for a region"""
    sheet_name: str
    category: str
    description: str
    range: str
    llm_calls: List[LLMCallMetadata]
    symbolic_operations_count: Dict[str, int]  # original, llm_raw, reordered
    region_image_path: str
    created_at: str


@dataclass
class SheetMetadata:
    """Metadata for a sheet"""
    sheet_name: str
    total_regions: int
    sheet_range: str
    llm_calls: List[LLMCallMetadata]
    total_operations: int
    sheet_image_path: str
    regions: List[str]  # region identifiers
    region_categories: Dict[str, str]  # region category dict: region -> category
    created_at: str


@dataclass
class WorkbookMetadata:
    """Metadata for the entire workbook"""
    workbook_name: str
    workbook_path: str
    total_sheets: int
    total_regions: int
    total_llm_calls: int
    total_llm_time_taken: float  # Time spent on LLM calls only
    total_experiment_time: float  # Total time for complete experiment
    total_tokens: TokenUsage
    experiment_config: Dict[str, Any]
    sheets: List[str]  # sheet names
    created_at: str
    experiment_start_time: float  # Internal tracking of start time
    output_path: str


class OutputManager:
    """Manages structured output for experiment runs.

    Creates a directory hierarchy for each experiment, saves LLM call
    metadata, symbolic operations, generated code, and summary reports.
    """

    def __init__(self, base_output_dir: str = "outputs"):
        self.base_output_dir = Path(base_output_dir)
        self.base_output_dir.mkdir(exist_ok=True)
        self.current_experiment_dir: Optional[Path] = None
        self.workbook_metadata: Optional[WorkbookMetadata] = None

    def start_experiment(self, workbook_path: str, experiment_config: Dict[str, Any]) -> Path:
        """Start a new experiment and create the directory structure"""
        workbook_name = Path(workbook_path).stem
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        experiment_name = f"{workbook_name}_{timestamp}"

        self.current_experiment_dir = self.base_output_dir / experiment_name
        self.current_experiment_dir.mkdir(parents=True, exist_ok=True)

        # Create directory structure
        (self.current_experiment_dir / "sheets").mkdir(exist_ok=True)
        (self.current_experiment_dir / "code").mkdir(exist_ok=True)

        wb = openpyxl.load_workbook(workbook_path)
        experiment_start_time = time.time()
        self.workbook_metadata = WorkbookMetadata(
            workbook_name=workbook_name,
            workbook_path=str(Path(workbook_path).absolute()),
            total_sheets=len(wb.sheetnames),
            total_regions=0,
            total_llm_calls=0,
            total_llm_time_taken=0.0,
            total_experiment_time=0.0,
            total_tokens=TokenUsage(0, 0, 0),
            experiment_config=experiment_config,
            sheets=wb.sheetnames,
            created_at=datetime.now().isoformat(),
            experiment_start_time=experiment_start_time,
            output_path=str(self.current_experiment_dir.absolute())
        )

        return self.current_experiment_dir

    def create_sheet_structure(self, sheet_name: str) -> Path:
        """Create directory structure for a sheet"""
        if not self.current_experiment_dir:
            raise ValueError("No experiment started. Call start_experiment first.")

        sheet_dir = self.current_experiment_dir / "sheets" / sheet_name
        sheet_dir.mkdir(parents=True, exist_ok=True)
        (sheet_dir / "regions").mkdir(exist_ok=True)
        (sheet_dir / "code").mkdir(exist_ok=True)

        return sheet_dir

    def create_region_structure(self, sheet_name: str, region_range: str) -> Path:
        """Create directory structure for a region"""
        sheet_dir = self.current_experiment_dir / "sheets" / sheet_name
        region_id = f"{region_range.replace(':', '_')}"
        region_dir = sheet_dir / "regions" / region_id
        region_dir.mkdir(parents=True, exist_ok=True)
        (region_dir / "symbolic").mkdir(exist_ok=True)
        (region_dir / "code").mkdir(exist_ok=True)
        (region_dir / "execution").mkdir(exist_ok=True)

        return region_dir

    def save_llm_response(
        self,
        response_text: str,
        prompt: str,
        model_name: str,
        time_taken: float,
        save_path: Path,
        file_prefix: str = "llm_output",
        usage: Optional[Dict[str, int]] = None,
    ) -> LLMCallMetadata:
        """Save LLM response and return metadata.

        Args:
            response_text: The raw text response from the LLM.
            prompt: The prompt that was sent.
            model_name: Name/identifier of the model used.
            time_taken: Wall-clock time for the call in seconds.
            save_path: Directory in which to save output files.
            file_prefix: Filename prefix for saved files.
            usage: Optional token usage dict with keys
                   ``total_tokens``, ``completion_tokens``, ``prompt_tokens``.

        Returns:
            LLMCallMetadata with call details.
        """
        if usage is None:
            usage = {"total_tokens": 0, "completion_tokens": 0, "prompt_tokens": 0}

        # Save raw text
        with open(save_path / f"{file_prefix}.txt", "w", encoding="utf-8") as f:
            f.write(response_text)

        # Save formatted markdown
        with open(save_path / f"{file_prefix}.md", "w", encoding="utf-8") as f:
            f.write("# LLM Response\n\n")
            f.write(f"**Timestamp:** {datetime.now().isoformat()}\n\n")
            f.write(f"**Model:** {model_name}\n\n")
            f.write(f"**Time Taken:** {time_taken:.2f} seconds\n\n")
            f.write(
                f"**Tokens Used:** {usage.get('total_tokens', 0)} "
                f"(completion: {usage.get('completion_tokens', 0)}, "
                f"prompt: {usage.get('prompt_tokens', 0)})\n\n"
            )
            f.write(f"## Prompt\n\n```\n{prompt}\n```\n\n")
            f.write(f"## Response\n\n{response_text}\n")

        # Create metadata
        metadata = LLMCallMetadata(
            timestamp=datetime.now().isoformat(),
            model_name=model_name,
            time_taken=time_taken,
            token_usage=TokenUsage(
                total_tokens=usage.get("total_tokens", 0),
                completion_tokens=usage.get("completion_tokens", 0),
                prompt_tokens=usage.get("prompt_tokens", 0),
            ),
            prompt_length=len(prompt),
            response_length=len(response_text),
            success=True,
        )

        return metadata

    def save_symbolic_operations(
        self,
        operations: List[str],
        save_path: Path,
        filename: str,
    ):
        """Save symbolic operations to file"""
        with open(save_path / f"{filename}.txt", "w", encoding="utf-8") as f:
            f.write("\n".join(operations))

    def save_code(
        self,
        code: str,
        save_path: Path,
        filename: str,
        language: str = "python",
    ):
        """Save generated code"""
        ext = "py" if language == "python" else "js"
        with open(save_path / f"{filename}.{ext}", "w", encoding="utf-8") as f:
            f.write(code)

    def save_region_metadata(
        self,
        sheet_name: str,
        region_data: Dict[str, Any],
        llm_calls: List[LLMCallMetadata],
        symbolic_counts: Dict[str, int],
    ) -> RegionMetadata:
        """Save region metadata"""
        region_dir = (
            self.current_experiment_dir
            / "sheets"
            / sheet_name
            / "regions"
            / f"{region_data['range'].replace(':', '_')}"
        )

        metadata = RegionMetadata(
            sheet_name=region_data["sheet_name"],
            category=region_data["category"],
            description=region_data["description"],
            range=region_data["range"],
            llm_calls=llm_calls,
            symbolic_operations_count=symbolic_counts,
            region_image_path=str(region_dir / "region_image.png"),
            created_at=datetime.now().isoformat(),
        )

        with open(region_dir / "metadata.json", "w", encoding="utf-8") as f:
            json.dump(asdict(metadata), f, indent=2)

        return metadata

    def save_sheet_metadata(
        self,
        sheet_name: str,
        sheet_data: Dict[str, Any],
        llm_calls: List[LLMCallMetadata],
        regions: List[str],
        region_categories: Dict[str, str],
    ) -> SheetMetadata:
        """Save sheet metadata"""
        sheet_dir = self.current_experiment_dir / "sheets" / sheet_name

        metadata = SheetMetadata(
            sheet_name=sheet_name,
            total_regions=len(regions),
            sheet_range=sheet_data.get("range", ""),
            llm_calls=llm_calls,
            total_operations=sheet_data.get("total_operations", 0),
            sheet_image_path=str(sheet_dir / f"{sheet_name}.png"),
            regions=regions,
            region_categories=region_categories,
            created_at=datetime.now().isoformat(),
        )

        with open(sheet_dir / "metadata.json", "w", encoding="utf-8") as f:
            json.dump(asdict(metadata), f, indent=2)

        return metadata

    def create_steps_sheet_file(self, sheet_name: str) -> Optional[Path]:
        """
        Create a combined steps_sheet.txt file for a sheet by concatenating
        reordered symbolic steps from all regions in the order specified in metadata.

        Args:
            sheet_name: Name of the sheet to process

        Returns:
            Path to the created steps_sheet.txt file, or None if creation failed
        """
        if not self.current_experiment_dir:
            raise ValueError("No experiment in progress")

        sheet_dir = self.current_experiment_dir / "sheets" / sheet_name
        metadata_path = sheet_dir / "metadata.json"

        if not metadata_path.exists():
            return None

        # Read sheet metadata to get region order
        with open(metadata_path, "r", encoding="utf-8") as f:
            sheet_metadata = json.load(f)

        regions = sheet_metadata.get("regions", [])
        regions_dir = sheet_dir / "regions"

        if not regions_dir.exists():
            return None

        # Concatenate reordered symbolic steps from all regions
        all_steps = []

        for region in regions:
            region_dir = regions_dir / region

            if not region_dir.exists():
                continue

            reordered_file = region_dir / "symbolic" / "reordered.txt"

            if not reordered_file.exists():
                continue

            # Read the reordered symbolic steps
            with open(reordered_file, "r", encoding="utf-8") as f:
                steps = f.read().strip()

            if steps:
                # Convert region_id back to range format for display (e.g., B2_I4 -> B2:I4)
                region_range = region.replace("_", ":", 1)
                all_steps.append(f"# Region: {region_range}")
                all_steps.append(steps)
                all_steps.append("")  # Empty line between regions

        # Write the concatenated steps to steps_sheet.txt
        if all_steps:
            output_file = sheet_dir / "steps_sheet.txt"
            with open(output_file, "w", encoding="utf-8") as f:
                f.write("\n".join(all_steps))
            return output_file

        return None

    def finalize_experiment(self):
        """Finalize experiment and save workbook metadata"""
        if not self.workbook_metadata or not self.current_experiment_dir:
            raise ValueError("No experiment in progress")

        # Update totals by aggregating from sheet metadata
        total_regions = 0
        total_llm_calls = 0
        total_llm_time = 0.0
        total_tokens = TokenUsage(0, 0, 0)

        for sheet_name in self.workbook_metadata.sheets:
            sheet_metadata_path = (
                self.current_experiment_dir / "sheets" / sheet_name / "metadata.json"
            )
            if sheet_metadata_path.exists():
                with open(sheet_metadata_path, "r") as f:
                    sheet_meta = json.load(f)
                    total_regions += sheet_meta["total_regions"]
                    total_llm_calls += len(sheet_meta["llm_calls"])
                    for call in sheet_meta["llm_calls"]:
                        total_llm_time += call["time_taken"]
                        total_tokens.total_tokens += call["token_usage"]["total_tokens"]
                        total_tokens.completion_tokens += call["token_usage"][
                            "completion_tokens"
                        ]
                        total_tokens.prompt_tokens += call["token_usage"]["prompt_tokens"]

        # Calculate total experiment time
        total_experiment_time = time.time() - self.workbook_metadata.experiment_start_time

        self.workbook_metadata.total_regions = total_regions
        self.workbook_metadata.total_llm_calls = total_llm_calls
        self.workbook_metadata.total_llm_time_taken = total_llm_time
        self.workbook_metadata.total_experiment_time = total_experiment_time
        self.workbook_metadata.total_tokens = total_tokens

        # Save workbook metadata (exclude internal tracking fields)
        metadata_dict = asdict(self.workbook_metadata)
        # Remove internal tracking field from saved metadata
        metadata_dict.pop("experiment_start_time", None)

        with open(self.current_experiment_dir / "metadata.json", "w", encoding="utf-8") as f:
            json.dump(metadata_dict, f, indent=2)

        # Generate summary report
        self._generate_summary_report()

    def _generate_summary_report(self):
        """Generate a markdown summary report"""
        report_path = self.current_experiment_dir / "summary_report.md"

        with open(report_path, "w", encoding="utf-8") as f:
            f.write(f"# Experiment Summary: {self.workbook_metadata.workbook_name}\n\n")
            f.write(f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            f.write("## Overview\n\n")
            f.write(f"- **Workbook:** {self.workbook_metadata.workbook_name}\n")
            f.write(f"- **Total Sheets:** {self.workbook_metadata.total_sheets}\n")
            f.write(f"- **Total Regions:** {self.workbook_metadata.total_regions}\n")
            f.write(f"- **Total LLM Calls:** {self.workbook_metadata.total_llm_calls}\n")
            f.write(
                f"- **Total Experiment Time:** "
                f"{self.workbook_metadata.total_experiment_time:.2f} seconds\n"
            )
            f.write(
                f"- **LLM Processing Time:** "
                f"{self.workbook_metadata.total_llm_time_taken:.2f} seconds\n"
            )
            f.write(
                f"- **Total Tokens:** {self.workbook_metadata.total_tokens.total_tokens}\n\n"
            )

            f.write("## Configuration\n\n")
            f.write(
                f"```json\n{json.dumps(self.workbook_metadata.experiment_config, indent=2)}\n```\n\n"
            )

            f.write("## Sheets\n\n")
            for sheet_name in self.workbook_metadata.sheets:
                f.write(f"- [{sheet_name}](sheets/{sheet_name}/)\n")

            f.write("\n## Logs\n\n")
            logs_dir = self.current_experiment_dir / "logs"
            if logs_dir.exists():
                log_files = list(logs_dir.glob("*.log"))
                if log_files:
                    f.write(f"- **Log Files:** {len(log_files)} file(s)\n")
                    for log_file in sorted(log_files):
                        f.write(f"  - [{log_file.name}](logs/{log_file.name})\n")
                else:
                    f.write("- No log files found\n")
            else:
                f.write("- Logs directory not found\n")

            f.write("\n## Output Structure\n\n")
            f.write("```\n")
            self._write_directory_tree(f, self.current_experiment_dir, prefix="")
            f.write("```\n")

    def _write_directory_tree(self, f, path: Path, prefix: str = "", max_depth: int = 4):
        """Write directory tree to file"""
        if max_depth <= 0:
            return

        items = sorted(path.iterdir(), key=lambda x: (x.is_file(), x.name))
        for i, item in enumerate(items):
            is_last = i == len(items) - 1
            current_prefix = "└── " if is_last else "├── "
            f.write(f"{prefix}{current_prefix}{item.name}\n")

            if item.is_dir() and not item.name.startswith("."):
                next_prefix = prefix + ("    " if is_last else "│   ")
                self._write_directory_tree(f, item, next_prefix, max_depth - 1)
