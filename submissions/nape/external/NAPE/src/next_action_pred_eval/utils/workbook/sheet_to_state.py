"""
Sheet to State Conversion.

Convert openpyxl workbooks to state dictionaries.
"""

import logging
from datetime import datetime, date, time, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Theme color resolution
# ---------------------------------------------------------------------------

def _parse_theme_colors(wb) -> List[str]:
    """Extract the 12 theme colors from a workbook's theme XML.

    Returns a list of 12 hex color strings (without '#'), e.g. ['000000', 'FFFFFF', ...].
    Returns an empty list if the theme cannot be parsed.
    """
    try:
        theme_bytes = getattr(wb, 'loaded_theme', None)
        if not theme_bytes:
            return []
        from lxml import etree
        root = etree.fromstring(theme_bytes)
        ns = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}
        clr_scheme = root.find('.//a:themeElements/a:clrScheme', ns)
        if clr_scheme is None:
            return []
        colors: List[str] = []
        for name in ['dk1', 'lt1', 'dk2', 'lt2',
                     'accent1', 'accent2', 'accent3', 'accent4',
                     'accent5', 'accent6', 'hlink', 'folHlink']:
            el = clr_scheme.find(f'a:{name}', ns)
            if el is None:
                colors.append('000000')
                continue
            srgb = el.find('a:srgbClr', ns)
            sys_clr = el.find('a:sysClr', ns)
            if srgb is not None:
                colors.append(srgb.get('val', '000000'))
            elif sys_clr is not None:
                colors.append(sys_clr.get('lastClr', sys_clr.get('val', '000000')))
            else:
                colors.append('000000')
        return colors
    except Exception:
        return []


def _apply_tint(hex_color: str, tint: float) -> str:
    """Apply a tint/shade modifier to a hex color string.

    Positive tint → lighten towards white, negative → darken towards black.
    """
    if tint == 0.0:
        return hex_color
    try:
        r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
        if tint > 0:
            r = int(r + (255 - r) * tint)
            g = int(g + (255 - g) * tint)
            b = int(b + (255 - b) * tint)
        else:
            factor = 1 + tint  # tint is negative, so this is < 1
            r = int(r * factor)
            g = int(g * factor)
            b = int(b * factor)
        r, g, b = max(0, min(255, r)), max(0, min(255, g)), max(0, min(255, b))
        return f"{r:02X}{g:02X}{b:02X}"
    except Exception:
        return hex_color


def _resolve_openpyxl_color(color_obj, theme_colors: List[str]) -> Optional[str]:
    """Resolve an openpyxl Color object to a hex string (without '#').

    Handles rgb, theme, and indexed color types. Returns None if unresolvable.
    """
    if color_obj is None:
        return None

    color_type = getattr(color_obj, 'type', None)

    # --- Explicit RGB ---
    if color_type == 'rgb':
        rgb = color_obj.rgb
        if isinstance(rgb, str):
            return rgb
        return None

    # --- Theme-based ---
    if color_type == 'theme' and theme_colors:
        theme_idx = color_obj.theme
        if isinstance(theme_idx, int) and 0 <= theme_idx < len(theme_colors):
            base = theme_colors[theme_idx]
            tint = getattr(color_obj, 'tint', 0.0)
            if not isinstance(tint, (int, float)):
                tint = 0.0
            return _apply_tint(base, tint)
        return None

    # --- Indexed ---
    if color_type == 'indexed':
        idx = color_obj.indexed
        if isinstance(idx, int):
            from openpyxl.styles.colors import COLOR_INDEX
            if 0 <= idx < len(COLOR_INDEX):
                val = COLOR_INDEX[idx]
                if isinstance(val, str):
                    return val
        return None

    # --- Fallback: try .rgb if it's a plain string ---
    rgb = getattr(color_obj, 'rgb', None)
    if isinstance(rgb, str):
        return rgb

    return None


def _get_workbook_defaults(
    wb, theme_colors: List[str]
) -> Dict[str, Any]:
    """Extract the workbook's actual default formatting from the Normal style.

    Reads the Normal named style to determine the workbook's real default font
    (name, size, color) instead of assuming Calibri/11/#000000.

    Returns a dict with keys: font_name, font_size, font_color (hex with '#' or None).
    """
    defaults = {
        "font_name": "Calibri",
        "font_size": 11,
        "font_color": None,  # resolved hex string with '#', or None
    }
    try:
        for ns in wb._named_styles:
            if ns.name == "Normal":
                if ns.font.name:
                    defaults["font_name"] = ns.font.name
                if ns.font.size is not None:
                    defaults["font_size"] = ns.font.size
                if ns.font.color:
                    resolved = _resolve_openpyxl_color(ns.font.color, theme_colors)
                    if resolved:
                        if len(resolved) == 8:  # ARGB
                            resolved = resolved[2:]
                        defaults["font_color"] = f"#{resolved}"
                break
    except Exception:
        pass
    return defaults


def workbook_to_state(
    workbook_path: Union[str, Path],
    sheets: Optional[List[str]] = None
) -> Dict[str, Any]:
    """
    Convert an Excel workbook to a state dictionary.

    Args:
        workbook_path: Path to the Excel workbook
        sheets: Optional list of sheet names to include. If None, includes all sheets.

    Returns:
        State dictionary in the standard format:
        {
            "worksheets": {
                "Sheet1": {
                    "cells": {
                        "A1": {"value": ..., "formula": ..., "Format": {...}},
                        ...
                    },
                    "worksheetProperties": {"merged_cells": [...]}
                }
            }
        }
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border
    from openpyxl.chartsheet import Chartsheet

    logger.debug(f"Loading workbook from {workbook_path}")

    wb = load_workbook(str(workbook_path), data_only=False)
    state: Dict[str, Any] = {"worksheets": {}}

    # Parse theme colors once for the whole workbook
    theme_colors = _parse_theme_colors(wb)

    # Determine actual workbook defaults from the Normal style
    wb_defaults = _get_workbook_defaults(wb, theme_colors)

    sheet_names = sheets if sheets else wb.sheetnames

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in workbook")
            continue

        ws = wb[sheet_name]

        # Skip chart sheets — they have no cells or merged_cells
        if isinstance(ws, Chartsheet):
            logger.debug(f"Skipping chart sheet '{sheet_name}'")
            continue

        sheet_state = _extract_sheet_state(ws, theme_colors, wb_defaults)
        state["worksheets"][sheet_name] = sheet_state

    # Store workbook defaults in state metadata so downstream consumers
    # (e.g. ExcelParser) can use them instead of hardcoded constants.
    state["workbook_defaults"] = wb_defaults

    wb.close()
    logger.debug(f"Loaded {len(state['worksheets'])} sheets")
    return state


def _extract_sheet_state(
    ws,
    theme_colors: Optional[List[str]] = None,
    wb_defaults: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Extract state from a single worksheet."""
    from openpyxl.utils import get_column_letter

    theme_colors = theme_colors or []
    cells: Dict[str, Dict[str, Any]] = {}
    merged_cells: List[Dict[str, int]] = []

    # Extract merged cells
    for merged_range in ws.merged_cells.ranges:
        merged_cells.append({
            "start_row": merged_range.min_row,
            "start_col": merged_range.min_col,
            "end_row": merged_range.max_row,
            "end_col": merged_range.max_col
        })

    # Extract cell data
    for row in ws.iter_rows():
        for cell in row:
            cell_data = _extract_cell_data(cell, theme_colors, wb_defaults)
            if cell_data:
                cell_addr = f"{get_column_letter(cell.column)}{cell.row}"
                cells[cell_addr] = cell_data

    return {
        "cells": cells,
        "worksheetProperties": {"merged_cells": merged_cells}
    }


def _extract_cell_data(
    cell,
    theme_colors: Optional[List[str]] = None,
    wb_defaults: Optional[Dict[str, Any]] = None,
) -> Optional[Dict[str, Any]]:
    """Extract data from a single cell."""
    from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula

    data: Dict[str, Any] = {}

    # Value
    if cell.value is not None:
        value = cell.value
        # Handle openpyxl formula objects (array formulas, data tables)
        # These are not JSON-serializable — extract the formula text instead.
        if isinstance(value, ArrayFormula):
            formula_text = getattr(value, 'text', None)
            if formula_text:
                data["formula"] = formula_text
            # Skip storing the raw object as "value"
        elif isinstance(value, DataTableFormula):
            # DataTableFormula has no useful text representation; skip it
            pass
        # Convert datetime/time/timedelta to strings for JSON serialization
        elif isinstance(value, datetime):
            data["value"] = value.isoformat()
        elif isinstance(value, date):
            data["value"] = value.isoformat()
        elif isinstance(value, time):
            data["value"] = value.isoformat()
        elif isinstance(value, timedelta):
            total_seconds = int(value.total_seconds())
            hours, remainder = divmod(total_seconds, 3600)
            minutes, seconds = divmod(remainder, 60)
            data["value"] = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
        else:
            data["value"] = value

    # Formula
    if cell.data_type == 'f' or (hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith('=')):
        # Check for actual formula
        if hasattr(cell, '_value') and isinstance(cell._value, str) and cell._value.startswith('='):
            data["formula"] = cell._value
        elif isinstance(cell.value, str) and cell.value.startswith('='):
            data["formula"] = cell.value

    # Number format
    if cell.number_format and cell.number_format != 'General':
        data["number_format"] = cell.number_format

    # Format (font, fill, alignment, borders)
    fmt = _extract_format(cell, theme_colors, wb_defaults)
    if fmt:
        data["Format"] = fmt

    return data if data else None


def _extract_format(
    cell,
    theme_colors: Optional[List[str]] = None,
    wb_defaults: Optional[Dict[str, Any]] = None,
) -> Optional[Dict[str, Any]]:
    """Extract formatting information from a cell."""
    theme_colors = theme_colors or []
    fmt: Dict[str, Any] = {}

    # Font
    font_data = _extract_font(cell.font, theme_colors, wb_defaults)
    if font_data:
        fmt["font"] = font_data

    # Fill
    fill_data = _extract_fill(cell.fill, theme_colors)
    if fill_data:
        fmt["fill"] = fill_data

    # Alignment
    alignment_data = _extract_alignment(cell.alignment)
    if alignment_data:
        for key, value in alignment_data.items():
            fmt[key] = value

    # Borders
    border_data = _extract_borders(cell.border, theme_colors)
    if border_data:
        fmt["borders"] = border_data

    return fmt if fmt else None


def _extract_font(
    font,
    theme_colors: Optional[List[str]] = None,
    wb_defaults: Optional[Dict[str, Any]] = None,
) -> Optional[Dict[str, Any]]:
    """Extract font properties.

    Compares against the workbook's actual Normal-style defaults (font name,
    size, color) rather than hardcoded constants.  Falls back to Calibri / 11
    when *wb_defaults* is not provided.
    """
    if font is None:
        return None

    theme_colors = theme_colors or []
    wb_defaults = wb_defaults or {}
    data: Dict[str, Any] = {}

    default_font_name = wb_defaults.get("font_name", "Calibri")
    default_font_size = wb_defaults.get("font_size", 11)
    default_font_color = wb_defaults.get("font_color")  # e.g. "#000000" or None

    # Only include non-default values
    if font.name and font.name != default_font_name:
        data["name"] = font.name
    if font.size is not None and font.size != default_font_size:
        data["size"] = font.size
    if font.bold:
        data["bold"] = True
    if font.italic:
        data["italic"] = True
    if font.underline and font.underline != 'none':
        data["underline"] = font.underline
    if font.color:
        color = _resolve_openpyxl_color(font.color, theme_colors)
        if color and color != '00000000':
            if len(color) == 8:  # ARGB format
                color = color[2:]  # Remove alpha
            hex_color = f"#{color}"
            # Skip if this matches the workbook's default font color
            # or the universal default (#000000)
            skip = hex_color.upper() == "#000000"
            if not skip and default_font_color:
                skip = hex_color.upper() == default_font_color.upper()
            if not skip:
                data["color"] = hex_color

    return data if data else None


def _extract_fill(fill, theme_colors: Optional[List[str]] = None) -> Optional[Dict[str, Any]]:
    """Extract fill properties."""
    if fill is None or fill.fill_type is None:
        return None

    theme_colors = theme_colors or []
    data: Dict[str, Any] = {}

    if fill.fill_type == 'solid' and fill.fgColor:
        color = _resolve_openpyxl_color(fill.fgColor, theme_colors)
        if color and color != '00000000':
            if len(color) == 8:
                color = color[2:]
            data["fgColor"] = f"#{color}"
            data["patternType"] = "solid"

    return data if data else None


def _extract_alignment(alignment) -> Optional[Dict[str, Any]]:
    """Extract alignment properties."""
    if alignment is None:
        return None

    data: Dict[str, Any] = {}

    if alignment.horizontal and alignment.horizontal != 'general':
        data["horizontalAlignment"] = alignment.horizontal.capitalize()
    if alignment.vertical and alignment.vertical != 'bottom':
        data["verticalAlignment"] = alignment.vertical.capitalize()
    if alignment.wrap_text:
        data["wrapText"] = True
    if alignment.text_rotation and alignment.text_rotation != 0:
        data["textOrientation"] = alignment.text_rotation

    return data if data else None


def _extract_borders(border, theme_colors: Optional[List[str]] = None) -> Optional[Dict[str, Any]]:
    """Extract border properties."""
    if border is None:
        return None

    theme_colors = theme_colors or []
    data: Dict[str, Any] = {}

    for side_name in ['left', 'right', 'top', 'bottom']:
        side = getattr(border, side_name, None)
        if side and side.style and side.style != 'none':
            side_data: Dict[str, Any] = {"lineStyle": side.style}
            color = None
            if side.color:
                color = _resolve_openpyxl_color(side.color, theme_colors)
                if color:
                    if len(color) == 8:
                        color = color[2:]
                    color = f"#{color}"
            # Default to black when color is unresolvable (e.g. theme-based)
            side_data["color"] = color or "#000000"
            data[side_name] = side_data

    return data if data else None
