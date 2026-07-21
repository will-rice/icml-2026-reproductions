"""
State to Sheet Conversion.

Convert state dictionaries to openpyxl workbooks.
"""

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

logger = logging.getLogger(__name__)


def state_to_workbook(
    state: Dict[str, Any],
    output_path: Optional[Union[str, Path]] = None
):
    """
    Convert a state dictionary to an openpyxl workbook.

    Args:
        state: State dictionary in the standard format
        output_path: Optional path to save the workbook. If None, returns the workbook object.

    Returns:
        openpyxl Workbook object (if output_path is None)
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    logger.debug("Creating workbook from state")

    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames and len(state.get("worksheets", {})) > 0:
        del wb['Sheet']

    worksheets = state.get("worksheets", {})

    for sheet_name, sheet_data in worksheets.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)

        _apply_sheet_state(ws, sheet_data)

    if output_path:
        wb.save(str(output_path))
        logger.debug(f"Saved workbook to {output_path}")
        wb.close()
        return None
    else:
        return wb


def _apply_sheet_state(ws, sheet_data: Dict[str, Any]) -> None:
    """Apply state data to a worksheet."""
    from openpyxl.utils import get_column_letter, column_index_from_string

    cells = sheet_data.get("cells", {})
    properties = sheet_data.get("worksheetProperties", {})

    # Apply cell data
    for cell_addr, cell_data in cells.items():
        _apply_cell_data(ws, cell_addr, cell_data)

    # Apply merged cells
    merged_cells = properties.get("merged_cells", [])
    for merge_info in merged_cells:
        start_row = merge_info["start_row"]
        start_col = merge_info["start_col"]
        end_row = merge_info["end_row"]
        end_col = merge_info["end_col"]

        start_cell = f"{get_column_letter(start_col)}{start_row}"
        end_cell = f"{get_column_letter(end_col)}{end_row}"
        merge_range = f"{start_cell}:{end_cell}"

        try:
            ws.merge_cells(merge_range)
        except Exception as e:
            logger.warning(f"Failed to merge cells {merge_range}: {e}")


def _apply_cell_data(ws, cell_addr: str, cell_data: Dict[str, Any]) -> None:
    """Apply data to a single cell."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    cell = ws[cell_addr]

    # Value (prefer formula if both exist)
    if "formula" in cell_data:
        cell.value = cell_data["formula"]
    elif "value" in cell_data:
        # Preserve date values as-is
        cell.value = cell_data["value"]

    # Number format
    if "number_format" in cell_data:
        cell.number_format = cell_data["number_format"]

    # Format
    fmt = cell_data.get("Format", {})
    if fmt:
        _apply_format(cell, fmt)


def _apply_format(cell, fmt: Dict[str, Any]) -> None:
    """Apply formatting to a cell."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    # Font
    font_data = fmt.get("font", {})
    if font_data:
        cell.font = _create_font(font_data)

    # Fill
    fill_data = fmt.get("fill", {})
    if fill_data:
        cell.fill = _create_fill(fill_data)

    # Alignment
    # Map alignment values to openpyxl-expected casing
    _h_align_map = {
        "centercontinuous": "centerContinuous",
        "centeracrossselection": "centerContinuous",
    }
    _v_align_map = {}
    alignment_kwargs = {}
    if "horizontalAlignment" in fmt:
        val = fmt["horizontalAlignment"].lower()
        alignment_kwargs["horizontal"] = _h_align_map.get(val, val)
    if "verticalAlignment" in fmt:
        val = fmt["verticalAlignment"].lower()
        alignment_kwargs["vertical"] = _v_align_map.get(val, val)
    if "wrapText" in fmt:
        alignment_kwargs["wrap_text"] = fmt["wrapText"]
    if "textOrientation" in fmt:
        alignment_kwargs["text_rotation"] = fmt["textOrientation"]

    if alignment_kwargs:
        cell.alignment = Alignment(**alignment_kwargs)

    # Borders
    border_data = fmt.get("borders", {})
    if border_data:
        cell.border = _create_border(border_data)


def _create_font(font_data: Dict[str, Any]):
    """Create an openpyxl Font from font data."""
    from openpyxl.styles import Font
    from openpyxl.styles.colors import Color

    kwargs = {}

    if "name" in font_data:
        kwargs["name"] = font_data["name"]
    if "size" in font_data:
        kwargs["size"] = font_data["size"]
    if "bold" in font_data:
        kwargs["bold"] = font_data["bold"]
    if "italic" in font_data:
        kwargs["italic"] = font_data["italic"]
    if "underline" in font_data:
        underline = font_data["underline"]
        if underline in [True, 'single']:
            kwargs["underline"] = 'single'
        elif underline == 'double':
            kwargs["underline"] = 'double'
        elif underline in [False, 'none', None]:
            pass
        else:
            kwargs["underline"] = underline
    if "color" in font_data:
        color = font_data["color"]
        if color:
            if color.startswith("#"):
                color = color[1:]
            kwargs["color"] = color

    return Font(**kwargs) if kwargs else Font()


def _create_fill(fill_data: Dict[str, Any]):
    """Create an openpyxl PatternFill from fill data."""
    from openpyxl.styles import PatternFill

    if not fill_data:
        return PatternFill()

    fg_color = fill_data.get("fgColor", "")
    pattern_type = fill_data.get("patternType", "solid")

    if fg_color:
        if fg_color.startswith("#"):
            fg_color = fg_color[1:]
        return PatternFill(
            start_color=fg_color,
            end_color=fg_color,
            fill_type=pattern_type
        )

    return PatternFill()


def _create_border(border_data: Dict[str, Any]):
    """Create an openpyxl Border from border data."""
    from openpyxl.styles import Border, Side

    kwargs = {}

    for side_name in ['left', 'right', 'top', 'bottom']:
        side_data = border_data.get(side_name)
        if side_data:
            kwargs[side_name] = _create_side(side_data)

    return Border(**kwargs) if kwargs else Border()


def _create_side(side_data: Dict[str, Any]):
    """Create an openpyxl Side from side data."""
    from openpyxl.styles import Side

    style = side_data.get("lineStyle", "thin")
    color = side_data.get("color", "000000")

    if color and color.startswith("#"):
        color = color[1:]

    return Side(style=style, color=color)
