"""
InputSplitter — post-processes multi-cell SetInput operations from OperationMerger,
splitting oversized 2D INPUTs into smaller, semantically meaningful blocks.

Pipeline position: after OperationMerger, before OperationSequencer.

Steps applied to each multi-cell SetInput:
1. Empty row split — split on fully-empty rows
2. Empty column split — split on fully-empty columns
3. Trim to bounding box (remove outer all-empty rows/cols)
4. Edge peel — peel edge rows/cols where trailing consecutive empty cells
   from the end >= half the row/col length
5. Formula column extraction — extract columns containing any '=' values
   as individual single-cell ops
6. Small/sparse decomposition — blocks with cells <= split_threshold decompose
   to individual ops; density < 0.5 also decomposes
"""

import logging
from typing import Any, Dict, List, Tuple

from openpyxl.utils import get_column_letter

from next_action_pred_eval.core.cell_range import CellRange
from next_action_pred_eval.core.operations.value_ops import SetInput
from next_action_pred_eval.generation.sequencing.base import BaseTransformer, SequencingContext

logger = logging.getLogger(__name__)


class InputSplitter(BaseTransformer):
    """
    Splits oversized multi-cell SetInput operations into smaller blocks.

    Config:
        enabled: bool (default True)
        split_threshold: int (default 4) — blocks with this many cells or fewer
            are decomposed into individual single-cell ops
        sparse_density: float (default 0.5) — blocks with density below this
            threshold are decomposed
    """

    DEFAULT_CONFIG = {
        "enabled": True,
        "split_threshold": 4,
        "sparse_density": 0.5,
    }

    def transform(self, context: SequencingContext) -> SequencingContext:
        if self.can_skip(context):
            return context

        split_threshold = self.config.get("split_threshold", 4)
        sparse_density = self.config.get("sparse_density", 0.5)

        result = []
        total_split = 0

        for op in context.operations:
            if not isinstance(op, SetInput) or op.is_inverse:
                result.append(op)
                continue

            # Only split multi-cell 2D INPUTs
            rows, cols = op.cell_range.get_dimensions()
            if rows <= 1 and cols <= 1:
                result.append(op)
                continue

            if not isinstance(op.value, list) or not all(isinstance(r, list) for r in op.value):
                result.append(op)
                continue

            r1, c1, _, _ = op.cell_range.get_coordinates()
            sheet = op.cell_range.sheet
            grid = op.value  # 2D list: grid[row_offset][col_offset]

            sub_blocks = self._split_input(grid, r1, c1, sheet, split_threshold, sparse_density)

            if len(sub_blocks) == 1 and sub_blocks[0] == (r1, c1, rows, cols, grid):
                # No splitting happened
                result.append(op)
            else:
                total_split += 1
                for block_r, block_c, block_rows, block_cols, block_grid in sub_blocks:
                    result.extend(
                        _grid_to_ops(block_r, block_c, block_rows, block_cols, block_grid, sheet)
                    )

        if total_split > 0:
            self.log(context, f"Split {total_split} multi-cell INPUTs")

        return context.copy_with_operations(result)

    def _split_input(
        self,
        grid: List[List],
        abs_row: int,
        abs_col: int,
        sheet: str,
        split_threshold: int,
        sparse_density: float,
    ) -> List[Tuple[int, int, int, int, List[List]]]:
        """
        Apply the splitting heuristics to a 2D grid.

        Returns list of (abs_row, abs_col, rows, cols, sub_grid) tuples.
        """
        blocks = [(abs_row, abs_col, len(grid), len(grid[0]), grid)]

        # Step 1: Empty row split
        blocks = _apply_to_blocks(blocks, _split_on_empty_rows)

        # Step 2: Empty column split
        blocks = _apply_to_blocks(blocks, _split_on_empty_cols)

        # Step 3: Trim to bounding box
        blocks = _apply_to_blocks(blocks, _trim_to_bounding_box)

        # Step 4: Edge peel
        blocks = _apply_to_blocks(blocks, _edge_peel)

        # Step 5: Formula column extraction
        blocks = _apply_to_blocks(blocks, _extract_formula_columns)

        # Step 6: Small/sparse decomposition
        blocks = _apply_to_blocks(
            blocks,
            lambda r, c, rows, cols, g: _decompose_if_small_or_sparse(
                r, c, rows, cols, g, split_threshold, sparse_density
            ),
        )

        return blocks


# ============================================================================
# Block manipulation helpers
# ============================================================================

Block = Tuple[int, int, int, int, List[List]]  # (abs_row, abs_col, rows, cols, grid)


def _apply_to_blocks(blocks: List[Block], fn) -> List[Block]:
    """Apply a splitting function to each block, collecting results."""
    result = []
    for r, c, rows, cols, grid in blocks:
        result.extend(fn(r, c, rows, cols, grid))
    return result


def _is_empty(val) -> bool:
    """Check if a cell value is empty (None or empty string)."""
    return val is None or val == "" or val == []


def _split_on_empty_rows(
    abs_row: int, abs_col: int, rows: int, cols: int, grid: List[List]
) -> List[Block]:
    """Split on fully-empty rows."""
    result = []
    block_start = None

    for r in range(rows):
        row_empty = all(_is_empty(grid[r][c]) for c in range(cols))
        if row_empty:
            if block_start is not None:
                sub = [grid[i] for i in range(block_start, r)]
                result.append((abs_row + block_start, abs_col, r - block_start, cols, sub))
                block_start = None
        else:
            if block_start is None:
                block_start = r

    if block_start is not None:
        sub = [grid[i] for i in range(block_start, rows)]
        result.append((abs_row + block_start, abs_col, rows - block_start, cols, sub))

    return result if result else [(abs_row, abs_col, rows, cols, grid)]


def _split_on_empty_cols(
    abs_row: int, abs_col: int, rows: int, cols: int, grid: List[List]
) -> List[Block]:
    """Split on fully-empty columns."""
    result = []
    block_start = None

    for c in range(cols):
        col_empty = all(_is_empty(grid[r][c]) for r in range(rows))
        if col_empty:
            if block_start is not None:
                width = c - block_start
                sub = [[grid[r][cc] for cc in range(block_start, c)] for r in range(rows)]
                result.append((abs_row, abs_col + block_start, rows, width, sub))
                block_start = None
        else:
            if block_start is None:
                block_start = c

    if block_start is not None:
        width = cols - block_start
        sub = [[grid[r][cc] for cc in range(block_start, cols)] for r in range(rows)]
        result.append((abs_row, abs_col + block_start, rows, width, sub))

    return result if result else [(abs_row, abs_col, rows, cols, grid)]


def _trim_to_bounding_box(
    abs_row: int, abs_col: int, rows: int, cols: int, grid: List[List]
) -> List[Block]:
    """Trim all-empty outer rows and columns."""
    # Find top
    top = 0
    while top < rows and all(_is_empty(grid[top][c]) for c in range(cols)):
        top += 1
    # Find bottom
    bot = rows
    while bot > top and all(_is_empty(grid[bot - 1][c]) for c in range(cols)):
        bot -= 1
    # Find left
    left = 0
    while left < cols and all(_is_empty(grid[r][left]) for r in range(top, bot)):
        left += 1
    # Find right
    right = cols
    while right > left and all(_is_empty(grid[r][right - 1]) for r in range(top, bot)):
        right -= 1

    if top >= bot or left >= right:
        return []  # entirely empty — drop

    if top == 0 and bot == rows and left == 0 and right == cols:
        return [(abs_row, abs_col, rows, cols, grid)]  # no trimming needed

    sub = [[grid[r][c] for c in range(left, right)] for r in range(top, bot)]
    return [(abs_row + top, abs_col + left, bot - top, right - left, sub)]


def _edge_peel(
    abs_row: int, abs_col: int, rows: int, cols: int, grid: List[List]
) -> List[Block]:
    """
    Peel edge rows/columns where trailing consecutive empty cells from the end
    are >= half the row/col length. Peeled edges become individual single-cell ops.

    For top/bottom rows: count consecutive empty cells from the right end.
    For left/right cols: count consecutive empty cells from the bottom end.
    """
    peeled = []
    top = 0
    bot = rows
    left = 0
    right = cols

    # Peel from top
    while top < bot - 1:  # keep at least 1 row
        trailing = _trailing_empties_in_row(grid[top], left, right)
        row_width = right - left
        if trailing >= row_width / 2:
            # Peel this row — emit non-empty cells as individual ops
            for c in range(left, right):
                if not _is_empty(grid[top][c]):
                    peeled.append((
                        abs_row + top, abs_col + c, 1, 1,
                        [[grid[top][c]]]
                    ))
            top += 1
        else:
            break

    # Peel from bottom
    while bot > top + 1:  # keep at least 1 row
        trailing = _trailing_empties_in_row(grid[bot - 1], left, right)
        row_width = right - left
        if trailing >= row_width / 2:
            for c in range(left, right):
                if not _is_empty(grid[bot - 1][c]):
                    peeled.append((
                        abs_row + bot - 1, abs_col + c, 1, 1,
                        [[grid[bot - 1][c]]]
                    ))
            bot -= 1
        else:
            break

    # Peel from left
    while left < right - 1:  # keep at least 1 col
        trailing = _trailing_empties_in_col(grid, left, top, bot)
        col_height = bot - top
        if trailing >= col_height / 2:
            for r in range(top, bot):
                if not _is_empty(grid[r][left]):
                    peeled.append((
                        abs_row + r, abs_col + left, 1, 1,
                        [[grid[r][left]]]
                    ))
            left += 1
        else:
            break

    # Peel from right
    while right > left + 1:  # keep at least 1 col
        trailing = _trailing_empties_in_col(grid, right - 1, top, bot)
        col_height = bot - top
        if trailing >= col_height / 2:
            for r in range(top, bot):
                if not _is_empty(grid[r][right - 1]):
                    peeled.append((
                        abs_row + r, abs_col + right - 1, 1, 1,
                        [[grid[r][right - 1]]]
                    ))
            right -= 1
        else:
            break

    if top == 0 and bot == rows and left == 0 and right == cols:
        return [(abs_row, abs_col, rows, cols, grid)]  # nothing peeled

    # Build remaining core block
    core_rows = bot - top
    core_cols = right - left
    if core_rows > 0 and core_cols > 0:
        sub = [[grid[r][c] for c in range(left, right)] for r in range(top, bot)]
        peeled.append((abs_row + top, abs_col + left, core_rows, core_cols, sub))

    return peeled if peeled else [(abs_row, abs_col, rows, cols, grid)]


def _trailing_empties_in_row(row: List, start: int, end: int) -> int:
    """Count consecutive empty cells from the right end of a row slice."""
    count = 0
    for c in range(end - 1, start - 1, -1):
        if _is_empty(row[c]):
            count += 1
        else:
            break
    return count


def _trailing_empties_in_col(grid: List[List], col: int, top: int, bot: int) -> int:
    """Count consecutive empty cells from the bottom end of a column slice."""
    count = 0
    for r in range(bot - 1, top - 1, -1):
        if _is_empty(grid[r][col]):
            count += 1
        else:
            break
    return count


def _extract_formula_columns(
    abs_row: int, abs_col: int, rows: int, cols: int, grid: List[List]
) -> List[Block]:
    """
    Extract columns where the majority of non-empty cells are formulas ('=' prefix).
    These become individual single-cell ops; remaining columns stay as blocks.
    """
    if cols <= 1:
        return [(abs_row, abs_col, rows, cols, grid)]

    formula_cols = set()
    for c in range(cols):
        formula_count = 0
        non_empty_count = 0
        for r in range(rows):
            val = grid[r][c]
            if not _is_empty(val):
                non_empty_count += 1
                if isinstance(val, str) and val.startswith("="):
                    formula_count += 1
        # Extract column if majority of non-empty cells are formulas
        if non_empty_count > 0 and formula_count > non_empty_count / 2:
            formula_cols.add(c)

    if not formula_cols:
        return [(abs_row, abs_col, rows, cols, grid)]

    result = []
    # Emit formula columns as individual single-cell ops
    for c in sorted(formula_cols):
        for r in range(rows):
            if not _is_empty(grid[r][c]):
                result.append((abs_row + r, abs_col + c, 1, 1, [[grid[r][c]]]))

    # Build contiguous blocks from remaining columns
    non_formula = [c for c in range(cols) if c not in formula_cols]
    if non_formula:
        # Group into contiguous runs
        runs = []
        run_start = non_formula[0]
        prev = non_formula[0]
        for c in non_formula[1:]:
            if c == prev + 1:
                prev = c
            else:
                runs.append((run_start, prev + 1))
                run_start = c
                prev = c
        runs.append((run_start, prev + 1))

        for run_s, run_e in runs:
            width = run_e - run_s
            sub = [[grid[r][c] for c in range(run_s, run_e)] for r in range(rows)]
            result.append((abs_row, abs_col + run_s, rows, width, sub))

    return result


def _decompose_if_small_or_sparse(
    abs_row: int,
    abs_col: int,
    rows: int,
    cols: int,
    grid: List[List],
    split_threshold: int,
    sparse_density: float,
) -> List[Block]:
    """Decompose small blocks (cells <= threshold) or sparse blocks (density < threshold)."""
    total = rows * cols
    if total <= 1:
        return [(abs_row, abs_col, rows, cols, grid)]

    non_empty = sum(1 for r in range(rows) for c in range(cols) if not _is_empty(grid[r][c]))
    density = non_empty / total if total > 0 else 0

    if total <= split_threshold or density < sparse_density:
        result = []
        for r in range(rows):
            for c in range(cols):
                if not _is_empty(grid[r][c]):
                    result.append((abs_row + r, abs_col + c, 1, 1, [[grid[r][c]]]))
        return result if result else []

    return [(abs_row, abs_col, rows, cols, grid)]


def _grid_to_ops(
    abs_row: int,
    abs_col: int,
    rows: int,
    cols: int,
    grid: List[List],
    sheet: str,
) -> List[SetInput]:
    """Convert a block back into SetInput operation(s)."""
    if rows == 1 and cols == 1:
        # Single cell
        val = grid[0][0]
        if _is_empty(val):
            return []
        range_str = f"{get_column_letter(abs_col)}{abs_row}"
        return [SetInput(
            cell_range=CellRange(sheet=sheet, range=range_str),
            value=val,
        )]

    # Multi-cell — check if all empty
    has_content = any(not _is_empty(grid[r][c]) for r in range(rows) for c in range(cols))
    if not has_content:
        return []

    start = f"{get_column_letter(abs_col)}{abs_row}"
    end = f"{get_column_letter(abs_col + cols - 1)}{abs_row + rows - 1}"
    range_str = f"{start}:{end}"
    return [SetInput(
        cell_range=CellRange(sheet=sheet, range=range_str),
        value=grid,
    )]
