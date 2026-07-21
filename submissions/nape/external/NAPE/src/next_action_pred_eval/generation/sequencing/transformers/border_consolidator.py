"""
Border Consolidator - Consolidates individual border operations into efficient patterns

Handles:
- Detecting BORDER_ALL patterns (all 4 sides with same border)
- Detecting BORDER_OUTSIDE patterns (only outer edges of a range)
- Handling merged cells (skip interior borders for merged regions)
"""

from typing import List, Dict, Any, Optional, Set, Tuple
from collections import defaultdict

from next_action_pred_eval.generation.sequencing.base import BaseTransformer, SequencingContext
from next_action_pred_eval.core.operation import Operation
from next_action_pred_eval.core.cell_range import CellRange


def get_border_key(border_value: Dict) -> str:
    """Create a unique key for border properties."""
    weight = border_value.get('weight') or ''
    style = border_value.get('style') or ''
    color = border_value.get('color') or ''
    return f"{weight}|{style}|{color}"


def cells_form_rectangle(cells: Set[Tuple[int, int]]) -> Tuple[bool, Tuple[int, int, int, int]]:
    """
    Check if a set of cells forms a contiguous rectangle.

    Args:
        cells: Set of (row, col) tuples

    Returns:
        Tuple of (is_rectangle, (min_row, min_col, max_row, max_col))
    """
    if not cells:
        return False, (0, 0, 0, 0)

    rows = {c[0] for c in cells}
    cols = {c[1] for c in cells}

    min_row, max_row = min(rows), max(rows)
    min_col, max_col = min(cols), max(cols)

    # Check if all cells in the bounding box are present
    expected_cells = (max_row - min_row + 1) * (max_col - min_col + 1)
    if len(cells) != expected_cells:
        return False, (min_row, min_col, max_row, max_col)

    return True, (min_row, min_col, max_row, max_col)


def _merge_horizontal_strips(
    cells: Set[Tuple[int, int]],
    sheet: str,
    side: str,
    border_value: Dict,
) -> List:
    """Merge cells into horizontal strips (same row, contiguous columns).

    Used for BORDER_TOP and BORDER_BOTTOM, which can only merge horizontally
    because applying these to a range only affects the edge row.
    """
    from next_action_pred_eval.core.operations import SetBorder
    from openpyxl.utils import get_column_letter

    # Group by row
    by_row: Dict[int, List[int]] = defaultdict(list)
    for row, col in cells:
        by_row[row].append(col)

    ops = []
    for row in sorted(by_row):
        cols = sorted(by_row[row])
        # Find contiguous strips
        start = cols[0]
        end = cols[0]
        for c in cols[1:]:
            if c == end + 1:
                end = c
            else:
                # Emit strip
                if start == end:
                    range_str = f"{get_column_letter(start)}{row}"
                else:
                    range_str = f"{get_column_letter(start)}{row}:{get_column_letter(end)}{row}"
                ops.append(SetBorder(
                    cell_range=CellRange(sheet=sheet, range=range_str),
                    side=side,
                    value=border_value,
                ))
                start = c
                end = c
        # Emit last strip
        if start == end:
            range_str = f"{get_column_letter(start)}{row}"
        else:
            range_str = f"{get_column_letter(start)}{row}:{get_column_letter(end)}{row}"
        ops.append(SetBorder(
            cell_range=CellRange(sheet=sheet, range=range_str),
            side=side,
            value=border_value,
        ))
    return ops


def _merge_vertical_strips(
    cells: Set[Tuple[int, int]],
    sheet: str,
    side: str,
    border_value: Dict,
) -> List:
    """Merge cells into vertical strips (same column, contiguous rows).

    Used for BORDER_LEFT and BORDER_RIGHT, which can only merge vertically
    because applying these to a range only affects the edge column.
    """
    from next_action_pred_eval.core.operations import SetBorder
    from openpyxl.utils import get_column_letter

    # Group by column
    by_col: Dict[int, List[int]] = defaultdict(list)
    for row, col in cells:
        by_col[col].append(row)

    ops = []
    for col in sorted(by_col):
        rows = sorted(by_col[col])
        # Find contiguous strips
        start = rows[0]
        end = rows[0]
        for r in rows[1:]:
            if r == end + 1:
                end = r
            else:
                # Emit strip
                if start == end:
                    range_str = f"{get_column_letter(col)}{start}"
                else:
                    range_str = f"{get_column_letter(col)}{start}:{get_column_letter(col)}{end}"
                ops.append(SetBorder(
                    cell_range=CellRange(sheet=sheet, range=range_str),
                    side=side,
                    value=border_value,
                ))
                start = r
                end = r
        # Emit last strip
        if start == end:
            range_str = f"{get_column_letter(col)}{start}"
        else:
            range_str = f"{get_column_letter(col)}{start}:{get_column_letter(col)}{end}"
        ops.append(SetBorder(
            cell_range=CellRange(sheet=sheet, range=range_str),
            side=side,
            value=border_value,
        ))
    return ops


def is_cell_in_merged_interior(
    row: int,
    col: int,
    merged_ranges: List[Tuple[int, int, int, int]]
) -> bool:
    """
    Check if a cell is in the interior of a merged range (not the anchor).

    Args:
        row: Cell row
        col: Cell column
        merged_ranges: List of (min_row, min_col, max_row, max_col) tuples

    Returns:
        True if cell is in a merged range but not the anchor
    """
    for min_row, min_col, max_row, max_col in merged_ranges:
        if min_row <= row <= max_row and min_col <= col <= max_col:
            # Check if it's the anchor (top-left cell)
            if row == min_row and col == min_col:
                return False  # It's the anchor, not interior
            return True  # It's in the merged range but not the anchor
    return False


def get_merged_range_for_cell(
    row: int,
    col: int,
    merged_ranges: List[Tuple[int, int, int, int]]
) -> Optional[Tuple[int, int, int, int]]:
    """
    Get the merged range that contains a cell, if any.

    Args:
        row: Cell row
        col: Cell column
        merged_ranges: List of (min_row, min_col, max_row, max_col) tuples

    Returns:
        The merged range tuple if cell is in a merged range, else None
    """
    for merge in merged_ranges:
        min_row, min_col, max_row, max_col = merge
        if min_row <= row <= max_row and min_col <= col <= max_col:
            return merge
    return None


def get_expected_borders_for_cell(
    row: int,
    col: int,
    range_bounds: Tuple[int, int, int, int],
    merged_ranges: List[Tuple[int, int, int, int]]
) -> Set[str]:
    """
    Get the expected border sides for a cell in a BORDER_ALL range, accounting for merged cells.

    When BORDER_ALL is applied to a range and then cells are merged:
    - Non-merged cells retain all 4 borders
    - Merged cells only retain their outer perimeter borders

    Args:
        row: Cell row
        col: Cell column
        range_bounds: (min_row, min_col, max_row, max_col) of the BORDER_ALL range
        merged_ranges: List of merged range tuples

    Returns:
        Set of expected border sides ('left', 'right', 'top', 'bottom')
    """
    range_min_row, range_min_col, range_max_row, range_max_col = range_bounds

    # Check if cell is in a merged range
    merge = get_merged_range_for_cell(row, col, merged_ranges)

    if merge is None:
        # Non-merged cell: expect all 4 borders
        return {'left', 'right', 'top', 'bottom'}

    merge_min_row, merge_min_col, merge_max_row, merge_max_col = merge

    # Cell is in a merged range - only expect borders on the merged range's perimeter
    expected = set()

    # Left border: only if cell is in leftmost column of merged range
    if col == merge_min_col:
        expected.add('left')

    # Right border: only if cell is in rightmost column of merged range
    if col == merge_max_col:
        expected.add('right')

    # Top border: only if cell is in topmost row of merged range
    if row == merge_min_row:
        expected.add('top')

    # Bottom border: only if cell is in bottommost row of merged range
    if row == merge_max_row:
        expected.add('bottom')

    return expected


def check_border_all_with_merges(
    sides_data: Dict[str, Set[Tuple[int, int]]],
    range_bounds: Tuple[int, int, int, int],
    merged_ranges: List[Tuple[int, int, int, int]]
) -> bool:
    """
    Check if border data matches BORDER_ALL pattern for a range, accounting for merged cells.

    Args:
        sides_data: Dict mapping side name to set of (row, col) cells with that border
        range_bounds: (min_row, min_col, max_row, max_col) to check
        merged_ranges: List of merged range tuples

    Returns:
        True if the border pattern matches BORDER_ALL with merges
    """
    min_row, min_col, max_row, max_col = range_bounds

    # Check each cell in the range
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            expected_sides = get_expected_borders_for_cell(row, col, range_bounds, merged_ranges)

            # Verify each expected border is present
            for side in expected_sides:
                if (row, col) not in sides_data.get(side, set()):
                    return False

    return True


class BorderConsolidator(BaseTransformer):
    """
    Consolidates individual border operations into more efficient patterns.

    This transformer detects patterns like:
    1. BORDER_ALL: When all 4 sides of cells have the same border
    2. BORDER_OUTSIDE: When only the outer edges of a range have borders

    It also handles merged cells by treating them as single units.

    Config:
        enabled: bool - Whether consolidation is active
        detect_border_all: bool - Detect and create BORDER_ALL patterns
        detect_border_outside: bool - Detect and create BORDER_OUTSIDE patterns
        respect_merged_cells: bool - Skip interior borders for merged cells
    """

    DEFAULT_CONFIG = {
        "enabled": True,
        "detect_border_all": True,
        "detect_border_outside": True,
        "respect_merged_cells": True,
        "noise_tolerance": True,
    }

    def transform(self, context: SequencingContext) -> SequencingContext:
        if self.can_skip(context):
            return context

        from next_action_pred_eval.core.operations import SetBorder, MergeCells
        from openpyxl.utils import range_boundaries, get_column_letter

        # Get merged ranges from operations
        merged_ranges = []
        if self.config.get("respect_merged_cells", True):
            for op in context.operations:
                if isinstance(op, MergeCells) and not op.is_inverse:
                    try:
                        min_col, min_row, max_col, max_row = range_boundaries(op.cell_range.range)
                        merged_ranges.append((min_row, min_col, max_row, max_col))
                    except:
                        continue

        # Separate border and non-border operations
        border_ops = [op for op in context.operations if isinstance(op, SetBorder) and not op.is_inverse]
        non_border_ops = [op for op in context.operations if not isinstance(op, SetBorder) or op.is_inverse]

        if not border_ops:
            return context

        # Composite-side ops (all, outside, insideHorizontal, insideVertical)
        # are already well-consolidated by the upstream BorderMerger.
        # Pass them through directly — decomposing and re-consolidating
        # is lossy and can dramatically inflate the op count.
        COMPOSITE_SIDES = {'all', 'outside', 'insideHorizontal', 'insideVertical'}
        passthrough_ops = [op for op in border_ops if op.side in COMPOSITE_SIDES]
        individual_border_ops = [op for op in border_ops if op.side not in COMPOSITE_SIDES]

        if not individual_border_ops:
            # Nothing to consolidate — all ops are already composite
            result_ops = non_border_ops + passthrough_ops
            self.log(
                context,
                f"Border consolidation: {len(border_ops)} → {len(passthrough_ops)} border ops (all composite, passthrough)"
            )
            return context.copy_with_operations(result_ops)

        # Group individual-side border operations by sheet and border properties
        # Key: (sheet, border_key) -> side -> set of (row, col) cells
        grouped: Dict[Tuple[str, str], Dict[str, Set[Tuple[int, int]]]] = defaultdict(
            lambda: defaultdict(set)
        )

        for op in individual_border_ops:
            border_key = get_border_key(op.value)
            sheet = op.cell_range.sheet

            try:
                min_col, min_row, max_col, max_row = range_boundaries(op.cell_range.range)

                # Only individual sides (left/right/top/bottom) reach here;
                # composite sides (all/outside/inside*) are in passthrough_ops.
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        if op.side == 'left':
                            if col == min_col:
                                grouped[(sheet, border_key)]['left'].add((row, col))
                        elif op.side == 'right':
                            if col == max_col:
                                grouped[(sheet, border_key)]['right'].add((row, col))
                        elif op.side == 'top':
                            if row == min_row:
                                grouped[(sheet, border_key)]['top'].add((row, col))
                        elif op.side == 'bottom':
                            if row == max_row:
                                grouped[(sheet, border_key)]['bottom'].add((row, col))
            except:
                continue

        # Consolidate border patterns
        consolidated_ops = []
        processed_cells: Dict[Tuple[str, str], Dict[str, Set[Tuple[int, int]]]] = defaultdict(
            lambda: defaultdict(set)
        )

        for (sheet, border_key), sides_data in grouped.items():
            # Parse border properties from key
            weight, style, color = border_key.split('|')
            border_value = {
                'weight': weight or None,
                'style': style or None,
                'color': color or None
            }

            # Check for BORDER_ALL pattern: same cells have all 4 sides (accounting for merged cells)
            if self.config.get("detect_border_all", True):
                all_sides = {'left', 'right', 'top', 'bottom'}
                noise_tolerance = self.config.get("noise_tolerance", True)
                min_sides = 3 if noise_tolerance else 4

                # Count sides per cell for noise-tolerant detection
                cell_side_count: Dict[Tuple[int, int], int] = defaultdict(int)
                for side in all_sides:
                    for cell in sides_data.get(side, set()):
                        cell_side_count[cell] += 1

                # Cells with enough sides for BORDER_ALL
                candidate_cells = {c for c, cnt in cell_side_count.items() if cnt >= min_sides}

                if candidate_cells and len(candidate_cells) > 1:
                    is_rect, bounds = cells_form_rectangle(candidate_cells)
                    if is_rect:
                        min_row, min_col, max_row, max_col = bounds
                        range_str = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

                        all_op = SetBorder(
                            cell_range=CellRange(sheet=sheet, range=range_str),
                            side='all',
                            value=border_value
                        )
                        consolidated_ops.append(all_op)

                        for side in all_sides:
                            processed_cells[(sheet, border_key)][side].update(candidate_cells)

                # Now check for BORDER_ALL with merged cells
                # When borders are applied to a range and then cells are merged,
                # merged cells only retain their outer perimeter borders
                if merged_ranges and self.config.get("respect_merged_cells", True):
                    # Get all cells that have any border
                    all_bordered_cells = set()
                    for side_cells in sides_data.values():
                        all_bordered_cells.update(side_cells)

                    # Remove already processed cells
                    for side in all_sides:
                        all_bordered_cells -= processed_cells[(sheet, border_key)][side]

                    if all_bordered_cells:
                        # Find the bounding box of remaining bordered cells
                        rows = {c[0] for c in all_bordered_cells}
                        cols = {c[1] for c in all_bordered_cells}
                        candidate_bounds = (min(rows), min(cols), max(rows), max(cols))

                        # Check if this could be a BORDER_ALL with merged cells
                        if check_border_all_with_merges(sides_data, candidate_bounds, merged_ranges):
                            cand_min_row, cand_min_col, cand_max_row, cand_max_col = candidate_bounds

                            # Verify it's more than a single cell
                            if cand_max_row > cand_min_row or cand_max_col > cand_min_col:
                                range_str = f"{get_column_letter(cand_min_col)}{cand_min_row}:{get_column_letter(cand_max_col)}{cand_max_row}"

                                all_op = SetBorder(
                                    cell_range=CellRange(sheet=sheet, range=range_str),
                                    side='all',
                                    value=border_value
                                )
                                consolidated_ops.append(all_op)

                                # Mark all cells in the range as processed
                                for row in range(cand_min_row, cand_max_row + 1):
                                    for col in range(cand_min_col, cand_max_col + 1):
                                        for side in all_sides:
                                            processed_cells[(sheet, border_key)][side].add((row, col))

            # Check for BORDER_OUTSIDE pattern
            if self.config.get("detect_border_outside", True):
                # Find cells that have only outside borders (edges of a range)
                # For a range to have BORDER_OUTSIDE, we need:
                # - Left border on leftmost column
                # - Right border on rightmost column
                # - Top border on topmost row
                # - Bottom border on bottommost row

                left_cells = sides_data.get('left', set()) - processed_cells[(sheet, border_key)]['left']
                right_cells = sides_data.get('right', set()) - processed_cells[(sheet, border_key)]['right']
                top_cells = sides_data.get('top', set()) - processed_cells[(sheet, border_key)]['top']
                bottom_cells = sides_data.get('bottom', set()) - processed_cells[(sheet, border_key)]['bottom']

                # With noise tolerance, allow one side to be empty (3 of 4 present)
                noise_tol = self.config.get("noise_tolerance", True)
                non_empty = sum(bool(s) for s in [left_cells, right_cells, top_cells, bottom_cells])
                min_non_empty = 3 if noise_tol else 4

                if non_empty >= min_non_empty:
                    # Find the bounding box from all available edge cells
                    all_edge_cells = left_cells | right_cells | top_cells | bottom_cells
                    rows = {c[0] for c in all_edge_cells}
                    cols = {c[1] for c in all_edge_cells}

                    min_row, max_row = min(rows), max(rows)
                    min_col, max_col = min(cols), max(cols)

                    # Verify this is an OUTSIDE pattern:
                    # - All left_cells should be in min_col
                    # - All right_cells should be in max_col
                    # - All top_cells should be in min_row
                    # - All bottom_cells should be in max_row
                    is_outside = (
                        all(c[1] == min_col for c in left_cells) and
                        all(c[1] == max_col for c in right_cells) and
                        all(c[0] == min_row for c in top_cells) and
                        all(c[0] == max_row for c in bottom_cells)
                    )

                    if is_outside:
                        # Check if we have complete edges
                        expected_left = {(r, min_col) for r in range(min_row, max_row + 1)}
                        expected_right = {(r, max_col) for r in range(min_row, max_row + 1)}
                        expected_top = {(min_row, c) for c in range(min_col, max_col + 1)}
                        expected_bottom = {(max_row, c) for c in range(min_col, max_col + 1)}

                        noise_tolerance = self.config.get("noise_tolerance", True)
                        if noise_tolerance:
                            # Allow one missing edge (3 out of 4 complete)
                            complete_edges = sum([
                                left_cells >= expected_left,
                                right_cells >= expected_right,
                                top_cells >= expected_top,
                                bottom_cells >= expected_bottom,
                            ])
                            is_valid_outside = complete_edges >= 3
                        else:
                            is_valid_outside = (
                                left_cells >= expected_left and right_cells >= expected_right and
                                top_cells >= expected_top and bottom_cells >= expected_bottom
                            )

                        if is_valid_outside:
                            range_str = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

                            # Create BORDER_OUTSIDE operation
                            outside_op = SetBorder(
                                cell_range=CellRange(sheet=sheet, range=range_str),
                                side='outside',
                                value=border_value
                            )
                            consolidated_ops.append(outside_op)

                            # Mark cells as processed
                            processed_cells[(sheet, border_key)]['left'].update(expected_left)
                            processed_cells[(sheet, border_key)]['right'].update(expected_right)
                            processed_cells[(sheet, border_key)]['top'].update(expected_top)
                            processed_cells[(sheet, border_key)]['bottom'].update(expected_bottom)

        # Add remaining border operations that weren't consolidated
        for (sheet, border_key), sides_data in grouped.items():
            weight, style, color = border_key.split('|')
            border_value = {
                'weight': weight or None,
                'style': style or None,
                'color': color or None
            }

            for side, cells in sides_data.items():
                remaining_cells = cells - processed_cells[(sheet, border_key)][side]
                if remaining_cells:
                    # First try: if remaining cells form a contiguous rectangle,
                    # emit a single rectangular op (e.g. BORDER_TOP | A1:G1758).
                    is_rect, bounds = cells_form_rectangle(remaining_cells)
                    if is_rect:
                        r_min_row, r_min_col, r_max_row, r_max_col = bounds
                        if r_min_row == r_max_row and r_min_col == r_max_col:
                            range_str = f"{get_column_letter(r_min_col)}{r_min_row}"
                        else:
                            range_str = (
                                f"{get_column_letter(r_min_col)}{r_min_row}"
                                f":{get_column_letter(r_max_col)}{r_max_row}"
                            )
                        consolidated_ops.append(SetBorder(
                            cell_range=CellRange(sheet=sheet, range=range_str),
                            side=side,
                            value=border_value,
                        ))
                    # Fallback: merge into contiguous strips.
                    # BORDER_TOP/BOTTOM → horizontal strips (per row).
                    # BORDER_LEFT/RIGHT → vertical strips (per column).
                    elif side in ('top', 'bottom'):
                        strip_ops = _merge_horizontal_strips(
                            remaining_cells, sheet, side, border_value
                        )
                        consolidated_ops.extend(strip_ops)
                    elif side in ('left', 'right'):
                        strip_ops = _merge_vertical_strips(
                            remaining_cells, sheet, side, border_value
                        )
                        consolidated_ops.extend(strip_ops)
                    else:
                        # Fallback for unexpected sides
                        strip_ops = [
                            SetBorder(
                                cell_range=CellRange(
                                    sheet=sheet,
                                    range=f"{get_column_letter(col)}{row}",
                                ),
                                side=side,
                                value=border_value,
                            )
                            for row, col in remaining_cells
                        ]
                        consolidated_ops.extend(strip_ops)

        # Combine non-border ops with passthrough composite ops and consolidated individual ops
        all_border_ops = passthrough_ops + consolidated_ops
        result_ops = non_border_ops + all_border_ops

        self.log(
            context,
            f"Border consolidation: {len(border_ops)} → {len(all_border_ops)} border ops "
            f"({len(passthrough_ops)} composite passthrough, {len(consolidated_ops)} consolidated)"
        )

        return context.copy_with_operations(result_ops)
