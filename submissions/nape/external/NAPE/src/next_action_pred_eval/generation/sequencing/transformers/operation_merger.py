"""
Operation Merger - Merges operations to reduce operation count

Strategies:
- standard: Use existing merge_operations logic (rectangle + border + overlay)
- format_paste_optimization: Detect similar formatting and optimize with PasteFrom operations

Scope:
- global: Merge all operations together
- per_region: Merge operations within each region independently
"""

from typing import List, Dict, Any, Tuple, Optional
from collections import defaultdict, Counter
import json
import re
from next_action_pred_eval.generation.sequencing.base import BaseTransformer, SequencingContext
from next_action_pred_eval.generation.sequencing.utils import is_operation_in_region, get_operation_bounds
from next_action_pred_eval.core.operation import Operation
from next_action_pred_eval.core.cell_range import CellRange
from next_action_pred_eval.core.operations.paste_ops import adjust_formula_references


def create_operation_fingerprint(op: Operation, collision_count: int = 0) -> str:
    """
    Create a unique fingerprint for an operation.

    Fingerprint format: <OpType>:<range>:<value_summary>:<is_inverse>[#collision_count]

    Args:
        op: Operation to fingerprint
        collision_count: Counter for handling identical operations (0 = first occurrence)

    Returns:
        Unique fingerprint string
    """
    from next_action_pred_eval.core.operations import SetInput, PasteFrom, AutoFill, SetFontProperty, SetFillColor, SetBorder, SetAlignment

    op_type = op.__class__.__name__
    range_str = str(op.cell_range)
    inverse_flag = "INV" if op.is_inverse else "REG"

    # Summarize value for fingerprint
    if isinstance(op, SetInput):
        value_summary = f"{op.value}"[:50] if op.value else "None"
    elif isinstance(op, PasteFrom):
        value_summary = f"{op.source_range}→{op.value}"
    elif isinstance(op, (SetFontProperty, SetFillColor, SetBorder, SetAlignment)):
        # Format operations: hash the value dict
        value_summary = str(hash(json.dumps(op.value, sort_keys=True)))[:16]
    else:
        value_summary = str(hash(str(op.value)))[:16]

    base_fingerprint = f"{op_type}:{range_str}:{value_summary}:{inverse_flag}"

    if collision_count > 0:
        return f"{base_fingerprint}#{collision_count}"
    return base_fingerprint


def assign_fingerprints(operations: List[Operation]) -> Dict[Operation, str]:
    """
    Assign unique fingerprints to operations, handling collisions.

    Args:
        operations: List of operations

    Returns:
        Dict mapping operation → fingerprint
    """
    fingerprint_map = {}
    fingerprint_counts = Counter()

    for op in operations:
        # Generate base fingerprint
        base_fp = create_operation_fingerprint(op, collision_count=0)

        # Handle collisions
        collision_count = fingerprint_counts[base_fp]
        fingerprint_counts[base_fp] += 1

        final_fp = create_operation_fingerprint(op, collision_count=collision_count)
        fingerprint_map[op] = final_fp

    return fingerprint_map


def normalize_operation_to_pivot(op: Operation, region_range: str) -> Optional[Operation]:
    """
    Normalize operation to (1,1) pivot for comparison (using 1-based Excel coordinates).

    Args:
        op: Operation to normalize
        region_range: Region's range string (e.g., "A1:C10")

    Returns:
        New operation with coordinates translated to (1,1) pivot, or None if error
    """
    from openpyxl.utils import range_boundaries, get_column_letter

    try:
        # Get region's top-left corner (1-based indices from openpyxl)
        # range_boundaries returns: (min_col, min_row, max_col, max_row)
        region_min_col, region_min_row, _, _ = range_boundaries(region_range)

        # Get operation's coordinates (1-based indices)
        # get_coordinates returns: (start_row, start_col, end_row, end_col)
        op_min_row, op_min_col, op_max_row, op_max_col = op.cell_range.get_coordinates()

        # Translate to (1,1) pivot
        # Formula: new_coord = original_coord - region_min + 1
        norm_min_row = op_min_row - region_min_row + 1
        norm_min_col = op_min_col - region_min_col + 1
        norm_max_row = op_max_row - region_min_row + 1
        norm_max_col = op_max_col - region_min_col + 1

        # Create normalized range string (all coords should be >= 1 now)
        norm_range = f"{get_column_letter(norm_min_col)}{norm_min_row}:{get_column_letter(norm_max_col)}{norm_max_row}"

        # Create copy with normalized range
        return op.model_copy(update={
            'cell_range': CellRange(sheet=op.cell_range.sheet, range=norm_range)
        })
    except Exception as e:
        return None


def calculate_formatting_similarity(ops1: List[Operation], ops2: List[Operation]) -> float:
    """
    Calculate similarity score between two sets of normalized operations.

    Args:
        ops1: First set of normalized operations
        ops2: Second set of normalized operations

    Returns:
        Similarity score 0.0-1.0 (1.0 = identical)
    """
    from next_action_pred_eval.core.operations import SetFontProperty, SetFillColor, SetAlignment, SetBorder

    # Group operations by type and range
    def group_ops(ops):
        groups = defaultdict(list)
        for op in ops:
            if isinstance(op, (SetFontProperty, SetFillColor, SetAlignment, SetBorder)):
                # For SetFontProperty, include property name in key
                if isinstance(op, SetFontProperty):
                    key = (type(op).__name__, str(op.cell_range), op.property)
                else:
                    key = (type(op).__name__, str(op.cell_range))
                groups[key].append(op)
        return groups

    groups1 = group_ops(ops1)
    groups2 = group_ops(ops2)

    if not groups1 or not groups2:
        return 0.0

    # Calculate overlap
    common_keys = set(groups1.keys()) & set(groups2.keys())
    all_keys = set(groups1.keys()) | set(groups2.keys())

    if not all_keys:
        return 0.0

    # Compare values for common keys
    matching_ops = 0
    total_ops = len(all_keys)

    for key in common_keys:
        # Compare operation values
        ops1_values = {json.dumps(op.value, sort_keys=True) if isinstance(op.value, dict) else str(op.value) for op in groups1[key]}
        ops2_values = {json.dumps(op.value, sort_keys=True) if isinstance(op.value, dict) else str(op.value) for op in groups2[key]}

        if ops1_values == ops2_values:
            matching_ops += 1

    return matching_ops / total_ops if total_ops > 0 else 0.0


def detect_common_headers(
    source_region: Dict,
    target_region: Dict,
    source_ops: List[Operation],
    target_ops: List[Operation]
) -> Tuple[Optional[str], Optional[str]]:
    """
    Auto-detect common header rows and columns between regions.

    Args:
        source_region: Source region metadata
        target_region: Target region metadata
        source_ops: Operations in source region
        target_ops: Operations in target region

    Returns:
        Tuple of (common_header_rows_range, common_header_cols_range) or (None, None)
    """
    from next_action_pred_eval.core.operations import SetInput
    from openpyxl.utils import range_boundaries, get_column_letter

    # Extract SetInput operations
    source_inputs = [op for op in source_ops if isinstance(op, SetInput) and not op.is_inverse]
    target_inputs = [op for op in target_ops if isinstance(op, SetInput) and not op.is_inverse]

    if not source_inputs or not target_inputs:
        return None, None

    # Get region boundaries
    src_range = source_region.get("range", "")
    tgt_range = target_region.get("range", "")

    if not src_range or not tgt_range:
        return None, None

    src_min_col, src_min_row, src_max_col, src_max_row = range_boundaries(src_range)
    tgt_min_col, tgt_min_row, tgt_max_col, tgt_max_row = range_boundaries(tgt_range)

    # Check first row (potential header row)
    common_header_row = None
    src_first_row = [op for op in source_inputs if op.cell_range.get_coordinates()[0] == src_min_row]
    tgt_first_row = [op for op in target_inputs if op.cell_range.get_coordinates()[0] == tgt_min_row]

    if src_first_row and tgt_first_row:
        # Compare values
        src_values = [op.value for op in src_first_row]
        tgt_values = [op.value for op in tgt_first_row]

        if src_values == tgt_values:
            # Build header row range for target
            common_header_row = f"{get_column_letter(tgt_min_col)}{tgt_min_row}:{get_column_letter(tgt_max_col)}{tgt_min_row}"

    # Check first column (potential header column)
    common_header_col = None
    src_first_col = [op for op in source_inputs if op.cell_range.get_coordinates()[1] == src_min_col]
    tgt_first_col = [op for op in target_inputs if op.cell_range.get_coordinates()[1] == tgt_min_col]

    if src_first_col and tgt_first_col:
        src_values = [op.value for op in src_first_col]
        tgt_values = [op.value for op in tgt_first_col]

        if src_values == tgt_values:
            common_header_col = f"{get_column_letter(tgt_min_col)}{tgt_min_row}:{get_column_letter(tgt_min_col)}{tgt_max_row}"

    return common_header_row, common_header_col


def is_formula_value(value: Any) -> bool:
    """Check if a value is an Excel formula (starts with '=')."""
    if isinstance(value, str):
        return value.strip().startswith('=')
    return False


def is_sequential_numeric(values: List[Any]) -> bool:
    """
    Check if values form a sequential numeric pattern (e.g., 1, 2, 3, 4, 5).

    Args:
        values: List of values to check (can be 1D or flattened from 2D)

    Returns:
        True if values form a sequential pattern with consistent step
    """
    if not values or len(values) < 3:
        return False

    # Try to convert all values to numbers
    numeric_values = []
    for v in values:
        if v is None or v == '':
            return False
        try:
            numeric_values.append(float(v))
        except (ValueError, TypeError):
            return False

    if len(numeric_values) < 3:
        return False

    # Check for consistent step
    step = numeric_values[1] - numeric_values[0]
    for i in range(2, len(numeric_values)):
        if abs((numeric_values[i] - numeric_values[i-1]) - step) > 0.0001:
            return False

    return True


def flatten_2d_values(value: Any) -> List[Any]:
    """Flatten a 2D array to 1D, or return single value as list."""
    if isinstance(value, list):
        if all(isinstance(row, list) for row in value):
            return [cell for row in value for cell in row]
        return value
    return [value]


def select_paste_type(
    target_ops: List[Operation],
    common_headers: Tuple[Optional[str], Optional[str]]
) -> str:
    """
    Select appropriate paste type based on target operations and headers.

    Args:
        target_ops: Operations in target region
        common_headers: Tuple of (header_row, header_col)

    Returns:
        One of: "paste_format", "paste_template", "paste_full"
    """
    from next_action_pred_eval.core.operations import SetInput

    # Check if target has data operations
    has_data = any(isinstance(op, SetInput) and not op.is_inverse for op in target_ops)

    if not has_data:
        # No data operations → only formatting
        return "paste_format"

    # Has data operations
    header_row, header_col = common_headers
    if header_row or header_col:
        # Has common headers → use template
        return "paste_template"
    else:
        # No common headers → full paste with inverse
        return "paste_full"


class OperationMerger(BaseTransformer):
    """
    Merges operations to reduce operation count.

    Config:
        scope: "global" | "per_region"
        strategies: List of strategies to apply (in order)

        # Standard merge params:
        merge_params:
            row_first: bool
            merge_borders: bool
            merge_inputs: bool
            sort_input_by_type: bool

        # Paste detection params (TODO):
        paste_detection_params:
            min_cells_threshold: int
            create_paste_operation: bool
    """

    DEFAULT_CONFIG = {
        "enabled": True,
        "scope": "global",  # global | per_region
        "strategies": ["standard"],  # standard | paste_detection

        "merge_params": {
            "row_first": True,
            "merge_borders": True,
            "merge_inputs": False,
            "merge_inputs_for_data_tables_only": False,  # If True, merge_inputs only applies to data_table regions
            "sort_input_by_type": True,
            "force_merge_pasted_ranges": False,  # If True, force merge SetValue/SetInput in pasted_ranges into single SetInput

            # Smart INPUT merge: only merge bulk data, keep formulas/sequences separate
            "smart_merge_inputs": False,  # If True, use intelligent merging (skip formulas, detect sequences)
            "smart_merge_inputs_threshold": 32,  # Only merge if cell count >= threshold
        },

        "paste_detection_params": {
            "min_similarity": 0.95,  # Minimum similarity score (0.0-1.0)
            "min_format_operations": 3,  # Minimum number of format operations
        }
    }

    def transform(self, context: SequencingContext) -> SequencingContext:
        if self.can_skip(context):
            return context

        from next_action_pred_eval.core.operations import SetInput
        scope = self.config.get("scope", "global")
        strategies = self.config.get("strategies", ["standard"])

        if scope == "global":
            merged_ops = self._merge_operations(context.operations, strategies, context)
            self.log(context, f"Merged {len(context.operations)} → {len(merged_ops)} operations globally")

            # Log paste constraint metadata
            if context.paste_full_target_regions:
                self.log(context, f"Stored {len(context.paste_full_target_regions)} paste_full target regions for ConstraintEnforcer")
            if context.paste_template_target_regions:
                self.log(context, f"Stored {len(context.paste_template_target_regions)} paste_template target regions for ConstraintEnforcer")

            return context.copy_with_operations(merged_ops)

        elif scope == "per_region":
            if not context.regions:
                # No regions, fall back to global
                merged_ops = self._merge_operations(context.operations, strategies, context)
                self.log(context, f"No regions, merged globally: {len(context.operations)} → {len(merged_ops)}")
                return context.copy_with_operations(merged_ops)

            return self._merge_per_region(context, strategies)

        else:
            self.log(context, f"Unknown scope '{scope}', skipping")
            return context

    def _merge_per_region(
        self,
        context: SequencingContext,
        strategies: List[str]
    ) -> SequencingContext:
        """Merge operations within each region independently"""

        # Group operations by region
        ops_by_region = defaultdict(list)

        for op in context.operations:
            region_id = self._find_region_for_operation(op, context)
            ops_by_region[region_id].append(op)

        # Build region_id -> region dict for quick lookup
        region_by_id = {r.get("id"): r for r in context.regions}

        # Merge each region
        merged_ops = []
        total_before = 0
        total_after = 0

        # Preserve region order
        seen_regions = set()
        for op in context.operations:
            region_id = self._find_region_for_operation(op, context)

            if region_id not in seen_regions:
                seen_regions.add(region_id)

                if region_id in ops_by_region:
                    region_ops = ops_by_region[region_id]
                    total_before += len(region_ops)

                    # Get region metadata for type-aware merging
                    region = region_by_id.get(region_id)
                    merged = self._merge_operations(region_ops, strategies, context, region=region)
                    total_after += len(merged)
                    merged_ops.extend(merged)

        self.log(context, f"Merged per-region: {total_before} → {total_after} ops across {len(context.regions)} regions")
        return context.copy_with_operations(merged_ops)

    def _merge_operations(
        self,
        operations: List[Operation],
        strategies: List[str],
        context: SequencingContext,
        region: Optional[Dict] = None
    ) -> List[Operation]:
        """Apply merge strategies to operations

        Args:
            operations: Operations to merge
            strategies: List of strategy names to apply
            context: Sequencing context
            region: Optional region dict with 'type' field for type-aware merging
        """

        if not operations:
            return []

        current_ops = operations

        for strategy in strategies:
            if strategy == "standard":
                current_ops = self._apply_standard_merge(current_ops, context=context, region=region)

            elif strategy == "paste_detection":
                current_ops = self._apply_paste_detection(current_ops, context)

            # More strategies can be added here

        return current_ops

    def _apply_standard_merge(
        self,
        operations: List[Operation],
        context: Optional[SequencingContext] = None,
        region: Optional[Dict] = None
    ) -> List[Operation]:
        """
        Apply standard merge logic (rectangle + border + overlay).
        Uses existing ExcelConverter merge_operations logic.

        Args:
            operations: Operations to merge
            context: Optional sequencing context for accessing pasted_ranges
            region: Optional region dict. If provided and merge_inputs_for_data_tables_only=True,
                    merge_inputs will be True only for regions with type="data_table"
        """
        from next_action_pred_eval.core.operation_merger import RectangleMerger, BorderMerger
        from next_action_pred_eval.core.operations import SetBorder, SetInput, SetFormula, PasteFrom, AutoFill, OPERATION_ORDER_DICT
        from openpyxl.utils import range_boundaries, get_column_letter

        merge_params = self.config.get("merge_params", {})

        # Determine merge_inputs based on region type if configured
        merge_inputs = merge_params.get("merge_inputs", False)
        if merge_params.get("merge_inputs_for_data_tables_only", False):
            if region is not None:
                # Per-region path: check this specific region
                merge_inputs = (region.get("type") == "data_table")
            elif context and context.regions:
                # Global path: merge inputs if all regions are data_table
                merge_inputs = all(
                    r.get("type") == "data_table" for r in context.regions
                )
            else:
                # No region info available — ignore the data-tables-only
                # restriction and fall back to the base merge_inputs setting
                pass

        # --- Smart INPUT Merge ---
        # If enabled, use intelligent merging: skip formulas, detect sequences
        smart_merge = merge_params.get("smart_merge_inputs", False)
        smart_threshold = merge_params.get("smart_merge_inputs_threshold", 32)

        # Extract operations that should be preserved (INVERSE, PasteFrom)
        inverse_ops = [op for op in operations if isinstance(op, SetInput) and op.is_inverse]
        paste_ops = [op for op in operations if isinstance(op, (PasteFrom, AutoFill))]
        preserved_ops = inverse_ops + paste_ops

        # Only merge the regular operations
        regular_ops = [op for op in operations if op not in preserved_ops]

        # Separate border and non-border operations
        border_ops = [op for op in regular_ops if isinstance(op, SetBorder)]
        non_border_ops = [op for op in regular_ops if not isinstance(op, SetBorder)]

        # Smart merge: separate operations into mergeable and non-mergeable
        if smart_merge and merge_inputs:
            mergeable_ops, non_mergeable_ops = self._partition_smart_merge(
                non_border_ops, smart_threshold
            )
            if context:
                self.log(context, f"Smart merge: {len(mergeable_ops)} mergeable, {len(non_mergeable_ops)} kept separate")
        else:
            mergeable_ops = non_border_ops
            non_mergeable_ops = []

        # Merge non-border operations using rectangle merger
        rectangle_merger = RectangleMerger()
        merged = rectangle_merger.merge(
            mergeable_ops,
            row_first=merge_params.get("row_first", True),
            merge_inputs=merge_inputs,
            sort_input_by_type=merge_params.get("sort_input_by_type", True)
        )

        # Post-merge threshold: split back merged SetInput rectangles below threshold
        # This runs after RectangleMerger so adjacent single cells are grouped first
        if smart_merge and merge_inputs and smart_threshold > 1:
            kept = []
            for op in merged:
                if isinstance(op, SetInput) and not op.is_inverse:
                    try:
                        min_col, min_row, max_col, max_row = range_boundaries(op.cell_range.range)
                        cell_count = (max_col - min_col + 1) * (max_row - min_row + 1)
                        if cell_count < smart_threshold and cell_count > 1:
                            # Below threshold — split back to individual ops
                            non_mergeable_ops.append(op)
                            continue
                    except Exception:
                        pass
                kept.append(op)
            merged = kept

        # Add non-mergeable operations back (formulas, sequences, etc.)
        merged.extend(non_mergeable_ops)

        # Merge border operations if enabled
        if merge_params.get("merge_borders", True) and border_ops:
            border_merger = BorderMerger()
            merged.extend(border_merger.merge(border_ops))

        # Add back preserved operations
        merged.extend(preserved_ops)

        # Apply force_merge_pasted_ranges if enabled and context available
        if merge_params.get("force_merge_pasted_ranges", False) and context:
            merged = self._apply_force_merge_pasted_ranges(merged, context)

        # Sort (spatial ordering)
        merged = sorted(merged, key=lambda op: self._spatial_sort_key(op))

        return merged

    def _partition_smart_merge(
        self,
        operations: List[Operation],
        threshold: int = 32
    ) -> Tuple[List[Operation], List[Operation]]:
        """
        Partition operations into mergeable and non-mergeable based on smart criteria.

        Non-mergeable (kept separate):
        - SetFormula operations (formulas should not be naively merged)
        - SetInput/SetValue with formula values (strings starting with '=')
        - Sequential numeric data (e.g., 1, 2, 3, 4, 5)
        - Small data sets below threshold

        Mergeable (can be bulk merged):
        - Large data sets (>= threshold cells) with non-sequential, non-formula values

        Args:
            operations: List of operations to partition
            threshold: Minimum cell count to consider for bulk merging

        Returns:
            Tuple of (mergeable_ops, non_mergeable_ops)
        """
        from next_action_pred_eval.core.operations import SetInput, SetValue, SetFormula
        from openpyxl.utils import range_boundaries

        mergeable = []
        non_mergeable = []

        for op in operations:
            # SetFormula always kept separate
            if isinstance(op, SetFormula):
                non_mergeable.append(op)
                continue

            # Only process SetInput/SetValue
            if not isinstance(op, (SetInput, SetValue)):
                mergeable.append(op)  # Format ops can be merged
                continue

            # Check for formula values
            if isinstance(op, SetInput):
                values = flatten_2d_values(op.value)
                if any(is_formula_value(v) for v in values):
                    non_mergeable.append(op)
                    continue

                # Check for sequential numeric pattern
                if is_sequential_numeric(values):
                    non_mergeable.append(op)
                    continue

            elif isinstance(op, SetValue):
                # Single cell SetValue - check for formula
                if is_formula_value(op.value):
                    non_mergeable.append(op)
                    continue

            # Passed all checks - can be merged
            mergeable.append(op)

        return mergeable, non_mergeable

    def _apply_force_merge_pasted_ranges(
        self,
        operations: List[Operation],
        context: SequencingContext
    ) -> List[Operation]:
        """
        Force merge SetValue/SetInput operations within pasted_ranges into single SetInput operations.

        For each pasted_range in context.pasted_ranges, collects all SetValue/SetInput operations
        within that range and merges them into a single SetInput with a 2D value array.
        Empty cells within the range are filled with empty strings.

        Args:
            operations: List of operations to process
            context: Sequencing context with pasted_ranges metadata

        Returns:
            List of operations with pasted ranges merged
        """
        from next_action_pred_eval.core.operations import SetInput, SetValue, SetFormula
        from next_action_pred_eval.utils.cell_utils import expand_range
        from openpyxl.utils import range_boundaries, get_column_letter

        pasted_ranges = context.pasted_ranges
        if not pasted_ranges:
            return operations

        # Track which operations have been merged
        ops_to_remove = set()
        new_ops = []

        for pasted_range in pasted_ranges:
            range_str = pasted_range.get("range") if isinstance(pasted_range, dict) else pasted_range
            if not range_str:
                continue

            try:
                pasted_min_col, pasted_min_row, pasted_max_col, pasted_max_row = range_boundaries(range_str)
            except:
                continue

            # Find operations within this pasted range (only SetValue and SetInput, not SetFormula)
            ops_in_range = []
            for op in operations:
                if op in ops_to_remove:
                    continue
                if not isinstance(op, (SetValue, SetInput)):
                    continue
                # Skip formulas (SetFormula or SetInput with formula-like values)
                if isinstance(op, SetFormula):
                    continue

                try:
                    op_min_col, op_min_row, op_max_col, op_max_row = range_boundaries(op.cell_range.range)
                    # Check if operation is fully within pasted range
                    if (op_min_col >= pasted_min_col and op_max_col <= pasted_max_col and
                        op_min_row >= pasted_min_row and op_max_row <= pasted_max_row):
                        ops_in_range.append(op)
                except:
                    continue

            if not ops_in_range:
                continue

            # Find the minimal bounding box that contains all actual data
            # Start with the pasted_range bounds as the initial (maximum) assumption
            actual_min_row, actual_min_col = pasted_max_row, pasted_max_col
            actual_max_row, actual_max_col = pasted_min_row, pasted_min_col

            for op in ops_in_range:
                try:
                    op_min_col, op_min_row, op_max_col, op_max_row = range_boundaries(op.cell_range.range)
                    actual_min_col = min(actual_min_col, op_min_col)
                    actual_min_row = min(actual_min_row, op_min_row)
                    actual_max_col = max(actual_max_col, op_max_col)
                    actual_max_row = max(actual_max_row, op_max_row)
                except:
                    continue

            # Use the trimmed bounds instead of the full pasted_range
            min_col, min_row = actual_min_col, actual_min_row
            max_col, max_row = actual_max_col, actual_max_row

            # Build 2D value array for the trimmed range
            num_rows = max_row - min_row + 1
            num_cols = max_col - min_col + 1
            value_array = [["" for _ in range(num_cols)] for _ in range(num_rows)]

            # Get sheet name from first operation
            sheet_name = ops_in_range[0].cell_range.sheet

            # Fill in values from operations
            for op in ops_in_range:
                try:
                    op_min_col, op_min_row, op_max_col, op_max_row = range_boundaries(op.cell_range.range)

                    # Get the value(s) from the operation
                    if isinstance(op, SetValue):
                        op_value = op.value
                    else:  # SetInput
                        op_value = op.value

                    # Handle single cell vs range
                    if op_min_row == op_max_row and op_min_col == op_max_col:
                        # Single cell
                        row_idx = op_min_row - min_row
                        col_idx = op_min_col - min_col
                        value_array[row_idx][col_idx] = op_value if op_value is not None else ""
                    else:
                        # Range - op_value should be 2D array for SetInput
                        if isinstance(op_value, list):
                            for r_offset, row_vals in enumerate(op_value):
                                if isinstance(row_vals, list):
                                    for c_offset, val in enumerate(row_vals):
                                        row_idx = (op_min_row - min_row) + r_offset
                                        col_idx = (op_min_col - min_col) + c_offset
                                        if 0 <= row_idx < num_rows and 0 <= col_idx < num_cols:
                                            value_array[row_idx][col_idx] = val if val is not None else ""
                                else:
                                    # 1D array treated as single row
                                    row_idx = op_min_row - min_row
                                    col_idx = (op_min_col - min_col) + r_offset
                                    if 0 <= row_idx < num_rows and 0 <= col_idx < num_cols:
                                        value_array[row_idx][col_idx] = row_vals if row_vals is not None else ""
                        else:
                            # Single value for entire range
                            for r in range(op_min_row, op_max_row + 1):
                                for c in range(op_min_col, op_max_col + 1):
                                    row_idx = r - min_row
                                    col_idx = c - min_col
                                    if 0 <= row_idx < num_rows and 0 <= col_idx < num_cols:
                                        value_array[row_idx][col_idx] = op_value if op_value is not None else ""

                    ops_to_remove.add(op)
                except:
                    continue

            # Create merged SetInput operation with the trimmed range
            if ops_to_remove:
                # Build the trimmed range string from actual bounds
                trimmed_range_str = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
                merged_range = CellRange(sheet=sheet_name, range=trimmed_range_str)
                merged_op = SetInput(cell_range=merged_range, value=value_array)
                new_ops.append(merged_op)

        # Build final operations list
        result = [op for op in operations if op not in ops_to_remove]
        result.extend(new_ops)

        return result

    def _apply_paste_detection(
        self,
        operations: List[Operation],
        context: SequencingContext
    ) -> List[Operation]:
        """
        Detect and optimize similar formatting patterns with PasteFrom operations.

        Uses region metadata's similarly_formatted_regions to identify candidates,
        then analyzes similarity and generates appropriate paste operations with
        constraints for proper ordering.

        Args:
            operations: List of operations to optimize
            context: Sequencing context with region metadata

        Returns:
            Optimized list of operations with PasteFrom operations
        """
        from next_action_pred_eval.core.operations import SetInput, PasteFrom, SetFontProperty, SetFillColor, SetAlignment, SetBorder, SetNumberFormat, MergeCells

        # Helper: Check if operation is a formatting operation
        def is_formatting_op(op):
            return isinstance(op, (SetFontProperty, SetFillColor, SetAlignment, SetBorder, SetNumberFormat, MergeCells))

        # Helper: Assign fingerprint with collision handling
        def assign_fingerprint_safe(op, fp_map):
            fp = create_operation_fingerprint(op, 0)
            while fp in fp_map.values():
                collision_count = sum(1 for existing_fp in fp_map.values() if existing_fp.startswith(fp.split('#')[0]))
                fp = create_operation_fingerprint(op, collision_count)
            fp_map[op] = fp
            return fp

        # Helper: Find region by id or range
        def find_region(region_id):
            return next((r for r in context.regions if r.get("id") == region_id or r.get("range") == region_id), None)

        # 1. Assign fingerprints to all operations
        fingerprint_map = assign_fingerprints(operations)
        context.fingerprint_map.update(fingerprint_map)

        # 2. Get similar regions from metadata
        similar_regions = context.region_metadata.get("similarly_formatted_regions", [])

        if not similar_regions:
            self.log(context, "No similarly_formatted_regions found in metadata")
            return operations

        self.log(context, f"Found {len(similar_regions)} similar region groups in metadata")

        # Track operations to remove and add
        ops_to_remove = set()
        ops_to_add = []

        # 3. Process each group of similar regions
        for group in similar_regions:
            similar_region_ids = group.get("similar_regions", [])
            paste_type_hint = group.get("format_paste_type")  # Optional hint from region analysis

            # Check for force_paste_type from config (overrides hint)
            force_paste_type = self.config.get("paste_detection_params", {}).get("force_paste_type")
            if force_paste_type:
                paste_type_hint = force_paste_type
                self.log(context, f"Processing similarity group with {len(similar_region_ids)} regions, FORCED paste_type={paste_type_hint}")
            else:
                self.log(context, f"Processing similarity group with {len(similar_region_ids)} regions, paste_type_hint={paste_type_hint}")

            # Get execution and ordering parameters
            paste_execution = self.config.get("paste_detection_params", {}).get("paste_execution", "block")
            paste_full_ordering = self.config.get("paste_detection_params", {}).get("paste_full_ordering", "grouped")

            # Need at least 2 regions to apply paste optimization
            if len(similar_region_ids) < 2:
                continue

            # Track this group's source and targets for constraint ordering
            group_source_range = None
            group_target_ranges = []
            group_paste_type = None  # Track paste type for this group

            # For paste_full mode, collect all paste/inverse operations and their constraints
            group_paste_full_ops = []  # List of dicts with paste_op, inverse_op, and constraint info

            # For paste_template mode, collect all paste operations and their constraints
            group_paste_template_ops = []  # List of dicts with paste_op, paste_fp, and constraint info

            # For paste_format mode, collect all paste operations and their constraints
            group_paste_format_ops = []  # List of dicts with paste_op, paste_fp, and constraint info

            # Use first region as source template, apply to all others
            source_region_id = similar_region_ids[0]
            source_region = find_region(source_region_id)

            if not source_region:
                self.log(context, f"Source region not found: {source_region_id}")
                continue

            # Set group source range for constraint tracking
            group_source_range = source_region.get("range", "")

            # Process each target region
            for target_region_id in similar_region_ids[1:]:
                target_region = find_region(target_region_id)

                if not target_region:
                    self.log(context, f"Target region not found: {target_region_id}")
                    continue

                # Get operations for each region
                source_ops = self._get_region_operations(operations, source_region)
                target_ops = self._get_region_operations(operations, target_region)

                if not source_ops or not target_ops:
                    continue

                # 4. Normalize operations and calculate similarity (or use hint)
                source_range = source_region.get("range", "")
                target_range = target_region.get("range", "")

                # Get sheet from actual operations (more reliable than region metadata)
                source_sheet = source_ops[0].cell_range.sheet if source_ops else source_region.get('sheet', 'Sheet1')
                target_sheet = target_ops[0].cell_range.sheet if target_ops else target_region.get('sheet', 'Sheet1')

                # If paste type is provided as hint, trust it and skip similarity calculation
                if paste_type_hint:
                    paste_type = paste_type_hint
                    similarity = 1.0  # Assume high similarity if explicitly marked
                    self.log(context, f"Using paste type hint: {paste_type_hint} for {source_range} → {target_range}")
                else:
                    # Calculate similarity
                    norm_source_ops = [normalize_operation_to_pivot(op, source_range) for op in source_ops]
                    norm_target_ops = [normalize_operation_to_pivot(op, target_range) for op in target_ops]

                    # Filter out None values
                    norm_source_ops = [op for op in norm_source_ops if op is not None]
                    norm_target_ops = [op for op in norm_target_ops if op is not None]

                    similarity = calculate_formatting_similarity(norm_source_ops, norm_target_ops)

                    # 5. Check if similarity meets threshold
                    min_similarity = self.config.get("paste_detection_params", {}).get("min_similarity", 0.95)
                    min_format_ops = self.config.get("paste_detection_params", {}).get("min_format_operations", 3)

                    format_ops_count = sum(1 for op in target_ops if is_formatting_op(op))

                    if similarity < min_similarity or format_ops_count < min_format_ops:
                        continue

                    # 6. Detect common headers and select paste type
                    common_headers = detect_common_headers(source_region, target_region, source_ops, target_ops)
                    paste_type = select_paste_type(target_ops, common_headers)

                self.log(context, f"Found similar regions: {source_range} → {target_range} (similarity: {similarity:.2f})")
                self.log(context, f"Paste type: {paste_type}, common headers: {common_headers if not paste_type_hint else 'N/A (using hint)'}")

                # Track paste type for this group
                if group_paste_type is None:
                    group_paste_type = paste_type

                try:
                    # 7. Create PasteFrom operation
                    paste_value = {
                        "paste_format": "formats",
                        "paste_template": "all",
                        "paste_full": "all"
                    }.get(paste_type, "formats")

                    paste_op = PasteFrom(
                        cell_range=CellRange.from_string(f"{target_sheet}!{target_range}"),
                        source_range=f"{source_sheet}!{source_range}",
                        value=paste_value,
                        is_inverse=False
                    )

                    # Assign fingerprint with collision handling
                    paste_fp = assign_fingerprint_safe(paste_op, context.fingerprint_map)

                    # Defer adding to ops_to_add - we'll add in correct order later based on mode
                except Exception as e:
                    import traceback
                    traceback.print_exc()
                    raise

                # 8. Handle different paste types
                if paste_type == "paste_format":
                    # Remove only formatting operations from target
                    ops_to_remove.update(op for op in target_ops if is_formatting_op(op))

                    # Collect constraint information for this paste operation
                    constraint_info = {
                        'paste_op': paste_op,
                        'paste_fp': paste_fp,
                        'source_ops': [(op, fingerprint_map[op]) for op in source_ops if is_formatting_op(op) and op in fingerprint_map],
                        'target_ops': []  # paste_format doesn't have target operations to sequence after
                    }

                    group_paste_format_ops.append(constraint_info)

                elif paste_type == "paste_template":
                    # Detect common headers at CELL level (not entire row/column)
                    from openpyxl.utils import range_boundaries, get_column_letter

                    # Track regions for constraint purposes
                    context.paste_template_target_regions.append(target_region['range'])
                    context.paste_template_source_regions.append(source_region['range'])
                    group_target_ranges.append(target_region['range'])

                    src_min_col, src_min_row, src_max_col, src_max_row = range_boundaries(source_range)
                    tgt_min_col, tgt_min_row, tgt_max_col, tgt_max_row = range_boundaries(target_range)

                    # Detect common header CELLS (not entire rows/columns)
                    # Compare individual INPUT operations in first row and first column
                    source_header_cells = set()  # Cell ranges that are headers in source
                    target_header_cells = set()  # Cell ranges that are headers in target

                    # Get INPUT operations in first row and first column
                    src_first_row_ops = [op for op in source_ops if isinstance(op, SetInput) and not op.is_inverse
                                         and op.cell_range.get_coordinates()[0] == src_min_row]
                    tgt_first_row_ops = [op for op in target_ops if isinstance(op, SetInput) and not op.is_inverse
                                         and op.cell_range.get_coordinates()[0] == tgt_min_row]

                    src_first_col_ops = [op for op in source_ops if isinstance(op, SetInput) and not op.is_inverse
                                         and op.cell_range.get_coordinates()[1] == src_min_col]
                    tgt_first_col_ops = [op for op in target_ops if isinstance(op, SetInput) and not op.is_inverse
                                         and op.cell_range.get_coordinates()[1] == tgt_min_col]

                    # Compare values cell-by-cell in first row
                    for src_op in src_first_row_ops:
                        for tgt_op in tgt_first_row_ops:
                            # Check if they're at the same relative position
                            src_coords = src_op.cell_range.get_coordinates()
                            tgt_coords = tgt_op.cell_range.get_coordinates()

                            src_rel_col = src_coords[1] - src_min_col
                            tgt_rel_col = tgt_coords[1] - tgt_min_col

                            if src_rel_col == tgt_rel_col and src_op.value == tgt_op.value:
                                # Same position, same value = common header
                                source_header_cells.add(src_op.cell_range.range)
                                target_header_cells.add(tgt_op.cell_range.range)

                    # Compare values cell-by-cell in first column
                    for src_op in src_first_col_ops:
                        for tgt_op in tgt_first_col_ops:
                            src_coords = src_op.cell_range.get_coordinates()
                            tgt_coords = tgt_op.cell_range.get_coordinates()

                            src_rel_row = src_coords[0] - src_min_row
                            tgt_rel_row = tgt_coords[0] - tgt_min_row

                            if src_rel_row == tgt_rel_row and src_op.value == tgt_op.value:
                                source_header_cells.add(src_op.cell_range.range)
                                target_header_cells.add(tgt_op.cell_range.range)

                    # Collect constraint information for this paste operation
                    constraint_info = {
                        'paste_op': paste_op,
                        'paste_fp': paste_fp,
                        'source_ops': [],
                        'target_ops': [],
                        'source_header_cells': source_header_cells,
                        'target_header_cells': target_header_cells
                    }

                    for op in target_ops:
                        # Remove formatting operations
                        if is_formatting_op(op):
                            ops_to_remove.add(op)

                        # Remove header data operations (they'll be pasted from source)
                        # Keep non-header data operations (they differ between targets)
                        elif isinstance(op, SetInput) and not op.is_inverse and op in fingerprint_map:
                            # Check if this specific cell is a common header
                            is_header = op.cell_range.range in target_header_cells

                            if is_header:
                                # Remove header INPUT - will come from paste
                                ops_to_remove.add(op)
                            else:
                                # Keep non-header data, collect for constraints
                                constraint_info['target_ops'].append((op, fingerprint_map[op]))

                    # Collect source operations for constraints:
                    # - Header INPUT operations and formatting → before paste (keep them)
                    # - Non-header INPUT operations → after paste (source should be empty, targets have actual data)
                    for op in source_ops:
                        if op not in fingerprint_map:
                            continue

                        if is_formatting_op(op):
                            # All formatting before paste
                            constraint_info['source_ops'].append(('formatting', op, fingerprint_map[op]))
                        elif isinstance(op, SetInput) and not op.is_inverse:
                            # For paste_template, only keep operations in the first column (row labels) BEFORE paste
                            # All other data (including first row headers) should come AFTER paste
                            op_coords = op.cell_range.get_coordinates()
                            op_min_row = op_coords[0]
                            op_min_col = op_coords[1]
                            op_max_col = op_coords[3] if len(op_coords) > 3 else op_coords[1]

                            # Check if this is the first row (header row)
                            is_first_row = (op_min_row == src_min_row)

                            # Keep only if operation is entirely in first column AND not in first row
                            is_first_col_only = (op_min_col == src_min_col and op_max_col == src_min_col)

                            if is_first_col_only and not is_first_row:
                                # First column data (row labels), not header row - before paste
                                constraint_info['source_ops'].append(('row_labels', op, fingerprint_map[op]))
                            else:
                                # Not in first column only, or is header row - should come AFTER paste
                                # This includes first row headers (A5, B5) and all data cells (B6, C6, E7, etc.)
                                constraint_info['source_ops'].append(('data', op, fingerprint_map[op]))

                    group_paste_template_ops.append(constraint_info)

                elif paste_type == "paste_full":
                    # Detect common headers to exclude from INVERSE
                    common_headers = detect_common_headers(source_region, target_region, source_ops, target_ops)
                    header_row, header_col = common_headers
                    header_ranges = []
                    if header_row:
                        header_ranges.append(header_row)
                    if header_col:
                        header_ranges.append(header_col)

                    # Remove formatting operations from target
                    ops_to_remove.update(op for op in target_ops if is_formatting_op(op))

                    # Track this target region for constraint purposes
                    context.paste_full_target_regions.append(target_region['range'])
                    group_target_ranges.append(target_region['range'])

                    # Calculate INVERSE range (excluding headers)
                    inverse_range = target_range
                    if header_row or header_col:
                        # Need to adjust range to exclude headers
                        from openpyxl.utils import range_boundaries, get_column_letter
                        tgt_min_col, tgt_min_row, tgt_max_col, tgt_max_row = range_boundaries(target_range)

                        # Exclude header row (if present)
                        if header_row:
                            tgt_min_row += 1  # Start from row after header

                        # Exclude header column (if present)
                        if header_col:
                            tgt_min_col += 1  # Start from column after header

                        # Build adjusted range (only if there's data left after excluding headers)
                        if tgt_min_row <= tgt_max_row and tgt_min_col <= tgt_max_col:
                            inverse_range = f"{get_column_letter(tgt_min_col)}{tgt_min_row}:{get_column_letter(tgt_max_col)}{tgt_max_row}"
                        else:
                            # Entire region is headers, no INVERSE needed
                            inverse_range = None

                    # Create inverse operation to clear data after paste (excluding headers)
                    if inverse_range:
                        inverse_op = SetInput(
                            cell_range=CellRange.from_string(f"{target_sheet}!{inverse_range}"),
                            value=None,
                            is_inverse=True
                        )

                        # Assign fingerprint with collision handling
                        inverse_fp = assign_fingerprint_safe(inverse_op, context.fingerprint_map)

                        self.log(context, f"Created INVERSE operation for {inverse_range} (excluding headers)")

                        # Collect constraint information for this paste/inverse pair
                        constraint_info = {
                            'paste_op': paste_op,
                            'paste_fp': paste_fp,
                            'inverse_op': inverse_op,
                            'inverse_fp': inverse_fp,
                            'source_ops': [(op, fingerprint_map[op]) for op in source_ops if op in fingerprint_map],
                            'target_ops': [],
                            'header_ranges': header_ranges
                        }

                        # Process target operations
                        for op in target_ops:
                            if isinstance(op, SetInput) and not op.is_inverse and op in fingerprint_map:
                                is_header = any(self._range_overlaps(op.cell_range.range, hr) for hr in header_ranges) if header_ranges else False

                                if is_header:
                                    # Remove header INPUT - will come from paste
                                    ops_to_remove.add(op)
                                else:
                                    # Keep non-header data, will order after INVERSE
                                    constraint_info['target_ops'].append((op, fingerprint_map[op]))

                        group_paste_full_ops.append(constraint_info)
                    else:
                        # No INVERSE needed, just paste (all headers case)
                        self.log(context, f"No INVERSE needed for {target_range} (all headers)")

                        constraint_info = {
                            'paste_op': paste_op,
                            'paste_fp': paste_fp,
                            'inverse_op': None,
                            'inverse_fp': None,
                            'source_ops': [(op, fingerprint_map[op]) for op in source_ops if op in fingerprint_map],
                            'target_ops': [],
                            'header_ranges': header_ranges
                        }

                        # Remove all target INPUT operations (all are headers)
                        for op in target_ops:
                            if isinstance(op, SetInput) and not op.is_inverse:
                                ops_to_remove.add(op)

                        group_paste_full_ops.append(constraint_info)

            # ============================================================================
            # APPLY PASTE OPERATIONS AND CONSTRAINTS PER GROUP
            # ============================================================================

            # --- Case 1: paste_full ---
            if group_paste_type == "paste_full" and group_paste_full_ops:

                if paste_execution == "block":
                    # BLOCK MODE: Build operations based on ordering
                    block_ops = []
                    if paste_full_ordering == "grouped":
                        block_ops = [info['paste_op'] for info in group_paste_full_ops]
                        block_ops += [info['inverse_op'] for info in group_paste_full_ops if info['inverse_op']]
                    else:  # alternating
                        for info in group_paste_full_ops:
                            block_ops.append(info['paste_op'])
                            if info['inverse_op']:
                                block_ops.append(info['inverse_op'])

                    # Create immutable block
                    block_fps = self._create_immutable_block(
                        f"paste_full_block_{group_source_range}",
                        block_ops,
                        context,
                        ops_to_add
                    )

                    # Add boundary constraints
                    all_source_ops = [(op, fp) for info in group_paste_full_ops for op, fp in info['source_ops']]
                    all_target_ops = [(op, fp) for info in group_paste_full_ops for op, fp in info['target_ops']]
                    self._add_boundary_constraints(
                        block_fps[0] if block_fps else None,
                        block_fps[-1] if block_fps else None,
                        all_source_ops,
                        all_target_ops,
                        context
                    )

                elif paste_execution == "interleaved":
                    # INTERLEAVED MODE: Complete each region before next (no immutable blocks)
                    for i, info in enumerate(group_paste_full_ops):
                        # Add operations
                        ops_to_add.append(info['paste_op'])
                        if info['inverse_op']:
                            ops_to_add.append(info['inverse_op'])

                        # Constraints for this region
                        for src_op, src_fp in info['source_ops']:
                            context.add_constraint(src_fp, info['paste_fp'], "Source before paste")

                        if info['inverse_op']:
                            context.add_constraint(info['paste_fp'], info['inverse_fp'], "Paste before inverse")
                            for tgt_op, tgt_fp in info['target_ops']:
                                context.add_constraint(info['inverse_fp'], tgt_fp, "Inverse before target data")

                            # Sequential: complete region[i] before starting region[i+1]
                            if i < len(group_paste_full_ops) - 1:
                                next_info = group_paste_full_ops[i+1]
                                for tgt_op, tgt_fp in info['target_ops']:
                                    context.add_constraint(tgt_fp, next_info['paste_fp'], "Complete region before next")
                        else:
                            for tgt_op, tgt_fp in info['target_ops']:
                                context.add_constraint(info['paste_fp'], tgt_fp, "Paste before target data")

                            if i < len(group_paste_full_ops) - 1:
                                next_info = group_paste_full_ops[i+1]
                                for tgt_op, tgt_fp in info['target_ops']:
                                    context.add_constraint(tgt_fp, next_info['paste_fp'], "Complete region before next")

            # --- Case 2: paste_template ---
            elif group_paste_type == "paste_template" and group_paste_template_ops:
                # BLOCK MODE ONLY
                block_ops = [info['paste_op'] for info in group_paste_template_ops]

                # Create immutable block
                block_fps = self._create_immutable_block(
                    f"paste_template_block_{group_source_range}",
                    block_ops,
                    context,
                    ops_to_add
                )

                # Add boundary constraints (more complex for paste_template)
                first_fp = block_fps[0] if block_fps else None
                last_fp = block_fps[-1] if block_fps else None

                if first_fp:
                    # Source structure → block
                    for info in group_paste_template_ops:
                        for src_type, src_op, src_fp in info['source_ops']:
                            if src_type in ['formatting', 'row_labels']:
                                context.add_constraint(src_fp, first_fp, "Source structure before block")

                if last_fp:
                    # Block → source data and target data
                    for info in group_paste_template_ops:
                        for src_type, src_op, src_fp in info['source_ops']:
                            if src_type == 'data':
                                context.add_constraint(last_fp, src_fp, "Block before source data")
                        for tgt_op, tgt_fp in info['target_ops']:
                            context.add_constraint(last_fp, tgt_fp, "Block before target")

            # --- Case 3: paste_format ---
            elif group_paste_type == "paste_format" and group_paste_format_ops:

                if paste_execution == "block":
                    # BLOCK MODE: Create immutable block for all PASTE operations
                    block_ops = [info['paste_op'] for info in group_paste_format_ops]

                    # Create immutable block
                    block_fps = self._create_immutable_block(
                        f"paste_format_block_{group_source_range}",
                        block_ops,
                        context,
                        ops_to_add
                    )

                    # Add boundary constraints: source formatting → block
                    first_fp = block_fps[0] if block_fps else None
                    if first_fp:
                        for info in group_paste_format_ops:
                            for src_op, src_fp in info['source_ops']:
                                context.add_constraint(src_fp, first_fp, "Source formatting before block")

                elif paste_execution == "interleaved":
                    # INTERLEAVED MODE: Each paste operation happens independently
                    for info in group_paste_format_ops:
                        # Add operation
                        ops_to_add.append(info['paste_op'])

                        # Constraints: source formatting → paste
                        for src_op, src_fp in info['source_ops']:
                            context.add_constraint(src_fp, info['paste_fp'], "Source formatting before paste")

            # ============================================================================

            # Store this group's info for constraint enforcement
            if group_source_range and group_target_ranges:
                if paste_type_hint == "paste_full":
                    context.paste_full_groups.append({
                        'source_range': group_source_range,
                        'target_ranges': group_target_ranges,
                        'paste_execution': paste_execution,  # Pass execution mode
                        'paste_full_ordering': paste_full_ordering  # Pass ordering mode
                    })
                elif paste_type_hint == "paste_template":
                    context.paste_template_groups.append({
                        'source_range': group_source_range,
                        'target_ranges': group_target_ranges
                    })

        # 9. Apply changes to operations list
        result_ops = [op for op in operations if op not in ops_to_remove]
        result_ops.extend(ops_to_add)

        return result_ops

    def _get_region_operations(self, operations: List[Operation], region: Dict) -> List[Operation]:
        """Get all operations that belong to a region"""
        return [op for op in operations if is_operation_in_region(op, region, mode="contain")]

    def _create_immutable_block(
        self,
        block_id: str,
        operations: List[Operation],
        context: SequencingContext,
        ops_to_add: List[Operation]
    ) -> List[str]:
        """
        Create an immutable block from operations.
        Returns list of fingerprints in the block.
        """
        ops_to_add.extend(operations)
        block_fps = [context.fingerprint_map[op] for op in operations if op in context.fingerprint_map]
        if block_fps:
            context.add_immutable_block(block_id, block_fps)
        return block_fps

    def _add_boundary_constraints(
        self,
        first_fp: str,
        last_fp: str,
        source_ops: List[tuple],  # List of (op, fp) tuples
        target_ops: List[tuple],  # List of (op, fp) tuples
        context: SequencingContext
    ):
        """Add constraints: source → first_fp, last_fp → target"""
        if first_fp:
            for _, src_fp in source_ops:
                context.add_constraint(src_fp, first_fp, "Source before block")

        if last_fp:
            for _, tgt_fp in target_ops:
                context.add_constraint(last_fp, tgt_fp, "Block before target")

    def _range_overlaps(self, range1: str, range2: str) -> bool:
        """Check if two ranges overlap"""
        from openpyxl.utils import range_boundaries

        try:
            min_col1, min_row1, max_col1, max_row1 = range_boundaries(range1)
            min_col2, min_row2, max_col2, max_row2 = range_boundaries(range2)

            return not (max_col1 < min_col2 or min_col1 > max_col2 or
                       max_row1 < min_row2 or min_row1 > max_row2)
        except:
            return False

    def _spatial_sort_key(self, op: Operation):
        """Generate spatial sort key for operation"""
        try:
            min_col, min_row, max_col, max_row = get_operation_bounds(op)
            from next_action_pred_eval.core.operations import OPERATION_ORDER_DICT
            op_order = OPERATION_ORDER_DICT.get(type(op), 999)
            return (op.cell_range.sheet, min_row, min_col, max_row, max_col, op_order)
        except:
            return (op.cell_range.sheet, float('inf'), float('inf'), float('inf'), float('inf'), 999)

    def _find_region_for_operation(self, operation: Operation, context: SequencingContext):
        """Find which region an operation belongs to"""
        for region in context.regions:
            if is_operation_in_region(operation, region, mode="contain"):
                return region["id"]
        return None
