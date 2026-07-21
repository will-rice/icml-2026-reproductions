"""
AutoFill Detector — single-pass O(n) transformer that identifies consecutive
SetValue/SetFormula operations forming AutoFill-compatible patterns and collapses
them into source ops + a single AUTOFILL operation.

Detected patterns (all in one pass):
- Arithmetic number series (constant step from 2+ source values)
- Custom lists: months, days, Q-series (from 1+ source values)
- Text+number suffix (e.g., "Item 1" → "Item 2", from 1+ source values)
- Formula row/column offset (adjust_formula_references match, from 1 source)

Does NOT collapse:
- Constant-value copies (better handled by OperationMerger → INPUT)
- Non-contiguous cells or multi-cell operations
- Patterns requiring 3+ source values to infer (avg-of-diffs)

Placement: after OperationSequencer (so ops are spatially ordered),
before ConstraintEnforcer.
"""

import logging
from collections import defaultdict
from typing import Any, Dict, List, Optional, Set, Tuple

from next_action_pred_eval.generation.sequencing.base import BaseTransformer, SequencingContext

logger = logging.getLogger(__name__)


class AutoFillDetector(BaseTransformer):
    """
    Single-pass O(n) transformer that collapses consecutive SetValue/SetFormula
    ops into source ops + AUTOFILL when they form a recognized pattern.

    Config:
        enabled: bool (default True)
        min_fill: int (default 2) — minimum cells to fill beyond source
        detect_formulas: bool (default True) — collapse formula offset patterns
        detect_values: bool (default True) — collapse value series patterns
    """

    DEFAULT_CONFIG = {
        "enabled": True,
        "min_fill": 2,
        "detect_formulas": True,
        "detect_values": True,
    }

    def transform(self, context: SequencingContext) -> SequencingContext:
        if self.can_skip(context):
            return context

        from next_action_pred_eval.core.operations import SetValue, SetFormula, SetInput

        ops = context.operations
        min_fill = self.config.get("min_fill", 2)
        detect_formulas = self.config.get("detect_formulas", True)
        detect_values = self.config.get("detect_values", True)

        if not detect_formulas and not detect_values:
            return context

        # Phase 1: bucket single-cell ops by (sheet, col) and (sheet, row)
        # Handles SetValue, SetFormula, and single-cell SetInput.
        # SetInput is classified as formula (value starts with '=') or value.
        vertical_buckets: Dict[Tuple, List[Tuple[int, int]]] = defaultdict(list)
        horizontal_buckets: Dict[Tuple, List[Tuple[int, int]]] = defaultdict(list)

        for i, op in enumerate(ops):
            if not isinstance(op, (SetValue, SetFormula, SetInput)) or op.is_inverse:
                continue
            r1, c1, r2, c2 = op.cell_range.get_coordinates()
            if r1 != r2 or c1 != c2:  # skip multi-cell ops
                continue
            # Classify: formula or value
            if isinstance(op, SetFormula):
                bucket_type = "formula"
            elif isinstance(op, SetInput) and isinstance(op.value, str) and op.value.startswith("="):
                bucket_type = "formula"
            elif isinstance(op, (SetValue, SetInput)):
                bucket_type = "value"
            else:
                continue
            # Check config flags
            if bucket_type == "formula" and not detect_formulas:
                continue
            if bucket_type == "value" and not detect_values:
                continue
            sheet = op.cell_range.sheet
            vertical_buckets[(sheet, c1, bucket_type)].append((i, r1))
            horizontal_buckets[(sheet, r1, bucket_type)].append((i, c1))

        # Phase 2: find collapsible groups
        collapse_map: Dict[int, None] = {}  # op indices to remove (fill ops)
        insertions: Dict[int, Any] = {}     # after this index → insert AutoFill

        # Vertical groups first
        for key, entries in vertical_buckets.items():
            bucket_type = key[2]  # "formula" or "value"
            entries.sort(key=lambda x: x[1])
            runs = _find_contiguous_runs(entries)
            for run in runs:
                if len(run) < 1 + min_fill:
                    continue
                group_ops = [ops[idx] for idx, _ in run]
                group_indices = [idx for idx, _ in run]
                result = self._try_collapse(group_ops, 'vertical', bucket_type)
                if result:
                    source_size, autofill_op = result
                    for idx in group_indices[source_size:]:
                        collapse_map[idx] = None
                    insertions[group_indices[source_size - 1]] = autofill_op

        # Horizontal groups (skip ops already collapsed vertically)
        for key, entries in horizontal_buckets.items():
            bucket_type = key[2]  # "formula" or "value"
            entries = [(idx, coord) for idx, coord in entries if idx not in collapse_map]
            if len(entries) < 1 + min_fill:
                continue
            entries.sort(key=lambda x: x[1])
            runs = _find_contiguous_runs(entries)
            for run in runs:
                if len(run) < 1 + min_fill:
                    continue
                group_ops = [ops[idx] for idx, _ in run]
                group_indices = [idx for idx, _ in run]
                result = self._try_collapse(group_ops, 'horizontal', bucket_type)
                if result:
                    source_size, autofill_op = result
                    for idx in group_indices[source_size:]:
                        collapse_map[idx] = None
                    insertions[group_indices[source_size - 1]] = autofill_op

        if not collapse_map:
            return context

        # Phase 3: build output
        result = []
        for i, op in enumerate(ops):
            if i in collapse_map:
                continue
            result.append(op)
            if i in insertions:
                result.append(insertions[i])

        collapses = len(insertions)
        removed = len(collapse_map)
        self.log(context, f"Collapsed {removed} ops into {collapses} AUTOFILL ops")
        return context.copy_with_operations(result)

    def _try_collapse(
        self, group_ops: List, direction: str, bucket_type: str = "value"
    ) -> Optional[Tuple[int, Any]]:
        """Try to collapse a contiguous group into source + AUTOFILL.

        Returns (source_size, AutoFill_op) or None.
        """
        if bucket_type == "formula":
            return self._try_formula_pattern(group_ops, direction)
        return self._try_value_pattern(group_ops, direction)

    def _try_value_pattern(
        self, group_ops: List, direction: str
    ) -> Optional[Tuple[int, Any]]:
        """Check if SetValue ops form an AutoFill value pattern."""
        from next_action_pred_eval.core.operations.autofill_ops import (
            _detect_and_extend_values,
        )

        values = [op.value for op in group_ops]
        min_fill = self.config.get("min_fill", 2)

        # Try source_size=1 (custom lists, text+number)
        if len(values) >= 1 + min_fill:
            # Skip constant copy — OperationMerger handles that better
            if not all(v == values[0] for v in values):
                predicted = _detect_and_extend_values(values[:1], len(values) - 1, 1, 1)
                if _values_match(predicted, values[1:]):
                    return 1, self._make_autofill(group_ops, 1, direction)

        # Try source_size=2 (arithmetic series)
        if len(values) >= 2 + min_fill:
            predicted = _detect_and_extend_values(values[:2], len(values) - 2, 1, 2)
            if _values_match(predicted, values[2:]):
                # Skip if both source values are identical (constant copy)
                if values[0] != values[1]:
                    return 2, self._make_autofill(group_ops, 2, direction)

        return None

    def _try_formula_pattern(
        self, group_ops: List, direction: str
    ) -> Optional[Tuple[int, Any]]:
        """Check if SetFormula ops have formulas differing by consistent offset."""
        from next_action_pred_eval.core.operations.paste_ops import (
            adjust_formula_references,
        )

        formulas = [op.value for op in group_ops]
        base = formulas[0]
        min_fill = self.config.get("min_fill", 2)

        if len(formulas) < 1 + min_fill:
            return None

        for i in range(1, len(formulas)):
            if direction == "vertical":
                expected = adjust_formula_references(base, i, 0, "")
            else:
                expected = adjust_formula_references(base, 0, i, "")
            if expected != formulas[i]:
                return None

        return 1, self._make_autofill(group_ops, 1, direction)

    @staticmethod
    def _make_autofill(group_ops: List, source_size: int, direction: str):
        """Create an AutoFill operation from a group of ops."""
        from openpyxl.utils import get_column_letter
        from next_action_pred_eval.core.cell_range import CellRange
        from next_action_pred_eval.core.operations.autofill_ops import AutoFill

        sheet = group_ops[0].cell_range.sheet
        src_range = _ops_to_range_str(group_ops[:source_size])
        dest_range = _ops_to_range_str(group_ops)

        return AutoFill(
            cell_range=CellRange(sheet=sheet, range=dest_range),
            value=f"{sheet}!{src_range}",
        )


# ============================================================================
# Module-level helpers (no state, pure functions)
# ============================================================================


def _find_contiguous_runs(
    entries: List[Tuple[int, int]],
) -> List[List[Tuple[int, int]]]:
    """Given sorted (op_index, coordinate) pairs, find maximal contiguous runs.

    A run is contiguous when consecutive entries have coordinates differing by 1.
    """
    if not entries:
        return []
    runs: List[List[Tuple[int, int]]] = []
    current: List[Tuple[int, int]] = [entries[0]]
    for j in range(1, len(entries)):
        if entries[j][1] == current[-1][1] + 1:
            current.append(entries[j])
        else:
            runs.append(current)
            current = [entries[j]]
    runs.append(current)
    return runs


def _values_match(predicted: List, actual: List) -> bool:
    """Check if predicted values match actual values (with float tolerance)."""
    if len(predicted) != len(actual):
        return False
    for p, a in zip(predicted, actual):
        if p is None and a is None:
            continue
        if p is None or a is None:
            return False
        if isinstance(p, (int, float)) and isinstance(a, (int, float)):
            if abs(float(p) - float(a)) > 1e-9:
                return False
        elif p != a:
            return False
    return True


def _ops_to_range_str(ops: List) -> str:
    """Build a range string (e.g., 'A1:A5') from a list of single-cell ops."""
    from openpyxl.utils import get_column_letter

    coords = [op.cell_range.get_coordinates() for op in ops]
    min_row = min(c[0] for c in coords)
    min_col = min(c[1] for c in coords)
    max_row = max(c[2] for c in coords)
    max_col = max(c[3] for c in coords)

    start = f"{get_column_letter(min_col)}{min_row}"
    end = f"{get_column_letter(max_col)}{max_row}"
    return start if start == end else f"{start}:{end}"
