"""
Formatting Consolidator - Merges adjacent same-type/same-value formatting ops.

Problem: The parser often produces many small formatting ops on individual rows/cells
that a real user would apply in one selection. For example:

    FONT_NAME | A4:C4 | Times
    FONT_NAME | A6:C6 | Times
    FONT_NAME | A7:C7 | Times
    ... (25 more)

A real user would select A4:C30 and apply "Times" once.

This transformer detects such patterns and merges them into larger bounding-box ops.
It also filters out formatting applied to cells far beyond the populated data area.

Note: This runs AFTER the OperationMerger (which handles rectangle merging of inputs)
and AFTER the BorderConsolidator (which handles border-specific consolidation).
"""

import logging
from typing import Any, Dict, List, Optional, Set, Tuple
from collections import defaultdict

from next_action_pred_eval.generation.sequencing.base import (
    BaseTransformer,
    SequencingContext,
)
from next_action_pred_eval.core.cell_range import CellRange
from next_action_pred_eval.core.operation import Operation

logger = logging.getLogger(__name__)


class FormattingConsolidator(BaseTransformer):
    """
    Merges adjacent formatting operations with the same type and value
    into larger contiguous ranges, and optionally filters phantom formatting.

    Config:
        enabled: bool (default True)
        merge_formatting: bool (default True) — merge adjacent same-type/value format ops
        filter_phantom_formatting: bool (default True) — remove format ops beyond data area
        phantom_margin: int (default 3) — columns/rows past data bounding box to keep
        min_ops_to_merge: int (default 3) — minimum number of same-type/value ops to trigger merge
    """

    DEFAULT_CONFIG = {
        "enabled": True,
        "merge_formatting": True,
        "filter_phantom_formatting": True,
        "phantom_margin": 8,
        "min_ops_to_merge": 8,
    }

    def transform(self, context: SequencingContext) -> SequencingContext:
        if self.can_skip(context):
            return context

        operations = list(context.operations)
        initial_count = len(operations)

        # Step 1: Filter phantom formatting (ops beyond data area)
        if self.config.get("filter_phantom_formatting", True):
            operations = self._filter_phantom_formatting(operations, context)

        # Step 2: Merge adjacent same-type/value formatting ops
        if self.config.get("merge_formatting", True):
            operations = self._merge_formatting_ops(operations, context)

        self.log(
            context,
            f"Consolidated {initial_count} → {len(operations)} operations"
        )
        return context.copy_with_operations(operations)

    def _get_data_bounding_box(
        self, operations: List[Operation]
    ) -> Tuple[int, int, int, int]:
        """
        Compute the bounding box of cells that have actual data (INPUT/VALUE/FORMULA).

        Returns:
            (min_row, min_col, max_row, max_col) in 1-based coordinates
        """
        from next_action_pred_eval.core.operations import SetInput, SetValue, SetFormula

        min_row = float('inf')
        min_col = float('inf')
        max_row = 0
        max_col = 0

        for op in operations:
            if not isinstance(op, (SetInput, SetValue, SetFormula)):
                continue
            try:
                r1, c1, r2, c2 = op.cell_range.get_coordinates()
                min_row = min(min_row, r1)
                min_col = min(min_col, c1)
                max_row = max(max_row, r2)
                max_col = max(max_col, c2)
            except Exception:
                pass

        if max_row == 0:
            return (1, 1, 1, 1)
        return (int(min_row), int(min_col), int(max_row), int(max_col))

    def _filter_phantom_formatting(
        self, operations: List[Operation], context: SequencingContext
    ) -> List[Operation]:
        """
        Remove formatting ops on cells that have no data AND are far beyond
        the populated data area.

        Keeps formatting that is within `phantom_margin` columns/rows of the
        data bounding box.
        """
        from next_action_pred_eval.core.operations import (
            SetFontProperty,
            SetFillColor,
            SetAlignment,
            SetNumberFormat,
            SetWrapText,
            SetTextOrientation,
        )

        FORMAT_TYPES = (
            SetFontProperty, SetFillColor, SetAlignment,
            SetNumberFormat, SetWrapText, SetTextOrientation,
        )

        margin = self.config.get("phantom_margin", 3)
        data_bbox = self._get_data_bounding_box(operations)
        _, _, data_max_row, data_max_col = data_bbox

        # Collect cells that have data
        from next_action_pred_eval.core.operations import SetInput, SetValue, SetFormula
        data_cells: Set[Tuple[int, int]] = set()
        for op in operations:
            if isinstance(op, (SetInput, SetValue, SetFormula)):
                try:
                    r1, c1, r2, c2 = op.cell_range.get_coordinates()
                    for r in range(r1, r2 + 1):
                        for c in range(c1, c2 + 1):
                            data_cells.add((r, c))
                except Exception:
                    pass

        result = []
        filtered_count = 0

        for op in operations:
            if not isinstance(op, FORMAT_TYPES):
                result.append(op)
                continue

            try:
                r1, c1, r2, c2 = op.cell_range.get_coordinates()
            except Exception:
                result.append(op)
                continue

            # Check if the op is entirely beyond the data area + margin
            if (r1 > data_max_row + margin and r2 > data_max_row + margin) or \
               (c1 > data_max_col + margin and c2 > data_max_col + margin):
                # Check if ANY cell in this range has data
                has_data = False
                for r in range(r1, min(r2 + 1, r1 + 5)):  # Sample a few cells
                    for c in range(c1, min(c2 + 1, c1 + 5)):
                        if (r, c) in data_cells:
                            has_data = True
                            break
                    if has_data:
                        break

                if not has_data:
                    filtered_count += 1
                    continue

            result.append(op)

        if filtered_count > 0:
            self.log(
                context,
                f"Filtered {filtered_count} phantom formatting ops beyond data area"
            )

        return result

    def _merge_formatting_ops(
        self, operations: List[Operation], context: SequencingContext
    ) -> List[Operation]:
        """
        Merge adjacent same-type/value formatting ops into bounding-box ops.

        Groups formatting ops by (sheet, op_type, property, value) signature,
        then for each group with >= min_ops_to_merge ops, checks if the cells
        form (or nearly form) a contiguous rectangle and merges them.
        """
        from next_action_pred_eval.core.operations import (
            SetFontProperty,
            SetFillColor,
            SetAlignment,
            SetNumberFormat,
            SetWrapText,
            SetTextOrientation,
        )
        from openpyxl.utils import get_column_letter

        FORMAT_TYPES = (
            SetFontProperty, SetFillColor, SetAlignment,
            SetNumberFormat, SetWrapText, SetTextOrientation,
        )

        min_ops = self.config.get("min_ops_to_merge", 8)

        # Separate formatting ops from non-formatting
        format_ops = []
        non_format_ops = []
        for op in operations:
            if isinstance(op, FORMAT_TYPES) and not op.is_inverse:
                format_ops.append(op)
            else:
                non_format_ops.append(op)

        if len(format_ops) < min_ops:
            return operations

        # Build cell-level format map: cell → {property_key: property_value}
        # Used to detect conflicts when merging would overwrite gap cells
        cell_format_map: Dict[Tuple[int, int], Dict[str, str]] = defaultdict(dict)
        for op in format_ops:
            prop_key = self._format_property_key(op)
            prop_val = self._format_property_value(op)
            try:
                r1, c1, r2, c2 = op.cell_range.get_coordinates()
                for r in range(r1, r2 + 1):
                    for c in range(c1, c2 + 1):
                        cell_format_map[(r, c)][prop_key] = prop_val
            except Exception:
                pass

        # Group by signature: (sheet, op_class, property_if_any, value)
        groups: Dict[str, List[Operation]] = defaultdict(list)
        for op in format_ops:
            sig = self._format_op_signature(op)
            groups[sig].append(op)

        merged_ops = []
        ops_consumed: Set[int] = set()  # track by id

        for sig, group_ops in groups.items():
            if len(group_ops) < min_ops:
                continue

            # Collect all cells covered by these ops
            cells: Set[Tuple[int, int]] = set()
            sheet = None
            for op in group_ops:
                sheet = op.cell_range.sheet
                try:
                    r1, c1, r2, c2 = op.cell_range.get_coordinates()
                    for r in range(r1, r2 + 1):
                        for c in range(c1, c2 + 1):
                            cells.add((r, c))
                except Exception:
                    pass

            if not cells or not sheet:
                continue

            # Check if cells form a (near-)rectangle
            rows = {c[0] for c in cells}
            cols = {c[1] for c in cells}
            min_row, max_row = min(rows), max(rows)
            min_col, max_col = min(cols), max(cols)
            bbox_area = (max_row - min_row + 1) * (max_col - min_col + 1)
            actual_area = len(cells)

            # Merge if cells fill at least 80% of the bounding box
            if actual_area < 0.8 * bbox_area:
                continue

            # Check that gap cells don't already have this property with a
            # DIFFERENT value — we must not overwrite existing formatting
            template_op = group_ops[0]
            prop_key = self._format_property_key(template_op)
            prop_val = self._format_property_value(template_op)
            has_conflict = False
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    if (r, c) not in cells:  # gap cell
                        cell_props = cell_format_map.get((r, c), {})
                        if prop_key in cell_props and cell_props[prop_key] != prop_val:
                            has_conflict = True
                            break
                if has_conflict:
                    break
            if has_conflict:
                continue

            # Create merged op with bounding box range
            range_str = (
                f"{get_column_letter(min_col)}{min_row}:"
                f"{get_column_letter(max_col)}{max_row}"
            )
            merged_range = CellRange(sheet=sheet, range=range_str)

            # Create the merged operation from the first op in the group
            merged_op = template_op.model_copy(
                update={'cell_range': merged_range}
            )
            merged_ops.append(merged_op)

            # Mark all ops in this group as consumed
            for op in group_ops:
                ops_consumed.add(id(op))

        if not ops_consumed:
            return operations

        # Rebuild: non-format ops + unconsumed format ops + merged ops
        result = list(non_format_ops)
        for op in format_ops:
            if id(op) not in ops_consumed:
                result.append(op)
        result.extend(merged_ops)

        merge_count = sum(1 for _ in merged_ops)
        consumed_count = len(ops_consumed)
        self.log(
            context,
            f"Merged {consumed_count} formatting ops into {merge_count} bounding-box ops"
        )

        return result

    @staticmethod
    def _format_op_signature(op: Operation) -> str:
        """
        Create a signature for a formatting op so same-type/value ops group together.
        """
        import json
        from next_action_pred_eval.core.operations import (
            SetFontProperty,
            SetAlignment,
            SetNumberFormat,
        )

        sheet = op.cell_range.sheet
        op_type = type(op).__name__

        if isinstance(op, SetFontProperty):
            value_str = f"{op.property}::{op.value}"
        elif isinstance(op, SetAlignment):
            value_str = f"{op.alignment_type}::{op.value}"
        elif isinstance(op, SetNumberFormat):
            value_str = str(op.value)
        else:
            val = op.value
            if isinstance(val, dict):
                value_str = json.dumps(val, sort_keys=True)
            else:
                value_str = str(val)

        return f"{sheet}::{op_type}::{value_str}"

    @staticmethod
    def _format_property_key(op: Operation) -> str:
        """
        Property key identifying WHICH formatting attribute (without value).

        E.g. for FONT_NAME op → "SetFontProperty::name"
        """
        from next_action_pred_eval.core.operations import (
            SetFontProperty,
            SetAlignment,
        )

        op_type = type(op).__name__
        if isinstance(op, SetFontProperty):
            return f"{op_type}::{op.property}"
        elif isinstance(op, SetAlignment):
            return f"{op_type}::{op.alignment_type}"
        return op_type

    @staticmethod
    def _format_property_value(op: Operation) -> str:
        """The value part of a formatting op, as a comparable string."""
        import json
        from next_action_pred_eval.core.operations import (
            SetFontProperty,
            SetAlignment,
            SetNumberFormat,
        )

        if isinstance(op, SetFontProperty):
            return str(op.value)
        elif isinstance(op, SetAlignment):
            return str(op.value)
        elif isinstance(op, SetNumberFormat):
            return str(op.value)
        val = op.value
        if isinstance(val, dict):
            return json.dumps(val, sort_keys=True)
        return str(val)
